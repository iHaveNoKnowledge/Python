
from loguru import logger
from decimal import Decimal
import locale
from concurrent.futures import ThreadPoolExecutor
import threading
import sys
import os
from xml.dom.minidom import Document
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.abstract_event_listener import AbstractEventListener
from selenium.webdriver.support.events import EventFiringWebDriver, AbstractEventListener
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import UnexpectedAlertPresentException
from selenium import webdriver
from webdriver_auto_update.chrome_app_utils import ChromeAppUtils
from webdriver_auto_update.webdriver_manager import WebDriverManager

import re
import win32com.client as comclt
import time
import pandas as pd
import numpy as np
from tkinter import font
from tkinter import ttk
from tkinter import filedialog
from tkinter import messagebox
from tkinter import *
from customtkinter import *
from pypdf import PdfReader
from openpyxl import load_workbook
from PIL import Image, ImageTk
import base64

import traceback
from bs4 import BeautifulSoup
import httpcore
from googletrans import Translator
import requests
session = requests.Session()

import datetime
import pytz

# * images
icon_path = os.path.join(os.path.dirname(__file__), 'imgs', 'kheedluang.ico')
arrow_icon = os.path.join(os.path.dirname(__file__), 'imgs', 'Arrow.gif')
stop_icon = os.path.join(os.path.dirname(__file__), 'imgs', 'stop.jpg')

# * user interface
# * dataframe table
# from test_auto_cus_name_MKII import *

# * selenium
# from ....python_modules3.SMCO.cusNameFixer import cusNameFixer, currencyRemover, addressExtractor, cusNameFixer2, cusNameFixer3


locale.setlocale(locale.LC_ALL, 'en_us')


current_directory = os.getcwd()
print("current_directory:", current_directory)
address_file = "Addresscleaner_TambonData.xlsx"
file_path = os.path.join(current_directory, address_file)
directory_of_file = os.path.dirname(file_path)
print("file located:", directory_of_file)
# sys.path.append(os.path.dirname(os.getcwd()))

# * ปรับ https ให้ตัว translate
setattr(httpcore, 'SyncHTTPTransport', 'AsyncHTTPProxy')


class MyApp:
    def __init__(self, root):
        self.root = root
        self.dev_account = ["62078", "61651", "62302"]
        self.is_bot_running = BooleanVar(value=False)
        # self.validate_input_variable = self.root.register(self.validate_input)
        self.user_id = StringVar(value="")
        self.user_pw = StringVar(value="")
        self.result = ""
        self.is_accel_mode = BooleanVar()
        self.is_accel_mode_activated = BooleanVar(value=False)
        self.table_location = ""
        self.marketplace_target = StringVar(value="MarketPlace")
        self.bg_by_market_place = {'SHOPEE': '#ee4d2d', 'LAZADA': '#201adb', '': '#747474'}
        self.cus_order = StringVar(value="")
        self.tax_bool = BooleanVar(value=False)
        self.tax_num = StringVar(value="")
        self.is_tax = StringVar(value="")
        self.tax_branch_num = StringVar(value="")
        self.cus_name = StringVar(value="")
        self.cus_account_name = StringVar(value="")
        self.cus_address = ""
        self.cus_remark = ""
        self.order_note = ""
        self.cus_province = StringVar(value="")
        self.cus_district = StringVar(value="")
        self.cus_sub_district = StringVar(value="")
        self.cus_tel = StringVar(value="")
        self.cus_email = StringVar(value="")
        self.cus_cur_status = StringVar(value="")
        self.is_forbid = False
        self.cus_ship_cost = DoubleVar(value=0)
        self.cus_seller_voucher = DoubleVar(value=0)
        self.cus_purchase_time = StringVar(value="")
        self.cus_arrow_btn = '/html/body/div[1]/div[2]/div[2]/div[2]/div[2]/div[1]/div[2]/div/div/div[6]/form/div/span/span[1]/span/span[2]'
        self.cusNameInput = '/html/body/span/span/span[1]/input'
        self.cusSearchSMCO = '/html/body/div[1]/div[2]/div[2]/div[2]/div[2]/div[1]/div[2]/div/div/div[7]/a'
        self.cusCreateBtn = '/html/body/div[1]/div[2]/div[11]/div/div/div[2]/div/form/div[1]/div[2]/button[1]'
        self.cusNameLi1 = '/html/body/span/span/span[2]/ul/li'
        self.cus_name_dropdown_ul = '/html/body/span/span/span[2]/ul'
        # self.bot_state = BooleanVar(value=False)
        self.cookies = {'vatinfo': {
            'JSESSIONID': '',
        }}
        self.is_gui_busy = BooleanVar(value=False)
        self.bot = Bot_POS(self.root, self)
        # self.cus_masked_name = StringVar(value="")
        # self.cus_masked_tel = StringVar(value="")

        self.scale_factor = self.adjust_scale(self.root, 1000, 900)
        self.create_main_window()
        self.scale_widget(self.root, self.scale_factor)
        self.get_dataframe()
        self.mimic_list_item_states = []
        logger.add("autopageMKII_log.log",
                   format="{time} {level} {message}", level="INFO")

    def demonic_cp_selection(self):
        self.bot.demonic_cp_bot(
            self.demonicCp_itemNo.get(), self.demonicCp_cpNo.get())

    def validate_input(self, value):
        pattern = r'[A-z]'
        if re.fullmatch(pattern, value) is None:
            return False

        return True

    def on_canvas_configure(self, event):
        self.canvas_width = event.width
        self.canvas_height = event.height
        self.root_frame.config(width=self.canvas_width,
                               height=self.canvas_height)

    def adjust_scale(self, root, base_width, base_height):
        # Get current screen resolution
        screen_width = root.winfo_screenwidth()
        screen_height = root.winfo_screenheight()

        # Calculate scale factors
        width_scale = screen_width / base_width
        height_scale = screen_height / base_height

        # Use the smaller scale factor to maintain aspect ratio
        scale_factor = min(width_scale, height_scale)

        return scale_factor

    def scale_widget(self, widget, scale_factor):
        if isinstance(widget, (CTkLabel, CTkButton, CTkEntry, CTkFrame)):
            current_width = widget.cget("width")
            current_height = widget.cget("height")
            new_width = int(current_width * scale_factor)
            new_height = int(current_height * scale_factor)
            widget.configure(width=new_width, height=new_height)

        if isinstance(widget, CTk):
            for child in widget.winfo_children():
                self.scale_widget(child, scale_factor)

    def create_main_window(self):
        self.root.geometry("1000x900+400+300")
        self.root.title("Autosamatic ver0.396.1")
        self.root.configure(bg="#444")

        # #* BG CANVAS ##################################################################################
        self.canvas = Canvas(self.root, bg="#444")

        # * Scrollbar For Root ##################################################################################
        self.root_scrollbar_y = Scrollbar(
            self.canvas, command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=self.root_scrollbar_y.set)
        self.root_scrollbar_y.pack(side=RIGHT, fill="y")

        self.root_scrollbar_x = Scrollbar(
            self.root, command=self.canvas.xview, orient='horizontal')
        self.root_scrollbar_x.pack(side=BOTTOM, fill="x")

        self.canvas.config(xscrollcommand=self.root_scrollbar_x.set)
        self.canvas.configure(xscrollcommand=self.root_scrollbar_x.set)
        self.canvas.pack(side="left", fill="both", expand=True)

        # #* FRAMES #####################################################################################################
        # self.root_frame = Frame(self.canvas,  bg="pink") ใช้ได้แต่รอก่อน
        # self.canvas.create_window((0, 0), window=self.root_frame, anchor="nw") ใช้ได้แต่รอก่อน

        # > Frame1 Order Entry
        self.entry_frame = Frame(self.canvas, padx=5, pady=5, bg="#444",
                                 borderwidth=1, relief="groove", highlightbackground="#ccc")
        self.entry_frame.pack(side='top', pady=(10, 10))

        # > Frame2 Log Frame
        self.log_frame = Frame(self.canvas, bg="#444")
        self.log_frame.pack(side='bottom', pady=(0, 20))

        # > Frame3 ImportFile Status and Bot Status
        self.import_file_frame = Frame(self.canvas, bg="#444")
        self.import_file_frame.pack(
            side='top', anchor=W, padx=(5, 5), pady=(5, 0))

        # > Frame4 Customer Details
        self.order_details_frame = Frame(self.canvas, bg="#444", )
        self.order_details_frame.pack(
            side='top', anchor=W, padx=(5, 5), pady=(5, 0))

        # > Frame7 For Customer's Invoice Details
        self.invoice_details_frame = Frame(self.canvas, bg="#445")
        self.invoice_details_frame.pack(
            side='top', anchor=W, padx=(5, 5), pady=(5, 0))

        # > Frame5 Products Lists
        self.products_list_frame = Frame(self.canvas, bg="#445")
        self.products_list_frame.pack(
            side='top', padx=(5, 5), pady=(5, 5), fill=X)

        # > Frame6 Margetplace(MP) Products Lists
        self.mp_products_list_frame = Frame(self.canvas, bg="#444")
        self.mp_products_list_frame.pack(side='top',  padx=(
            5, 5), pady=(5, 5), fill="x")

        # > Frame7 The Upper Log Frame Demonic Frame
        self.demonic_frame = Frame(self.canvas, bg="#444")
        self.demonic_frame.pack(side='bottom', pady=(0, 2))

        # * Create widgets in the main window
        self.create_widgets()

        # * start the scrollbar
        self.canvas.update_idletasks()
        self.canvas.config(scrollregion=self.canvas.bbox("all"))
        self.canvas.bind_all("<MouseWheel>", lambda event: self.canvas.yview_scroll(
            int(-1*(event.delta/120)), "units"))
        # self.canvas.bind("<Configure>", self.on_canvas_configure) ใช้ได้แต่รอก่อน

    def measure_text(self, text):
        return font.Font().measure(str(text).strip())

    def row_header_maker(self, list_of_cols):
        # * สร้าง header
        self.list_of_cols = list_of_cols
        self.colspan_amount = [1, 19, 2, 2, 2, 2]
        self.cols_location = [0, 1, 21, 23, 25, 27]
        self.cols_width = [5, 100, 10, 10, 10, 10]
        # self.cols_width = [1, 22, 2, 2, 2, 2]
        self.entry_list = []
        i = 0
        for header in self.list_of_cols:
            self.mp_products_header = Entry(
                self.mp_products_list_frame, foreground="#000000", background="#fff", width=int(self.cols_width[i]))
            self.mp_products_header.insert(0, header)
            self.entry_list.append(self.mp_products_header)
            i += 1

        for idx, entry in enumerate(self.entry_list):
            entry.grid(
                row=0, column=self.cols_location[idx], columnspan=self.colspan_amount[idx], sticky='nsew')
            entry.configure(state="readonly")

    # * เป็นส่วนของ GUI ช่อง input ที่รับ order sn ในรูปแบบ file excel
    def accelmode_toggle(self):
        # * ระบุสถานะการแสดงผลของ GUI ที่ใช้สำหรับการ Input order แบบธรรมดา เพื่อทำการ remove มันแล้วแทนที่ ด้วย GUI ของ Accel mode
        order_label = self.inp1_label_order.winfo_ismapped()
        order_input = self.inp1_order_input.winfo_ismapped()
        order_btn = self.inp1_search_btn.winfo_ismapped()

        # * ถ้า Accel mode ทำงาน
        if self.is_accel_mode.get():
            # * gui โหมดธรรมดาทุกตัวทำงานอยู่
            if order_label and order_input and order_btn:
                # * ลบ gui โหมดธรรมดา ทิ้งรายตัว
                self.inp1_label_order.grid_remove()
                self.inp1_order_input.grid_remove()
                self.inp1_search_btn.grid_remove()

            # * เอา gui ของ accel mode มาแปะแทน
            self.accl_dir_label.grid(row=0, column=1, padx=5)
            self.accl_dir_namedisplay_on_btn.grid(row=0, column=3)
            self.add_trans_to_accel_file_btn.grid(row=0, column=4)
            self.accl_start_btn.grid(row=0, column=5, padx=5)

        # * ถ้า Accel mode ไม่ทำงาน
        else:
            # * ลบ gui ของ accel mode ทิ้งรายตัว
            self.accl_dir_label.grid_remove()
            self.accl_dir_namedisplay_on_btn.grid_remove()
            self.add_trans_to_accel_file_btn.grid_remove()
            self.accl_start_btn.grid_remove()

            # * เอา gui ของ โหมดธรรมดา มาแปะแทน
            self.inp1_label_order.grid(row=0, column=1, padx=5)
            self.inp1_order_input.grid(row=0, column=3)
            self.inp1_search_btn.grid(row=0, column=5, padx=5)

    def create_widgets(self):
        # * entry_frame !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
        # *  MarketPlace
        # * > Label
        self.marketplace_label = Label(
            self.entry_frame, textvariable=self.marketplace_target, bg="#747474", fg="#FFF", font='bazooka 10 bold')
        self.marketplace_label.grid(row=0, column=0, padx=5)

        # *  search order component !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
        # * > Labels
        self.inp1_label_order = Label(
            self.entry_frame, text="Order: ", bg="#FFF", width=10)
        self.inp1_label_order.grid(row=0, column=1, padx=5)
        # *> Inputs
        self.entered_order = StringVar()
        self.inp1_order_input = Entry(
            self.entry_frame, textvariable=self.entered_order, width=50)
        self.inp1_order_input.grid(row=0, column=3)
        # *> Buttons
        self.font = CTkFont(family='fixedsys', size=10, weight="bold")
        self.inp1_search_btn = CTkButton(
            self.entry_frame,
            font=self.font,
            text="Start",
            fg_color="#f5a91d",
            text_color="#1E1E1E",
            border_width=1.5,
            border_color="#7d4b19",
            command=self.search_order,
            width=10,
            height=25
        )
        self.inp1_search_btn.grid(row=0, column=5, padx=5)

        # *  search order Accel mode component !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
        # * พวกนี้มันต้อง add แบบ toggle เพราะมันต้องสลับกับโหมดปกติ
        # * > Labels
        self.accl_dir_label = Label(
            self.entry_frame, text=f"Accel File Dir ")

        # *> FileName Display on Button
        self.accl_dir_namedisplay_on_btn = Button(
            self.entry_frame, text=f"ยังไม่เลือก Accel File", command=self.select_accel_file, bg="#969696")

        # *> Buttons
        self.start_image = Image.open(arrow_icon)
        self.start_photo = ImageTk.PhotoImage(
            self.start_image.resize((30, 20)))
        # * fonts
        # self.font = CTkFont(family='fixedsys', size=10, weight="bold")
        self.accl_start_btn = CTkButton(
            self.entry_frame,
            font=self.font,
            text=f"Start",
            # image=self.start_photo,
            command=self.accel_search,
            fg_color="#81ed55",
            text_color="#1E1E1E",
            border_color="#2d8a37",
            border_width=1.5,
            width=30,
            height=25
        )

        # *  search order Stop Button
        self.accel_stop_btn = CTkButton(
            self.entry_frame,
            font=self.font,
            text=f"Stop",
            command=self.stop_operation,
            fg_color="#bf2d2a",
            text_color="#ffffff",
            border_width=1.5,
            border_color="#732844",
            width=28,
            height=25
        )
        self.accel_stop_btn.grid(row=0, column=6, padx=5)

        # todo WIP transfer to accel
        # *  add transfers to accel mode component !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
        # *พวกนี้มันต้อง add แบบ toggle เพราะมันต้องสลับกับโหมดปกติ
        # *> add transfer Button
        self.add_trans_to_accel_file_btn = Button(
            self.entry_frame, text=f"เลือกใส่ Transfer", command=lambda: self.extract_sn_btn(self.accel_file_dir), bg="#969696")

        # *  Log in button component !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
        # * > A BTN to display the User_account
        self.btn_display = f"ID:{self.user_id.get()}" if self.user_id.get(
        ) and self.user_pw.get() else "Login"
        self.display_acc_btn = Button(
            self.entry_frame, text=self.btn_display, command=lambda: UserAccount(self.root, self))
        self.display_acc_btn.grid(row=0, column=7, padx=5)

        # * Accel mode
        # * > Checkbox for activation toggle
        self.accel_mode_checkbox = Checkbutton(
            self.entry_frame, text="Accel Mode", variable=self.is_accel_mode, command=self.accelmode_toggle)

        # for method
        # if self.user_id in self.dev_account and self.is_accel_mode.get():
        #     print("Accel mode Activated")
        # else:
        #     print("Normal mode")

        # * import_file_frame !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
        # *  Export File and Bot status location display component
        self.display_location_label = Label(
            self.import_file_frame, text=f"File located: ")
        self.display_location_label.grid(row=0, column=0, padx=(5, 0))

        self.display_location_result = Label(
            self.import_file_frame, text=f"ยังไม่เลือก Import File")
        self.display_location_result.grid(row=0, column=1, padx=(5, 0))

        self.display_location_result_btn = Button(
            self.import_file_frame, text=f"ใส่ Import File", command=self.select_excel, bg="#969696")
        self.display_location_result_btn.grid(row=0, column=2, padx=(5, 0))

        # >> bot status
        self.display_bot_status_label = Label(
            self.import_file_frame, text=f"Bot Status: ไม่มีการทำงาน (⸝⸝ᴗ﹏ᴗ⸝⸝) ᶻ 𝗓 𐰁", bg="#1f242e", fg="#ffec1f")
        self.display_bot_status_label.grid(row=0, column=3, padx=(5, 0))

        # * > Current Order display component
        # >> Labels
        self.label_current_order = Label(
            self.order_details_frame, text="Current Order: ", bg="#FFF",)
        self.label_current_order.grid(row=1, column=0, padx=(5, 0))

        self.display_current_order = Entry(
            self.order_details_frame, width=40, state="readonly",  borderwidth=0, textvariable=self.cus_order)
        self.display_current_order.grid(row=1, column=1, padx=(1, 0), sticky=W)

        # * > Current Status display component
        # >> Labels
        self.label_current_status = Label(
            self.order_details_frame, text="Status: ", bg="#FFF",)
        self.label_current_status.grid(
            row=1, column=2, padx=(5, 0), columnspan=1)
        self.display_current_status = Label(
            self.order_details_frame, width=20,  borderwidth=0, textvariable=self.cus_cur_status, fg="#000000", bg="#8fd4ff")

        self.display_current_status.grid(
            row=1, column=3, padx=(1, 0), sticky=W)

        # * > Is Tax?? display component
        # >> Labels
        self.label_is_tax = Label(
            self.order_details_frame, text="ใบกำกับ", bg="#FFF")
        self.label_is_tax.grid(row=2, column=2, padx=(5, 0), sticky='ew')
        # >> Value display
        self.display_is_tax = Label(
            self.order_details_frame,  borderwidth=0, textvariable=self.is_tax, foreground="#000000", background="#fff")
        self.display_is_tax.grid(row=2, column=3, padx=(1, 0), sticky='ew')

        # * > Tax Number display component
        # >> Labels
        self.label_tax_number = Label(
            self.order_details_frame, text="เลขผู้เสียภาษี", bg="#FFF")
        self.label_tax_number.grid(row=2, column=4, padx=(5, 0), sticky='ew')
        # >> Value display
        self.display_tax_number = Entry(
            self.order_details_frame, width=15,  borderwidth=0, textvariable=self.tax_num, foreground="#000000", background="#fff", readonlybackground="white", state="readonly")
        self.display_tax_number.grid(row=2, column=5, padx=(1, 0), sticky='ew')

        # * > Customer Email display component
        # >> Labels
        self.label_cus_email = Label(
            self.order_details_frame, text="Email", bg="#FFF")
        self.label_cus_email.grid(row=2, column=6, padx=(5, 0), sticky='ew')
        # >> Value display
        self.display_cus_email = Entry(
            self.order_details_frame, width=15,  borderwidth=0, textvariable=self.cus_email, foreground="#000000", background="#fff", readonlybackground="white", state="readonly")
        self.display_cus_email.grid(row=2, column=7, padx=(1, 0), sticky='ew')

        # * > Customer Name display component
        # * >> Labels
        self.label_cus_name = Label(
            self.order_details_frame, text="ชื่อ", bg="#FFF", height=1)
        self.label_cus_name.grid(row=2, column=0, padx=(
            5, 0), pady=(2, 2), sticky='ew')
        # * >> Value display
        self.display_cus_name = Entry(
            self.order_details_frame, width=40,  borderwidth=0, textvariable=self.cus_name, foreground="#000000", background="#fff", state="readonly")
        self.display_cus_name.grid(row=2, column=1, padx=(1, 0), sticky='ew')

        # * > Customer Address display component ส่วนแสดงผลที่อยู่ลูกค้า
        # * >>Address
        # >>> Labels
        self.label_cus_address = Label(
            self.invoice_details_frame, text="ที่อยู่: ", bg="#FFF", height=1,)
        self.label_cus_address.grid(
            row=3, column=0, padx=(5, 0), pady=(2, 2), sticky="nsew")
        # >>> Value display
        self.display_cus_address = Text(
            self.invoice_details_frame, width=44, height=5, borderwidth=0, foreground="#000000", background="#fff", state="disabled")
        self.display_cus_address.grid(
            row=3, column=1, padx=(1, 0), columnspan=2, sticky=W)
        self.display_cus_address.tag_add("left", "1.0", "1.end")

        # * >> Customer remark display component ส่วนแสดงผลหมายเหตุลูกค้า col 4-5
        # >>> Labels
        self.label_cus_remark = Label(
            self.invoice_details_frame, text="หมายเหตุจากผู้ซื้อ: ", bg="#FFF", height=1,)
        self.label_cus_remark.grid(row=3, column=4, padx=(
            5, 0), pady=(2, 2), sticky="nsew")
        # >>> Value display
        self.display_cus_remark = Text(
            self.invoice_details_frame, width=20, height=5, borderwidth=0, foreground="#000000", background="#fff", state="disabled")
        self.display_cus_remark.grid(
            row=3, column=5, padx=(1, 0), columnspan=1, sticky=W)
        self.display_cus_remark.tag_add("left", "1.0", "1.end")

        # * >> Order Note display component ส่วนแสดงผลหมายเหตุลูกค้า col 6-7
        # >>> Labels
        self.label_order_note = Label(
            self.invoice_details_frame, text="บันทึก: ", bg="#FFF", height=1,)
        self.label_order_note.grid(row=3, column=6, padx=(
            5, 0), pady=(2, 2), sticky="nsew")
        # >>> Value display
        self.display_order_note = Text(
            self.invoice_details_frame, width=20, height=5, borderwidth=0, foreground="#000000", background="#fff", state="disabled")
        self.display_order_note.grid(
            row=3, column=7, padx=(1, 0), columnspan=1, sticky=W)
        self.display_order_note.tag_add("left", "1.0", "1.end")

        # * > Customter Products List
        self.label_cus_products = Label(
            self.products_list_frame, text="รายการสินค้า: ", bg="#FFF", height=1)
        self.label_cus_products.pack()

        # * >> สร้าง Treeview widget
        self.tree = ttk.Treeview(self.products_list_frame, columns=(
            "Productname", "Price", "QTY"), show="headings")
        self.tree.column("Productname", anchor=W, width=350)
        self.tree.column("Price", width=self.measure_text("Price")+10)
        self.tree.column("QTY", width=self.measure_text("QTY")+10)
        self.tree.heading("Productname", text="Product")
        self.tree.heading("Price", text="Price")
        self.tree.heading("QTY", text="QTY")

        self.y_scrollbar = ttk.Scrollbar(
            self.products_list_frame, command=self.tree.yview)

        self.y_scrollbar.pack(side="right", fill="y")
        self.tree.pack(side='bottom', fill=X)
        self.tree.config(yscrollcommand=self.y_scrollbar.set)

        # * > Margetplace Products display Header
        headers = ['No.', 'สินค้าทั้งหมด', 'ราคาต่อชิ้น',
                   'จำนวน', 'ราคาขายสุทธิ', 'ราคารวมรีเบท']
        self.row_header_maker(headers)

        # * > demonic cp segment
        # * >> Label
        self.demonicCp_label = Label(
            self.demonic_frame, text="Ulti CP", bg="#FFF", height=1)
        self.demonicCp_label.grid(row=0, column=0)
        # * >> Inputs1
        self.demonicCp_itemNo = StringVar()
        self.demonicCp_itemNo_input = Entry(
            self.demonic_frame, textvariable=self.demonicCp_itemNo, width=10)
        self.demonicCp_itemNo_input.grid(row=0, column=3, padx=(0, 2))
        # * >> Inputs2
        self.demonicCp_cpNo = StringVar()
        self.demonicCp_cpNo_input = Entry(
            self.demonic_frame, textvariable=self.demonicCp_cpNo, width=10)
        self.demonicCp_cpNo_input.grid(row=0, column=4)
        # * >> Buttons
        self.demonicCp_btn = Button(
            self.demonic_frame, text="SonicBlow!!", bg="#969696", command=self.demonic_cp_selection, width=10)
        self.demonicCp_btn.grid(row=0, column=5)

        # * > Log windows component
        self.report_log = Text(self.log_frame, state=DISABLED, height=13)

        self.scrollbar = Scrollbar(
            self.log_frame, command=self.report_log.yview)
        self.scrollbar.pack(side="right", fill="y")

        self.report_log.pack(side='bottom', fill=X)
        self.report_log.config(yscrollcommand=self.scrollbar.set)

        ## * Create DataSourceSelector instance ###########
        self.data_source_selector = DataSourceSelector(self.root, self)
        self.user_account = UserAccount(self.root, self)

    def reset_all_display(self):
        self.result = ""
        self.table_location = ""
        self.cus_order.set("")
        self.tax_bool.set(False)
        self.tax_num.set("")
        self.cus_email.set("")
        self.is_tax.set("")
        self.cus_name.set("")
        self.cus_address = ""
        self.cus_remark = ""
        self.order_note = ""
        self.update_gui('', self.display_cus_address)
        self.cus_province.set("")
        self.cus_district.set("")
        self.cus_sub_district.set("")
        self.cus_tel.set("")
        self.cus_cur_status.set("")
        self.cus_account_name.set("")
        self.display_is_tax.config(
            background="#FFF", foreground="#000", font='Chiller 10 normal')
        for i in self.tree.get_children():
            self.tree.delete(i)

    def update_log(self, update_txt):
        self.update_txt = update_txt
        self.report_log.config(state=NORMAL)
        self.report_log.insert(END, self.update_txt + "\n")
        self.report_log.config(state=DISABLED)

    def update_mp_frame(self, data_list):
        data_list
        self.report_log.config(state=NORMAL)
        self.report_log.insert(END, self.update_txt + "\n")
        self.report_log.config(state=DISABLED)

    def define_marketplace(self):
        file_input = self.table_location
        df = pd.read_excel(file_input)
        search_words = ['shopee', 'lazada']
        matches = [
            word for col in df.columns for word in search_words if word.lower() in col.lower()]

        # # เอา Dataframe มา groupby
        # if matches[0].lower() == 'lazada':
        #     self.group_by_order(file_input)

        if all(item == matches[0] for item in matches):
            return matches[0].upper()
        else:
            raise ValueError(
                "Error: Cannot varify the marketplace from this file, check the file you've imported")

    # def read_accel_file(self, accel_file_dir):
    #     # * accel data frame เราจะใช้แปลงค่า
    #     self.accel_df_state = pd.read_excel(self.accel_file_dir, dtype=str)

    #     # * สองบรรทัดล่างนี้ คือลอง ทำให้ bot มัน auto sn แบบหลาย sku
    #     accel_file_columns = self.accel_df_state.columns.dropna().tolist()
    #     self.obj_data_from_accel_file = {
    #         col: self.accel_df_state[col].replace(" ", '').dropna().tolist() for col in accel_file_columns
    #     }

    #     self.accel_orders_list = self.accel_df_state['orders'].dropna().tolist()
    #     # self.sn_list = self.accel_df_state['sn'].dropna().tolist()
    #     self.CP_list = self.accel_df_state['cp'].dropna().tolist()
    #     print(self.accel_orders_list)
    #     print('self.obj_data_from_accel_file: ', self.obj_data_from_accel_file)
    #     # print(self.sn_list)
    #     print(self.CP_list)
    #     pass

    def select_accel_file(self):
        # * รับ dir ของไฟล์
        self.accel_file_dir = filedialog.askopenfilename(
            title="Select Accel File")
        if self.accel_file_dir:
            self.accl_dir_namedisplay_on_btn.config(
                text=f"{self.accel_file_dir.split('/')[-1]}")
        else:
            self.accl_dir_namedisplay_on_btn.config(
                text=f"ยังไม่เลือก Accel File")

        # * accel data frame เราจะใช้แปลงค่า
        self.accel_df_state = pd.read_excel(self.accel_file_dir, dtype=str)
        print("before self.accel_df_state: ", self.accel_df_state)
        self.accel_df_state.loc[self.accel_df_state.duplicated(
            subset=['orders']), 'orders'] = pd.NA
        self.accel_df_state['orders'].dropna(inplace=True)
        print("after self.accel_df_state: ", self.accel_df_state)

        # * สองบรรทัดล่างนี้ คือลอง ทำให้ bot มัน auto sn แบบหลาย sku
        accel_file_columns = self.accel_df_state.columns.dropna().tolist()
        self.obj_data_from_accel_file = {
            col: self.accel_df_state[col].replace(" ", '').dropna().tolist() for col in accel_file_columns
        }

        self.accel_orders_list = self.accel_df_state['orders'].dropna(
        ).tolist()
        # self.sn_list = self.accel_df_state['sn'].dropna().tolist()
        self.CP_list = self.accel_df_state['cp'].dropna().tolist()
        print(self.accel_orders_list)
        print('self.obj_data_from_accel_file: ', self.obj_data_from_accel_file)
        print(self.CP_list)

    # * update accel file ******************************************************************
    # * เมื่อมีการใช้ SN ภายใน Accel file
    def deduct_accel_file_data(self, order, sku_serials=[]):
        order = order.get()
        df = self.accel_df_state
        print("deduct_accel_file_data df มีมาก่อนเหรอ: ", df)
        print("deduct_accel_file_data order: ", order)
        print("deduct_accel_file_data ref: ",
              df.loc[df['orders'] == order, 'orders'])
        # * ใช้ loc ของ df โดยดูว่า column 'orders' == order ที่รับเข้ามาหรือไม่, โดยให้ดึงค่าจาก column orders
        has_order = df.loc[df['orders'] == order, 'orders']
        if not has_order.empty:
            df.loc[df['orders'] == order, 'orders'] = ''

        print("sku_serials ไม่ได้ได้ไง: ", sku_serials)
        if sku_serials:
            for sn in sku_serials:
                df.loc[df[sn['sku']] == sn['sn'], sn['sku']] = ''
        df.to_excel(self.accel_file_dir, sheet_name='Sheet1', index=False)

    # todo WIP transfer to accel
    def extract_sn_btn(self, accel_file_dir):
        if not accel_file_dir:
            print("select accel file first!!")
            return

        target_dirs: tuple = filedialog.askopenfilenames(
            title="Select SN PDF files")
        if len(target_dirs) != 0:
            for target_dir in target_dirs:
                self.sn_extractor(accel_file_dir, target_dir)
        else:
            print("You have not selected any transfer file, Extraction ends!!")
        self.accel_df_state = pd.read_excel(self.accel_file_dir, dtype=str)

    def sn_extractor(self, output_excel, target_dir):
        extracted_txt: str = ""
        reader = PdfReader(target_dir)
        # * สกัดเอา ข้อความออกมาจากไฟล์
        for page in reader.pages:
            extracted_txt += page.extract_text()

        pattern = r'^.*?(?=No\. Product Code Barcode Product Name Transfer No\. Order Ship Status)'
        extracted_txt = re.sub(pattern, '', extracted_txt, flags=re.DOTALL)
        extracted_txt = extracted_txt.lstrip()

        pattern2 = r'ผู้ส่งสินค้า.*?(?:No\. Product Code Barcode Product Name Transfer No\. Order Ship Status|วันที่ _ _ _ / _ _ _ / _ _ _)'
        extracted_txt = re.sub(pattern2, '', extracted_txt, flags=re.DOTALL)

        pattern_serial = r'Serial\s:'
        extracted_txt = re.sub(
            pattern_serial, '', extracted_txt, flags=re.DOTALL)

        pattern_sku_no = r'\d+\s{0,}(?=([A-Z0-9]{3}-[0-9]{6}))'
        extracted_txt = re.sub(
            pattern_sku_no, '', extracted_txt, flags=re.DOTALL)

        # * สกัดเอาค่าที่จำเป็นออกจากข้อความทั้งหมด
        # * Regular expression สำหรับการจับ SKU
        sku_pattern = r'([A-Z0-9]{3}-[0-9]{6})'

        # *Regular expression สำหรับการจับ serial numbers
        # serial_pattern = r'Shipped\s*([\w, \n]+)(?=(?:[A-Z0-9]{3}-[0-9]{6}|\nผู้ส่งสินค้า|$))'
        serial_pattern = r'(?:Shipped|Confirm)\s*([\w, \n]+)(?=(?:[A-Z0-9]{3}-[0-9]{6}|\nผู้ส่งสินค้า|$))'

        # * สกัด SKU
        product_codes = re.findall(sku_pattern, extracted_txt)

        # * สกัด serial numbers
        serial_numbers = re.findall(serial_pattern, extracted_txt, re.DOTALL)

        print(serial_numbers)

        cleaned_serial_numbers = []
        for serial in serial_numbers:
            # * ลบช่องว่างและเลขลำดับที่ไม่ต้องการออก
            # * ลบเลขลำดับที่ท้าย
            cleaned_serial = re.sub(r'\n', '', serial).strip()
            # cleaned_serial = re.sub(r'\s+', '', cleaned_serial)  #* ลบช่องว่างทั้งหมด
            cleaned_serial_numbers.append(cleaned_serial)

        # * แสดงผล
        print("Product Codes:")
        code_count = 0
        for code in product_codes:
            code_count += 1
            print(code_count, " ", code)

        print("\nSerial Numbers:")
        code_count = 0
        for serial in cleaned_serial_numbers:
            code_count += 1
            # * ลบช่องว่างและเพิ่มวงเล็บ [] รอบ Serial Numbers
            serial = serial.replace(" ", "")
            serial_list = serial.split(",")
            print(f"{code_count} {len(serial_list)} [{serial}]")

        # * จัดการ serial numbers ให้เป็น list ของแต่ละ SKU
        # serial_numbers_grouped = [serial.strip().replace('\n', '').replace(' ', '').split(',') for serial in serial_numbers]
        serial_numbers_grouped = [re.findall(
            r'\b[\w]+\b', serial) for serial in cleaned_serial_numbers]

        # ตรวจสอบข้อมูลที่ถูกสกัด
        print("SKU Matches:")
        print(len(product_codes), product_codes)
        print("Serial Numbers Grouped:")
        print(len(serial_numbers_grouped), serial_numbers_grouped)

        # * สร้าง DataFrame ที่แต่ละคอลัมน์เป็น SKU และแต่ละ row เป็น serial number
        data = {sku: serials for sku, serials in zip(
            product_codes, serial_numbers_grouped)}

        # ตรวจสอบ DataFrame ก่อนเขียนลงไฟล์
        # print("DataFrame:")

        # * เอาเข้าตาราง
        try:
            # โหลด workbook และ sheet ล่าสุด
            book = load_workbook(output_excel)
            sheet = book.active

            # หาคอลัมน์ล่าสุดที่มีข้อมูล
            last_column = sheet.max_column

            # เขียนข้อมูลลงใน Excel
            for col, (sku, serials) in enumerate(data.items(), start=last_column+1):
                sheet.cell(row=1, column=col, value=sku)
                for row, serial in enumerate(serials, start=2):
                    sheet.cell(row=row, column=col, value=serial)

            # บันทึกไฟล์
            book.save(output_excel)
            print(f"ข้อมูลถูกเพิ่มลงใน {output_excel} เรียบร้อยแล้ว")
        except Exception as e:
            print(f"เกิดข้อผิดพลาด: {e}")
            import traceback
            traceback.print_exc()

    def select_excel(self):
        self.result = "Excel"
        print("Select Excel")
        self.table_location = filedialog.askopenfilename(
            title="Select Shopee order toship file")
        # * ตัดเอาเฉพาะ ชื่อไฟล์
        self.display_location_result.config(
            text=f"{self.table_location.split('/')[-1]}")

        # * target should come before get dataframe
        self.marketplace_target.set(self.define_marketplace())
        result = self.marketplace_target.get()
        print("ต้องตีเว็บไหน", result)
        # self.canvas.config(bg=f'{self.bg_by_market_place[self.marketplace_target.get()}')
        self.entry_frame.config(
            bg=f'{self.bg_by_market_place[str(result)]}')
        self.marketplace_label.config(
            bg=f'{self.bg_by_market_place[str(result)]}')
        # self.import_file_frame.config(
        #     bg=f'{self.bg_by_market_place[self.marketplace_target.get()]}')

        # * หลังจากได้ไฟล์เข้ามาแล้ว (self.table_location) เราจะทำการสร้างเป็น dataframe ด้วย function get_data_frame()
        self.get_data_frame()
        print("Table Location:", self.table_location)
        self.update_log("แอดไฟล์")

    def group_by_order(self, file_input, dtype):
        df = pd.read_excel(file_input, dtype=dtype)
        #! สำคัญมาก ถ้าอยากให้ nan หาย เอา dfมาใช้ method fillna('', inplace=True) "//การใช้ Inplace ทำให้แก้ ที่ df โดยตรงโดยไม่ต้องเก็บค่าใหม่
        # df.fillna('', inplace=True)

        # เพิ่มส่วนที่ไม่มี และหาไม่ได้
        df['ส่วนลดจาก Shopee'], df['ประเภทใบกำกับภาษี'], df['โค้ดส่วนลดชำระโดย Shopee'], df[
            'ประเภทสาขา'], df['หมายเหตุจากผู้ซื้อ'], df['บันทึก'] = 0.00, "", 0, "", "", ""

        # กำหนด Datatype
        data_types = {'orderNumber': str, 'ส่วนลดจาก Shopee': float, 'ประเภทใบกำกับภาษี': str,
                      'โค้ดส่วนลดชำระโดย Shopee': float, 'ประเภทสาขา': str, 'หมายเหตุจากผู้ซื้อ': str, 'บันทึก': str, 'paidPrice': float, 'variation': str, 'billingAddr': str, 'createTime': str, 'branchNumber': str, 'billingAddr2': str, 'customerEmail': str, 'taxCode': str, 'billingAddr3': str, 'billingAddr4': str, 'billingAddr5': str, 'billingName': str, 'billingPhone': str, 'customerName': str, 'shippingFee': float, 'sellerDiscountTotal': float, 'unitPrice': float}
        df = df.astype(data_types)

        # อุดค่าว่างก่อนไม่งั้น จะใช้ size() ไม่ได้
        #! คำเตือน ถ้าใช้ parameter inplace= ห้ามเก็บค่าเข้าตัวแปรเดิมเด็ดขาด ไม่งั้นมันจะพัง มันจะ error nontype รัวๆ
        # df['variation'].fillna(value="-", inplace=True)
        # ตรวจสอบ columntype
        column_type = df.dtypes
        print("createTime: ", column_type['createTime'])
        print("createTimeerr?:", df['createTime'])
        print("column_type: ", column_type['variation'])
        for column in df.columns:
            # print("ทำไมคืนค่า0: ", column_type[column])
            if column_type[column] == 'float':
                df[column] = df[column].replace(np.nan, 0)
            elif column_type[column] == 'object':
                df[column] = df[column].replace('nan', '')
            elif column_type[column] == 'str':
                df[column] = df[column].replace('nan', '')
            else:
                df.fillna(np.nan, inplace=True)

        # เพิ่มส่วนที่ไม่มี แต่สามารถหาคำนวณเพิ่มเองได้
        result_count = df.groupby(['orderNumber', 'sellerSku', 'itemName',
                                  'unitPrice', 'variation']).size().reset_index(name='จำนวน')

        result_with_additional_columns_df = df.groupby('orderNumber').agg({
            'status': 'first',
            'ส่วนลดจาก Shopee': 'first',
            'ประเภทใบกำกับภาษี': 'first',
            'customerEmail': 'first',
            'โค้ดส่วนลดชำระโดย Shopee': 'first',
            'ประเภทสาขา': 'first',
            'หมายเหตุจากผู้ซื้อ': 'first',
            'บันทึก': 'first',
            'billingName': 'first',
            'billingAddr': 'first',
            'billingAddr2': 'first',
            'billingAddr4': 'first',
            'billingAddr3': 'first',
            'billingAddr5': 'first',
            'taxCode': 'first',
            'billingPhone': 'first',
            'customerName': 'first',
            'paidPrice': 'sum',
            'createTime': 'first',
            'branchNumber': 'first'
        })
        result_with_additional_columns_df = result_with_additional_columns_df.astype(
            {'billingAddr5': str, 'billingPhone': str})

        # เก็บไว้ก่อน['billingName', 'billingAddr', 'billingAddr2', 'billingAddr4', 'billingAddr3', 'billingAddr5', 'taxCode', 'billingPhone', 'customerName', 'paidPrice', 'createTime', 'branchNumber']

        # ** ปรับแต่ง Column สำหรับ LAZADA--------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
        # * สร้าง sum_column  ขึ้นมาใหม่ --------------------------------------------------------
        # *> 'ราคาขายสุทธิ'
        # total_per_order_df = df.groupby('orderNumber')[
        #     'unitPrice'].sum().reset_index(name='ราคาขายสุทธิ')
        result_count['ราคาขายสุทธิ'] = result_count["จำนวน"] * \
            result_count["unitPrice"]

        # *> 'ชื่อผู้รับ'
        result_count['ชื่อผู้รับ'] = df['billingName']

        # *> 'หมายเลขโทรศัพท์'
        result_count['หมายเลขโทรศัพท์'] = df['billingPhone']

        # *> 'โค้ดส่วนลดชำระโดยผู้ขาย'
        total_sellerDiscountTotal_df = df.groupby('orderNumber')[
            'sellerDiscountTotal'].sum().reset_index(name='โค้ดส่วนลดชำระโดยผู้ขาย')
        total_sellerDiscountTotal_df['โค้ดส่วนลดชำระโดยผู้ขาย'] *= -1

        # *> 'ค่าจัดส่งที่ชำระโดยผู้ซื้อ'
        total_shippingfee_df = df.groupby('orderNumber')['shippingFee'].sum(
        ).reset_index(name='ค่าจัดส่งที่ชำระโดยผู้ซื้อ')

        # * ปรับแต่งค่าใน Column
        # result_with_additional_columns_df = result_with_additional_columns_df['branchNumber'].map(lambda x: )

        # *  รวม dataframe เป็น dataframe ใหม่
        # merge1_df = pd.merge(result_count, total_per_order_df,
        #                      on='orderNumber', how='left')
        merge2_df = pd.merge(result_count, total_sellerDiscountTotal_df,
                             on='orderNumber', how='left')
        merge3_df = pd.merge(
            merge2_df, result_with_additional_columns_df, on='orderNumber', how='left')
        result_df = pd.merge(merge3_df, total_shippingfee_df,
                             on='orderNumber', how='left')
        # result_df = pd.concat([result_count, total_per_order_df,
        #                    total_sellerDiscountTotal_df, total_shippingfee_df], ignore_index=True)

        # * เราต้องการ column ที่มีชื่อต่างกัน แต่ข้อมูลเหมือนกัน เลยต้อง copy column เพิ่ม
        result_df['รายละเอียดที่อยู่'] = result_df['billingAddr'].copy()

        result_df['ประเภทสาขา'] = result_df['branchNumber'].copy()
        print("result_df d-type", type(result_df['ประเภทสาขา']))

        # * สกัดและหาเลขสาขา จากข้อมูลที่กรอกมั่วๆไร้ซึ่ง pattern จาก lazada exportfile และเก็บไว้ในตัวแปร extracted_branch_df สาขาจะแสดงเป็นเลข 5 หลักแทนช่องว่างด้วย 0 แต่สาขา 00000 จะแสดงเป็น "สำนักงานใหญ่"
        extracted_branch_df = result_df['ประเภทสาขา'].apply(
            self.find_branch)

        # * เปลี่ยน ค่าใน col branchNumber ให้กลายเป็นบอกเฉพาะเลขสาขาถ้าเป็นสาขาย่อย และ เป็นค่าว่างถ้าเป็นสำนักงานใหญ่
        result_df['branchNumber'] = extracted_branch_df.copy()
        result_df['branchNumber'] = result_df['branchNumber'].map(
            lambda row: "" if row == "สำนักงานใหญ่" else row)

        # * นำค่าที่สกัดและแปลงจากตัวแปร extracted_branch_df มาหาประเภทสาขา หาก ค่าใน cell เป็น"สำนักงานใหญ่" จะ return "สำนักงานใหญ่" ถ้าไม่ใช่ จะแสดงเป็น "สาขาย่อย" (มีค่าเป็นเลขสาขา จะ return เป็น สาขาย่อย)
        result_df['ประเภทสาขา'] = extracted_branch_df.map(
            lambda row: "สำนักงานใหญ่" if row == "สำนักงานใหญ่" else "สาขาย่อย")

        # * แยกย่อยออกมาจาก บรรทัดบนเพื่อ กรองส่วนที่ไม่มีเลขให้เป็นคำว่า สาขาย่อย เพราอันบน มันจะเปนสำนักงานใหญ่หมดถ้าหากหาค่าไม่ได้
        # result_df['ประเภทสาขา'] = result_df['taxCode'].apply(
        #     lambda row: "สาขาย่อย" if len(row) == 0 else "สำนักงานใหญ่")
        result_df['ประเภทสาขา'] = result_df['taxCode'].apply(
            lambda row: "สาขาย่อย" if (isinstance(row, str) and len(row) == 0) else "สำนักงานใหญ่")
        # result_df['ประเภทสาขา'] = result_df['taxCode'].apply(
        #     lambda row: print("ทำไมไม่ได้สาขาย่อยวะ", row) if (isinstance(row, str) and len(row) == 0) else print("ทำไมไม่ได้สาขาย่อยวะ", type(row)))
        result_df['ประเภทสาขา'] = result_df['taxCode'].apply(
            lambda row: "สาขาย่อย" if (pd.notna(row) and isinstance(row, str) and len(
                row) == 0) else "สำนักงานใหญ่" if isinstance(row, str) else "สาขาย่อย"
        )

        # * เปลี่ยนค่าใน Col billingAddrs ตัดภาษาอังกฤษออก เนื่องจาก ที่อยู่ที่ได้จาก exportfile laz จะมี pattern เป็น ไทย/ อังกิก เช่น "บางปะกง/ Bang Pakong"
        # >> addr4 = เขต/อำเภอ, addr3 = จังหวัด
        address_divs = ['billingAddr4', 'billingAddr3']
        for address_div in address_divs:
            result_df[f'{address_div}'] = result_df[f'{address_div}'].map(
                lambda row: row.split('/')[0].strip())

        # * เปลี่ยน Dtype ของ Column ['createTime'] (วันที่ทำการสั่งซื้อ) จาก Series ให้เป็นobjวันที่ เนื่องจากอันเดิมมันเอาไป Sort ไม่ได้ เวลาออกเป็นตาราง
        result_df['createTime'] = pd.to_datetime(
            result_df['createTime'], format='mixed', dayfirst=True)
        # * >  แปลง objวันที่ ให้กลายเป็น number ใน excel เพื่อให้แสดงผลใน cel เหมือนกับ exported file ของ shopee
        result_df['createTime'] = result_df['createTime'].dt.strftime(
            '%Y-%m-%d %H:%M')

        # * ตรวจสอบผลลัพธ์
        # print(f"""qty ใน lazada""")
        # print(result_count)

        # print("ราคาขายสุทธิ")
        # print(total_per_order_df)

        # * เปลี่ยนชื่อ column // เปลี่ยนชื่อ column // เปลี่ยนชื่อ column // เปลี่ยนชื่อ column // เปลี่ยนชื่อ column // เปลี่ยนชื่อ column // เปลี่ยนชื่อ column
        result_df.rename(columns={
            'orderNumber': 'หมายเลขคำสั่งซื้อ',
            'sellerSku': 'เลขอ้างอิง SKU (SKU Reference No.)',
            'itemName': 'ชื่อสินค้า',
            'unitPrice': 'ราคาขาย',
            'status': 'สถานะการสั่งซื้อ',
            'billingName': 'ชื่อ',
            'billingAddr': 'ที่อยู่สำหรับออกใบกำกับภาษีแบบเต็มรูป',
            'billingAddr2': 'แขวง/ตำบล',
            'billingAddr4': 'เขต/อำเภอ.1',
            'billingAddr3': 'จังหวัด.1',
            'billingAddr5': 'รหัสไปรษณีย์.1',
            'taxCode': 'หมายเลขประจำตัวผู้เสียภาษี',
            'billingPhone': 'หมายเลขโทรศัพท์สำหรับออกใบกำกับภาษี',
            'customerName': 'ชื่อผู้ใช้ (ผู้ซื้อ)',
            'paidPrice': 'จำนวนเงินทั้งหมด',
            'createTime': 'วันที่ทำการสั่งซื้อ',
            'branchNumber': 'รหัสประจำสาขา',
            'variation': 'ชื่อตัวเลือก',
            'customerEmail': 'อีเมลสำหรับรับใบกำกับภาษี'
        },
            inplace=True
        )
        result_df['หมายเลขคำสั่งซื้อ'] = result_df['หมายเลขคำสั่งซื้อ'].astype(
            str)

        print("ตารางใหม่")
        print(result_df)
        excel_file_path = "output_test.xlsx"
        result_df.to_excel(excel_file_path, index=False,
                           na_rep="", engine="openpyxl")
        return result_df

    def f(self, d):
        return '{0:n}'.format(d)

    def get_data_frame(self):
        print("มีป่าวหว่า", self.table_location)
        self.file_path = self.table_location
        print('self.marketplace_target.get()', self.marketplace_target.get())
        shopee = {'หมายเลขประจำตัวผู้เสียภาษี': str, 'รหัสไปรษณีย์.1': str, 'หมายเลขโทรศัพท์สำหรับออกใบกำกับภาษี': str, 'จำนวน': int, 'ค่าจัดส่งที่ชำระโดยผู้ซื้อ': float, 'โค้ดส่วนลดชำระโดยผู้ขาย': float, 'แขวง/ตำบล': str, 'ประเภทสาขา': str,
                  'สาขาย่อย': str, 'รหัสประจำสาขา': str, 'หมายเหตุจากผู้ซื้อ': str, 'บันทึก': str}
        lazada = {'หมายเลขประจำตัวผู้เสียภาษี': str, 'รหัสไปรษณีย์.1': str, 'หมายเลขโทรศัพท์สำหรับออกใบกำกับภาษี': str, 'จำนวน': int, 'ค่าจัดส่งที่ชำระโดยผู้ซื้อ': float, 'โค้ดส่วนลดชำระโดยผู้ขาย': float, 'แขวง/ตำบล': str, 'ประเภทสาขา': str,
                  'สาขาย่อย': str, 'รหัสประจำสาขา': str, 'หมายเหตุจากผู้ซื้อ': str, 'บันทึก': str, 'taxCode': str}
        self.columns_preset = shopee if self.marketplace_target.get(
        ) == 'SHOPEE' else lazada if self.marketplace_target.get() == 'LAZADA' else ''
        try:
            if self.marketplace_target.get() == 'SHOPEE':
                print("เปลี่ยน dtype เป็น form shopee")
                self.data_frame = pd.read_excel(
                    self.file_path, dtype=self.columns_preset)
                self.data_frame['หมายเลขประจำตัวผู้เสียภาษี'].astype(str)
            elif self.marketplace_target.get() == 'LAZADA':
                self.data_frame = self.group_by_order(
                    self.file_path, self.columns_preset)

                self.data_frame['โค้ดส่วนลดชำระโดยผู้ขาย'].astype(float)
                self.data_frame['หมายเลขประจำตัวผู้เสียภาษี'].astype(str)

            print("df มี type เป็นไร", type(self.data_frame))
            # print("self.data_frame หน้าตาเปนไง: ", self.data_frame) มันยาว
            if self.data_frame.empty:
                print("ไม่มี Data Frame")
            else:
                print("มี Data Frame")
        except FileNotFoundError:
            print("File not found.")
        except NameError as e:
            print(f"ตัวแปร '{e.name}' ไม่มีอยู่จริง")
        # except Exception as e:
        #     print(f"อะไรสักอย่างพัง {e}")

    def update_gui(self, address, widget):
        if address != "":
            input = address.strip()
        else:
            input = "-"

        self.cus_address = input
        widget.config(state=NORMAL)
        widget.delete(1.0, END)
        widget.insert(END, input)
        widget.config(state=DISABLED)

    def update_gui_remark(self):
        if self.cus_remark == "" or self.cus_remark == "nan":
            self.display_cus_remark.config(state=NORMAL)
            self.display_cus_remark.delete(1.0, END)
            self.display_cus_remark.insert(END, 'ไม่มี')
            self.display_cus_remark.config(state=DISABLED)

        else:
            self.display_cus_remark.config(state=NORMAL)
            self.display_cus_remark.delete(1.0, END)
            self.display_cus_remark.insert(END, self.cus_remark)
            self.display_cus_remark.config(state=DISABLED)

    def update_gui_note(self):
        if self.order_note == "" or self.order_note == "nan":
            self.display_order_note.config(state=NORMAL)
            self.display_order_note.delete(1.0, END)
            self.display_order_note.insert(END, 'ไม่มี')
            self.display_order_note.config(state=DISABLED)

        else:
            self.display_order_note.config(state=NORMAL)
            self.display_order_note.delete(1.0, END)
            self.display_order_note.insert(END, self.order_note)
            self.display_order_note.config(state=DISABLED)

    # * widget รายการสินค้า ///////////////////////////////////////////////
    def show_products(self, products_list):
        for i in self.tree.get_children():
            self.tree.delete(i)
        self.total_price = 0
        for product in products_list:
            product_name = product["เลขอ้างอิง SKU (SKU Reference No.)"]
            price = product["ราคาขายสุทธิ"]
            shopee_rebate = product['ส่วนลดจาก Shopee']
            price_plusrebate = price+shopee_rebate
            QTY = product['จำนวน']
            self.total_price += price_plusrebate
            self.tree.insert("", "end", values=(
                product_name, self.f(price_plusrebate), QTY))

        # * Shopee + ค่าขนส่ง แต่ Lazada ไม่ต้อง + ค่าขนส่งในบางกรณี
        if self.marketplace_target.get() == 'SHOPEE':
            self.total_price += self.cus_ship_cost.get()
            self.phase1_sum_price = self.total_price
            self.tree.insert("", "end", value=(
                "ค่าขนส่ง", self.f(self.cus_ship_cost.get()), 1))
            self.total_price -= self.cus_seller_voucher.get()
            self.tree.insert("", "end", value=(
                "Seller Voucher",  "-"+self.f(self.cus_seller_voucher.get()), 1))

            self.tree.insert("", "end", values=(
                "ราคาที่ต้องออก", self.f(self.total_price)))

            self.tree.insert("", "end", values=("Shopee Voucher",
                                                self.f(self.nondistortedData['โค้ดส่วนลดชำระโดย Shopee']*-1)))

            self.tree.insert("", "end", values=("ลูกค้าจ่ายทั้งหมด",
                                                self.f(self.nondistortedData['จำนวนเงินทั้งหมด'])))

        elif self.marketplace_target.get() == 'LAZADA':
            # * มันต้องมีทั้ง ราคาที่ต้องออกแบบ +ขนส่งกับ ไม่มีขนส่ง
            self.total_price -= self.cus_seller_voucher.get()
            self.tree.insert("", "end", value=(
                "Seller Voucher",  "-"+self.f(self.cus_seller_voucher.get()), 1))
            # > แบบไม่มีขนส่ง
            self.total_price_no_ship_cost = self.total_price
            self.tree.insert("", "end", values=(
                "ราคาที่ต้องออก(Noขนส่ง)", self.f(self.total_price_no_ship_cost)))

            # > แบบมีขนส่ง
            self.total_price_with_ship = self.total_price + self.cus_ship_cost.get()
            self.tree.insert("", "end", value=(
                "ค่าขนส่ง", self.f(self.cus_ship_cost.get()), 1))
            self.tree.insert("", "end", values=(
                "ราคาที่ต้องออก(+ขนส่ง)", self.f(self.total_price_with_ship)))

    def get_pure_address(self, cus_address):
        # สร้างรายชื่อของตำแหน่งที่พบคำใน customer_address
        keywords = ["เขต", "แขวง", "ต.", "ตำบล",
                    "อ.", "อำเภอ", "จ.", "จังหวัด"]
        positions = []

        for keyword in keywords:
            if keyword in cus_address:
                keyword_position = cus_address.find(keyword)
                positions.append(keyword_position)

        # หาตำแหน่งสูงสุดและตำแหน่งต่ำสุดในรายชื่อตำแหน่ง
        if positions:
            min_position = min(positions)
            max_position = max(positions)
        else:
            min_position = -1
            max_position = -1

        # ลบตั้งแต่ตำแหน่งที่พบคำไปจนสุดลงท้ายของ customer_address
        if min_position >= 0:
            truncated_address = cus_address[:min_position]
        else:
            truncated_address = cus_address

        print("Address_cleaned: ", truncated_address.strip())
        return truncated_address.strip()

    def clean_duplicate_parts(self, address):
        # ใช้ regex เพื่อค้นหาและลบคำย่อที่มีส่วนที่มากกว่าคำเต็ม
        pattern = r'(ต\..+?)\s+?(ตำบล|อ\..+?)\s+?(อำเภอ|จ\..+?)\s+?(จังหวัด)'

        matches = re.findall(pattern, address)
        if matches:
            cleaned_address = address
            for match in matches:
                full_word, abbr_word1, abbr_word2, abbr_word3 = match
                if len(full_word) > len(abbr_word1):
                    cleaned_address = cleaned_address.replace(
                        abbr_word1, full_word)
                if len(full_word) > len(abbr_word2):
                    cleaned_address = cleaned_address.replace(
                        abbr_word2, full_word)
                if len(full_word) > len(abbr_word3):
                    cleaned_address = cleaned_address.replace(
                        abbr_word3, full_word)
        else:
            cleaned_address = address
        print("After_Clean_dup: ", cleaned_address)
        return cleaned_address

    def clean_address(self, address):
        keywords = ["เขต", "แขวง", "ต.", "ตำบล",
                    "อ.", "อำเภอ", "จ.", "จังหวัด"]

        # ตรวจสอบว่าสตริงมีคำ "จังหวัด" และ ("เขต" หรือ "แขวง") หรือไม่
        if "จังหวัด" in address and any(keyword in address for keyword in ["เขต", "แขวง"]):
            # ลบคำ "จังหวัด" ออกจากสตริง
            address = address.replace("จังหวัด", "")

        if "\n" in address:
            address = address.replace('\n', " ")

        # เริ่มต้นโดยการแยกคำด้วยช่องว่าง
        parts = address.split()

        # สร้าง list เพื่อเก็บคำที่ไม่ใช่คำย่อ
        cleaned_parts = []

        for part in parts:
            # ตรวจสอบว่าคำนี้เป็นคำย่อหรือไม่
            is_abbreviation = any(part.startswith(keyword)
                                  for keyword in ["ต.", "อ.", "จ."])
            if not is_abbreviation:
                cleaned_parts.append(part)

        # นำคำที่ไม่ใช่คำย่อมาเชื่อมกลับเป็นสตริงใหม่
        cleaned_address = ' '.join(cleaned_parts)

        # ลบคำที่มีส่วนที่เหมือนกันออก
        cleaned_address = self.clean_duplicate_parts(cleaned_address)

        # แก้ไขเครื่องหมายช่องว่างที่เหลือหลังการลบคำ
        cleaned_address = cleaned_address.replace("  ", " ")

        return cleaned_address

    # ? WIP note_extractor ยังไม่ค่อยสมบูรณ์ เพราะ case ตัวอย่างมันนานๆเจอที
    def note_extractor(self):
        if self.order_note != 'nan':
            try:
                self.name_match = re.search(
                    r'ชื่อ\s*:?\s*(.*)', self.order_note)
            except:
                self.name_match = re.search(r'บริษัท.*', self.order_note)

            # * ถ้ากลับมาดูไม่ต้องสงสัยว่าแยกทำไม พอเขียนติดกันแล้วมันดูสับสน เลยแยกเฉยๆไม่มีไร (A1/2)
            if "บริษัท" in self.name_match.group():
                self.branch_match = re.search(
                    r'สาขา\s*:?\s*(.*)', self.order_note)
                self.tax_id_match = re.search(
                    r'Tax id\s*:?\s*(.*)', self.order_note)
                self.email_match = re.search(
                    r'email\s*:?\s*(.*)', self.order_note.lower())
                self.tel_match = re.search(
                    r'tel\s*:?\s*,?(.*)', self.order_note.lower())

            self.address_match = re.search(
                r'ที่อยู่\s*:?\s*(.*)', self.order_note)

            print("try: regexบันทึก: ", self.name_match)
            print("try: ใช้ group กับ regexบันทึก: ", self.name_match.group(1))

            # * เก็บค่าเข้าตัวแปร //#* ถ้ากลับมาดูไม่ต้องสงสัยว่าแยกทำไม พอเขียนติดกันแล้วมันดูสับสน เลยแยกเฉยๆไม่มีไรจะรวมกันก็ได้ (A2/2)
            if "บริษัท" in self.name_match.group():
                self.tax_branch_num.set(self.branch_match.group(
                    1)) if self.branch_match else self.tax_branch_num.set(self.tax_branch_num.get())
                self.tax_num.set(self.tax_id_match.group(
                    1)) if self.tax_id_match else self.tax_num.set(self.tax_num.get())
                self.cus_email.set(self.email_match.group(
                    1)) if self.email_match else self.cus_email.set(self.cus_email.get())
                self.cus_tel.set(self.tel_match.group(
                    1)) if self.tel_match else self.cus_tel.set(self.cus_tel.get())

            self.cus_name.set(self.name_match.group(
                1)) if self.name_match else self.cus_name.set(self.cus_name.get())
            self.note_extracted_address = self.address_match.group(
                1) if self.address_match else "-"
            self.cus_address = self.note_extracted_address
        else:
            print("no note to be extracted, note_extractor was not used")

    def translator(self, text):
        # ตรวจสอบว่าชื่อไม่ใช่ภาษาไทย, อังกฤษ, หรือตัวเลข
        #! Patterns เก่าๆ
        #! pattern = re.compile(r'^[a-zA-Z0-9ก-๙\s\W\_]+$')
        #! pattern = re.compile(r'^[a-zA-Z0-9ก-๙\s\W_!@#$%^&*()\-_=+/[\]{}|;:\'",<.>/?]+$') ต่างกันตรงที่มี "     กับไม่มี " ตอน testใน regex101 มันใส่ " แล้ววมันไม่ทำงาน ซึ่งพอลบออกมาก็ใช้งานได้ปกติ

        # * สำหรับแก้ปัญหาข้อที่ 38 // pattern ใหม่
        pattern = re.compile(
            r'^[a-zA-Z0-9ก-๙\s\W_!@#$%^&*()\-_=+/[\]{}|;:\',<.>/?]+$')
        is_usable = bool(re.match(pattern, text))
        if is_usable:
            return text
        else:

            try:
                translator = Translator()
                lang_src = translator.detect(text).lang
                print("Whare are you from: ", lang_src)
                translation = translator.translate(
                    text, src=lang_src, dest='en')
                print("Translated name", translation.text)
                return translation.text
            except:
                print("google ก็แปลให้ไม่ได้เอาชื่อ", text, "ไปแทนนะ")
                return text

    def cus_tel_fixer(self, tel):
        if len(tel) == 10:
            print("เบอร์มือถือ")
            return tel
        elif len(tel) == 9:
            print("ดูก่อนว่าเบอร์บ้านไหม")
            if tel[0] == "0":
                print('โอเคเบอร์บ้าน')
                return tel
            else:
                print("ลืมใส่เลข 0 แหละ")
                print("เติม 0 ให้", "0"+tel)
                return tel
        elif len(tel) > 10:
            if len(tel) == 11:
                tel_no_code = tel[-8:]
                print('เบอร์บ้านแต่ติดรหัสประเทศ')
            elif len(tel) == 12:
                tel_no_code = tel[-9:]
                print('เบอร์มือถือแต่ติดรหัสประเทศ')

            if tel_no_code[0] != 0:
                print("ตัดรหัสประเทศและเพิ่มเลข 0")
                fixed_tel = "0" + str(tel_no_code)
                print('')
                return fixed_tel
            else:
                print('holy shetttttttttttt+')

    def find_branch(self, input):
        # จะ method นี้ จะ return ไม่ "สำนักงานใหญ่" ก็ เลขสาขาที่เป็นเลข 5 หลัก เท่านั้น
        # ตัวแปร branch
        input = re.sub(r'\s+', '', str(input))
        branch = str(input).strip()

        pattern = re.compile(
            r"สำนักงานใหญ่|ใหญ่|สนงใหญ่|สนง\.ใหญ่|สนง|Head|สนญ|^0+$")
        match = pattern.findall(branch)
        # ตรวจสอบค่าของตัวแปร branch
        if match:
            return 'สำนักงานใหญ่'

          # เมื่อไม่มีสำนักงานใหญ่ ให้ดูว่ามีเลขไหม
        elif re.findall(r'[0-9]+', branch):
            #   เมื่อมีเลขให้ดูว่ามีคำว่าสาขากับเลขหรือไม่
            # print("2nd condition", branch)
            if re.search(r'สาขา.*[0-9]', branch):
                matches = re.search(r'[0-9]+', branch)
                match = matches[0]
                match = match.strip()
                if len(match) == 5 and match[0] == "0":
                    # print("ตัดสาขาออกแล้วมีเลขครบ 5 หลักพอดี", match)
                    return match

                elif len(match) < 5:
                    txt = "{:0>5}"
                    # print("เลขที่ได้หลังตัดสาขาออก มีหลักไม่ครบ เติมหลัก แล้ว Return")
                    # print("ปรับ Format เป็น 5 หลัก")
                    # print("return", txt.format(match))
                    match = txt.format(match)
                    return match

            elif re.search(r'0[0-9]{0,4}', branch) and (branch[0] == "0" and len(branch) <= 5):
                branch = re.search(r'0[0-9]{0,4}', branch)[0]
                # print(branch)
                # print("เลขสาขาตรงๆ ครบบ้างไม่ครบบ้าง")
                if len(branch) == 5:
                    # print("Return แค่ 5 ตัวท้าย")
                    # print("return", branch[-5:])
                    branch = branch[-5:]
                    return branch

                elif len(branch) < 5:
                    txt = "{:0>5}"
                    # print("ปรับผลลัพธ์ ให้ Return เป็น Format 5 หลัก")
                    # print("return", txt.format(branch))
                    return txt.format(branch)
                else:
                    # print("บิดเบี้ยว", branch)
                    # print("return สำนักงานใหญ่")
                    return "สำนักงานใหญ่"
            else:
                # print("not match any subcondition in the 2nd condition")
                # print("return สำนักงานใหญ่")
                return "สำนักงานใหญ่"

        else:
            # print("ไม่มีบอกสำนักงาน")
            # print("return ค่า สำนักงานใหญ่ แล้วกัน")
            return "สำนักงานใหญ่"

    def branch_to_branch_type(self, branch):
        if 'สำนักงานใหญ่' in branch:
            return 'สำนักงานใหญ่'
        else:
            return ''

    def cus_name_simplifyer(self, name):
        if name:
            self.cus_name.set(self.translator(
                re.sub(r'\s{2,}', " ", name.strip().replace('\u200b', ''))))

            # *  ตัดพวก non-ASCII values // ref https://stackoverflow.com/questions/20889996/how-do-i-remove-all-non-ascii-characters-with-regex-and-notepad
            self.cus_name.set(
                re.sub(r'[^\x00-\x25\x27-\x7F\wA-Zก-๙|/]+', '', self.cus_name.get().strip()))

            # * ปรับคำบอกประเภทการจดทะเบียนของใบกำกับ
            self.cus_name.set(self.tax_name_standardizer(self.cus_name.get()))
            print("name.get()หลังจากทำการ standarrdizer", self.cus_name.get())
        else:
            print("Customer Name is empty")

    def order_search(self, order,  on_complete):
        print("order_search ทำงาน")
        self.on_complete = on_complete
        self.order = order.strip()
        if len(self.order) < 14:
            raise ValueError("Order num is weird")
        self.cus_order.set(self.order)
        differential_col_data = ['เลขอ้างอิง SKU (SKU Reference No.)', 'ชื่อสินค้า',
                                 'ราคาขาย', 'จำนวน', 'ราคาขายสุทธิ', 'ส่วนลดจาก Shopee', 'ชื่อตัวเลือก']
        non_differential_col_data = [
            'หมายเลขคำสั่งซื้อ',
            'สถานะการสั่งซื้อ',
            'โค้ดส่วนลดชำระโดยผู้ขาย',
            'ค่าจัดส่งที่ชำระโดยผู้ซื้อ',
            'ประเภทใบกำกับภาษี',
            'ชื่อ',
            'ที่อยู่สำหรับออกใบกำกับภาษีแบบเต็มรูป',
            'แขวง/ตำบล',
            'เขต/อำเภอ.1',
            'จังหวัด.1',
            'รหัสไปรษณีย์.1',
            'หมายเลขประจำตัวผู้เสียภาษี',
            'หมายเลขโทรศัพท์สำหรับออกใบกำกับภาษี',
            'อีเมลสำหรับรับใบกำกับภาษี',
            'ชื่อผู้ใช้ (ผู้ซื้อ)',
            'จำนวนเงินทั้งหมด',
            'วันที่ทำการสั่งซื้อ',
            'โค้ดส่วนลดชำระโดย Shopee',
            'รายละเอียดที่อยู่',
            'ประเภทสาขา',
            'รหัสประจำสาขา',
            'หมายเหตุจากผู้ซื้อ',
            'บันทึก',
            'ชื่อผู้รับ',
            'หมายเลขโทรศัพท์'
        ]

        if self.order != "":
            print("self.order err?: ", self.order, type(self.order))
            if not self.data_frame[(self.data_frame["หมายเลขคำสั่งซื้อ"] == self.order)].empty:
                # ? self.filter_data จะเป็นการทำComparisionให้เรียบร้อยแล้วคืน DataFrame ที่กรองแล้วทันที --------------------ไวกว่า
                self.filter_data = self.data_frame[(
                    self.data_frame["หมายเลขคำสั่งซื้อ"] == self.order)]
                # ? self.target_row เป็น การหา เอาคอล "หมายเลขคำสั่งซื้อ" ทั้งหมดมาตรวจแล้วคืนค่าเป็น Boolean เท่านั้น ---------ช้ากว่า
                self.target_row = self.data_frame["หมายเลขคำสั่งซื้อ"] == self.order
                # print(
                #     "err?: ", self.data_frame[self.target_row]['สถานะการสั่งซื้อ'])
                self.cus_masked_name = self.data_frame[self.target_row]['ชื่อผู้รับ'].iloc[0]
                self.cus_masked_tel = self.data_frame[self.target_row]['หมายเลขโทรศัพท์'].iloc[0]
                self.order_status = self.data_frame[self.target_row]['สถานะการสั่งซื้อ'].iloc[0]

                # *  ของมีอะไรบ้าง
                # print("ของมีไรบ้าง: ", self.data_frame['ชื่อตัวเลือก'])
                # print("ของมีไรบ้าง: ", self.data_frame['ราคาขายสุทธิ'])
                # print("ของมีไรบ้าง: ", self.data_frame['ส่วนลดจาก Shopee'])
                self.items = self.data_frame[differential_col_data][self.target_row].to_dict(
                    'records')
                # ตัดช่องว่าง
                for row in self.items:
                    row['เลขอ้างอิง SKU (SKU Reference No.)'] = row['เลขอ้างอิง SKU (SKU Reference No.)'].replace(
                        ' ', '')

                self.nondistortedData = self.data_frame[self.target_row][non_differential_col_data].iloc[0].to_dict(
                )
                print('self.nondistortedData', self.nondistortedData)
                self.update_log(f"สินค้าที่มี")

                for row in self.items:
                    print("ตัวเลือก", str(row['ชื่อตัวเลือก']))
                    option = ""
                    if str(row['ชื่อตัวเลือก']) != "nan":
                        option = str(row['ชื่อตัวเลือก'])
                    self.update_log(f"SKU: {str(row['เลขอ้างอิง SKU (SKU Reference No.)'])} ชื่อสินค้า: {
                                    option} {str(row['ชื่อสินค้า'])} ")
                    self.update_log(f"ราคาขาย: {float(row['ราคาขาย']):,.2f} จำนวน: {int(row['จำนวน'])} ราคาขายสุทธิ: {
                                    float(row['ราคาขายสุทธิ']):,.2f} ส่วนลดจาก Shopee: {float(row['ส่วนลดจาก Shopee']):,.2f}")

                # * update list รายการสินค้า ช่องที่เลียนแบบ mimic list item like shopee ///////////////////////////////////////////////////////////////////////////////////////////////////////////////////
                self.widgets_no_col_lst = []
                self.widgets_product_col_lst = []
                self.widgets_prc_unit_lst = []
                self.widgets_qty_lst = []
                self.widgets_total_prc_lst = []
                self.widgets_total_rebt_prc_lst = []
                self.widgets_demonic_cp_btn_lst = []
                self.all_cols = [self.widgets_no_col_lst, self.widgets_product_col_lst, self.widgets_prc_unit_lst,
                                 self.widgets_qty_lst, self.widgets_total_prc_lst, self.widgets_total_rebt_prc_lst, self.widgets_demonic_cp_btn_lst]
                self.idx = 0
                self.mimic_list_item_states = []
                for row in self.items:
                    self.no_col_value_widget = Entry(
                        self.mp_products_list_frame, width=int(self.cols_width[0]))
                    self.no_col_value_widget.insert(0, self.idx+1)
                    self.widgets_no_col_lst.append(self.no_col_value_widget)
                    self.idx += 1

                    self.product_col_name_value_widget = Entry(
                        self.mp_products_list_frame, width=int(self.cols_width[1]))
                    self.product_col_name_value_widget.insert(0, f"{str(row['เลขอ้างอิง SKU (SKU Reference No.)'])}{
                                                              ' : ' + str(row['ชื่อตัวเลือก']) if not pd.isna(row['ชื่อตัวเลือก']) else ''} : {str(row['ชื่อสินค้า'])}")
                    self.widgets_product_col_lst.append(
                        self.product_col_name_value_widget)
                    self.mimic_list_item_states.append(
                        f"{str(row['เลขอ้างอิง SKU (SKU Reference No.)'])}")

                    self.price_unit_col_value_widget = Entry(
                        self.mp_products_list_frame, width=int(self.cols_width[2]))
                    self.price_unit_col_value_widget.insert(
                        0, f"{float(row['ราคาขาย']):,.2f}")
                    self.widgets_prc_unit_lst.append(
                        self.price_unit_col_value_widget)

                    self.qty_col_value_widget = Entry(
                        self.mp_products_list_frame, width=int(self.cols_width[3]))
                    self.qty_col_value_widget.insert(0, int(row['จำนวน']))
                    self.widgets_qty_lst.append(self.qty_col_value_widget)

                    self.total_price_col_value_widget = Entry(
                        self.mp_products_list_frame, width=int(self.cols_width[4]))
                    self.total_price_col_value_widget.insert(
                        0, f"{float(row['ราคาขายสุทธิ']):,.2f}")
                    self.widgets_total_prc_lst.append(
                        self.total_price_col_value_widget)

                    self.total_rebate_price_col_value_widget = Entry(
                        self.mp_products_list_frame, width=int(self.cols_width[5]))
                    self.total_rebate_price_col_value_widget.insert(
                        0, f"{float(row['ราคาขายสุทธิ'])+float(row['ส่วนลดจาก Shopee']):,.2f}")
                    self.widgets_total_rebt_prc_lst.append(
                        self.total_rebate_price_col_value_widget)

                    # # * ปุ่ม CP นรกใช้ไม่ได้เก็บไว้พิจารณา
                    # self.demonic_cp_btn = Button(self.mp_products_list_frame, text="xxx", bg="#969696", command=self.search_order, width=10)
                    # self.widgets_demonic_cp_btn_lst.append(self.demonic_cp_btn)

                # print("none ได้ไง:", self.widgets_no_col_lst)
                # print("ไม่สามารถ grid: ", self.all_cols)

                for col_idx, col_list in enumerate(self.all_cols):
                    for idxrow, col in enumerate(col_list):
                        col.grid(
                            row=idxrow+1, column=self.cols_location[col_idx], columnspan=self.colspan_amount[col_idx])
                        col.configure(state="readonly")

                # self.row_header_maker(self.items)

                # * ชื่อที่ต้องออกใบกำกับ
                try:
                    self.cus_name.set(self.translator(re.sub(
                        r'\s{2,}', " ", self.nondistortedData['ชื่อ'].strip().replace('\u200b', ''))))
                except:
                    # * ถ้าชื่อมันว่างมันจะ strip()
                    self.cus_name.set(
                        re.sub(
                            r"[\(\)]",
                            "",
                            self.nondistortedData['ชื่อผู้ใช้ (ผู้ซื้อ)']+" "+self.cus_masked_name+" "+self.cus_masked_tel)
                    )

                self.cus_name_simplifyer(self.cus_name.get())

                # * ประเภทใบกำกับภาษี
                # * เราดูว่าขอใบกำกับหรือไม่ จากที่ว่า 1)มีเลขผู้เสียภาษี 2)มี branch_type
                # * เลือก Column และ row ที่เฉพาะเจาะจง มาแสดงผล โดยการใช้ ['ชื่อคอลั่ม'].iloc[0]
                self.branch_type = str(self.nondistortedData['ประเภทสาขา'])
                print("รหัสประจำสาขา= ",
                      self.data_frame[self.target_row]['รหัสประจำสาขา'].iloc[0])
                branch = self.find_branch(
                    str(self.nondistortedData['รหัสประจำสาขา']))
                self.tax_branch_num.set(branch)

                print("self.data_frame[self.target_row]['หมายเลขประจำตัวผู้เสียภาษี'] กลายเป็น boolจริงเหรอ",
                      self.data_frame[self.target_row]['หมายเลขประจำตัวผู้เสียภาษี'])
                print(
                    "self.nondistortedData['หมายเลขประจำตัวผู้เสียภาษี'] พัง")
                print(bool(pd.isna(
                    self.data_frame[self.target_row]['หมายเลขประจำตัวผู้เสียภาษี'].iloc[0])))
                print(
                    pd.isna(self.data_frame[self.target_row]['หมายเลขประจำตัวผู้เสียภาษี'].iloc[0]))
                print(
                    pd.isna(self.data_frame[self.target_row]['หมายเลขประจำตัวผู้เสียภาษี']))
                print("ค่าจาก DFเพียวๆ: ",
                      self.data_frame[self.target_row]['หมายเลขประจำตัวผู้เสียภาษี'].iloc[0])
                print("ดูtype: ", type(
                    self.data_frame[self.target_row]['หมายเลขประจำตัวผู้เสียภาษี'].iloc[0]))

                # * ถ้า col ['หมายเลขประจำตัวผู้เสียภาษี'] ไม่ใช่ nan จะเก็บค่าลงใน tax_num_only

                if self.marketplace_target.get() == 'SHOPEE':
                    if not pd.isna(self.data_frame[self.target_row]['หมายเลขประจำตัวผู้เสียภาษี'].iloc[0]):
                        tax_num_only = re.sub(r'\D', '', str(
                            self.nondistortedData['หมายเลขประจำตัวผู้เสียภาษี']))
                    else:
                        tax_num_only = "ไม่มีเลข"

                elif self.marketplace_target.get() == 'LAZADA':
                    if self.data_frame[self.target_row]['หมายเลขประจำตัวผู้เสียภาษี'].iloc[0] != "":
                        tax_num_only = re.sub(
                            r'\D', '', str(self.nondistortedData['หมายเลขประจำตัวผู้เสียภาษี']))
                    else:
                        tax_num_only = "ไม่มีเลข"

                # ถ้าเลขใบกำกับเป็น nan หรือ tax_num_only ไม่มีค่า
                if tax_num_only == "ไม่มีเลข":
                    self.tax_bool.set(False)
                    self.is_tax.set("ไม่ขอใบกำกับ")
                    self.display_is_tax.config(
                        background="#6ec7ff", foreground="#000", font='Chiller 10 normal')
                    self.tax_num.set("")
                elif tax_num_only != "ไม่มีเลข" and len(tax_num_only) != 13:
                    if len(tax_num_only) > 13:
                        self.tax_bool.set(False)
                        self.is_tax.set("ขอ//เลขเกิน")
                    elif len(tax_num_only) < 13:
                        self.tax_bool.set(False)
                        self.is_tax.set("ขอ//เลขไม่ครบ")

                    self.display_is_tax.config(
                        background="#8502d1", foreground="#FFF", font='Chiller 10 normal'
                    )
                    self.tax_num.set(tax_num_only)

                else:
                    if "สำนักงานใหญ่" in self.branch_type:
                        self.tax_bool.set(True)
                        self.is_tax.set("ขอใบกำกับ สนงใหญ่")
                        self.display_is_tax.config(
                            background="#ff0000", foreground="#FFF", font='Chiller 10 bold'
                        )
                        self.tax_num.set(tax_num_only)
                    elif self.branch_type == "สาขาย่อย" and (not pd.isna(self.data_frame[self.target_row]['รหัสประจำสาขา'].iloc[0])):
                        self.tax_bool.set(True)
                        self.is_tax.set("ขอใบกำกับ สาขาย่อย")
                        self.display_is_tax.config(
                            background="#ff0055", foreground="#FFF", font='Chiller 10 bold'
                        )
                        self.tax_num.set(tax_num_only)
                    else:
                        self.tax_bool.set(True)
                        self.is_tax.set("ไม่ขอแต่มีเลข")
                        self.display_is_tax.config(
                            background="#ff9e36", foreground="#FFF", font='Chiller 12 bold'
                        )
                        self.tax_num.set(tax_num_only)

                if self.tax_bool.get() == True and len(tax_num_only) == 13:
                    pass

                # * ส่วนสำหรับการแสดงผล UI ------------------------------------------------------
                # self.address = self.filter_data.iat[0, 59]
                self.address = self.nondistortedData['รายละเอียดที่อยู่']
                self.cus_remark: str = str(
                    self.nondistortedData['หมายเหตุจากผู้ซื้อ'])
                self.order_note: str = str(self.nondistortedData['บันทึก'])
                self.cus_email.set(
                    str(self.nondistortedData['อีเมลสำหรับรับใบกำกับภาษี']))

                print("ตรวจหมายเหตุ: ", self.cus_remark)
                print("ตรวจบันทึก: ", self.order_note,
                      "type: ", type(self.order_note))

                # * ดึงบันทึกลูกค้า SHOPEE
                if self.marketplace_target.get() == 'SHOPEE':
                    try:
                        self.note_extractor()
                    except EXCEPTION as err:
                        print("Cannot Extract Note: ", err)
                    # self.cus_name.set()
                    # self.cus_name_simplifyer(self.name_match.group())
                elif self.marketplace_target.get() == 'LAZADA':
                    pass

                # * เอาที่อยู่มาโชว์ ใน UI
                # ? แบบที่1 แบ่ง Channel
                # if self.marketplace_target.get() == "SHOPEE":
                #     self.cleaned_address = f"""{self.get_pure_address(self.clean_address(self.address))} {self.nondistortedData['แขวง/ตำบล']} {
                #     self.nondistortedData['เขต/อำเภอ.1']} {self.nondistortedData['จังหวัด.1']} {self.nondistortedData['รหัสไปรษณีย์.1']}"""
                # else:
                #     self.cleaned_address = f"""{self.get_pure_address(self.clean_address(self.address))}"""
                # ? แบบที่2 ไม่แบ่ง Channel
                if self.marketplace_target.get() == 'LAZADA':
                    self.cleaned_address = f"""{self.get_pure_address(self.clean_address(self.address))} {self.nondistortedData['แขวง/ตำบล']} {
                        self.nondistortedData['เขต/อำเภอ.1']} {self.nondistortedData['จังหวัด.1']} {self.nondistortedData['รหัสไปรษณีย์.1']}"""

                    if "กรุงเทพ" in self.cleaned_address:
                        self.cleaned_address = self.cleaned_address.replace(
                            "จังหวัด", '')
                    self.search_result = {
                        "status": self.order_status,
                        "is_tax": self.tax_bool.get(),
                        "address": self.cleaned_address,
                        "details": self.nondistortedData,
                        "items": self.items
                    }

                if self.marketplace_target.get() == 'SHOPEE':
                    self.cleaned_address = ""
                    # * ถ้าขอใบกำกับค่อยใส่ ถ้าไม่ ก็ "" ไป
                    if self.tax_bool.get():
                        self.cleaned_address = f"""{self.get_pure_address(self.clean_address(self.address))} {self.nondistortedData['แขวง/ตำบล']} {
                            self.nondistortedData['เขต/อำเภอ.1']} {self.nondistortedData['จังหวัด.1']} {self.nondistortedData['รหัสไปรษณีย์.1']}"""

                    if "กรุงเทพ" in self.cleaned_address:
                        self.cleaned_address = self.cleaned_address.replace(
                            "จังหวัด", '')
                    # print("Addressที่คลีนแล้ว: ", self.cleaned_address)
                    self.search_result = {"status": self.order_status,
                                          "is_tax": self.tax_bool.get(), "address": self.cleaned_address, "details": self.nondistortedData, "items": self.items}

                # * สร้างสูตรสำหรับสร้าง input gui
                print("มีไอเทมไรบ้าง", self.search_result['items'])
                self.input_formula = []
                for item in self.search_result['items']:
                    amount = int(item['จำนวน'])
                    sku = str(item['เลขอ้างอิง SKU (SKU Reference No.)'])
                    result = {'sku': sku, 'qty': amount}
                    self.input_formula.append(result)
                    print("จำนวน", int(item['จำนวน']),
                          type(int(item['จำนวน'])))
                print("สูตรสร้าง input", self.input_formula)
                for idx, item in enumerate(self.input_formula):
                    print("รายการที่ ", idx+1, item['sku'])
                    for idx in range(item['qty']):
                        print("สร้างinputอันที่ ", idx+1)

                self.cus_account_name.set(re.sub(
                    r'[^\x00-\x25\x27-\x7F\wA-Zก-๙|/]+', '', self.nondistortedData['ชื่อผู้ใช้ (ผู้ซื้อ)']))
                self.cus_account_name.set(self.cus_account_name.get().strip())
                print("self.cus_account_name: ", self.cus_account_name.get())

                # * update display text ใน gui
                # * เลือกว่าจะใช้ที่อยู่ แบบรายcol หรือ แบบสำเร็จ ไปอัพเดทและแสดงผลที่อยู่ใน gui โดยอัพเดท the gui ด้วย method update_gui_address
                # * การจะเลือกรายcol ได้ต้องชัวร์ว่า col แขวง/ตำบลต้องไม่ใช่ค่าว่าง หรือต้องไม่ Return เป็น "nan"
                try:
                    if not str(self.nondistortedData['แขวง/ตำบล']) == "nan":
                        print("แขวง/ตำบล ไม่เท่ากับ nan: ",
                              self.nondistortedData['แขวง/ตำบล'])
                        # * Lazada กับ shopee มันแสดงผล address ไม่เหมือนกันเพราะ ตาราง Excel ที่มันให้มา
                        if self.marketplace_target.get() == "LAZADA":
                            self.update_gui(
                                re.sub(r'\s{2,}',
                                       " ",
                                       f"""{self.address} {self.nondistortedData['แขวง/ตำบล']} {self.nondistortedData['เขต/อำเภอ.1']} {self.nondistortedData['จังหวัด.1']} {self.nondistortedData['รหัสไปรษณีย์.1']}""".replace('\u200b', '')).strip(),
                                self.display_cus_address
                            )
                        else:
                            print("update gui address else")
                            self.update_gui(
                                re.sub(r'\s{2,}', " ", self.cleaned_address.replace(
                                    '\u200b', '')).strip(),
                                self.display_cus_address
                            )
                    else:
                        print("ถ้ามี nan")
                        self.update_gui(
                            re.sub(
                                r'\s{2,}',
                                " ",
                                self.nondistortedData['ที่อยู่สำหรับออกใบกำกับภาษีแบบเต็มรูป'].strip().replace(
                                    '\u200b', ''
                                )
                            ),
                            self.display_cus_address
                        )

                except Exception as err:
                    print("Cannot Update Address", err)
                    self.update_gui('-', self.display_cus_address)

                self.update_gui_remark()
                self.update_gui_note()

                # * เก็บค่ารายละเอียดที่อยู่
                if self.tax_bool.get():
                    self.cus_province.set(
                        self.nondistortedData['จังหวัด.1'].strip())
                    self.cus_district.set(
                        self.nondistortedData['เขต/อำเภอ.1'].strip())
                if self.cus_sub_district != "":
                    self.cus_sub_district.set(
                        self.nondistortedData['แขวง/ตำบล'])
                else:
                    self.cus_sub_district.set('')
                print("self.nondistortedData['หมายเลขโทรศัพท์สำหรับออกใบกำกับภาษี']: ",
                      self.nondistortedData['หมายเลขโทรศัพท์สำหรับออกใบกำกับภาษี'])
                print("self.nondistortedData['หมายเลขโทรศัพท์สำหรับออกใบกำกับภาษี'] bool?: ",
                      pd.isna(self.nondistortedData['หมายเลขโทรศัพท์สำหรับออกใบกำกับภาษี']))

                if not str(self.nondistortedData['หมายเลขโทรศัพท์สำหรับออกใบกำกับภาษี']) == "nan":
                    print("มีเบอร์โทร")
                    tel_for_set = self.cus_tel_fixer(
                        self.nondistortedData['หมายเลขโทรศัพท์สำหรับออกใบกำกับภาษี'])
                    self.cus_tel.set(tel_for_set)
                else:
                    print("ไม่มีเบอร์โทร")
                    self.cus_tel.set("1")

                self.cus_ship_cost.set(
                    self.nondistortedData['ค่าจัดส่งที่ชำระโดยผู้ซื้อ'])
                self.cus_seller_voucher.set(abs(
                    float(self.nondistortedData['โค้ดส่วนลดชำระโดยผู้ขาย'])))
                self.cus_purchase_time.set(
                    self.nondistortedData['วันที่ทำการสั่งซื้อ'])

                self.net_prices_list = []
                for item in self.items:
                    net_price = item['ราคาขายสุทธิ'] + item['ส่วนลดจาก Shopee']
                    self.net_prices_list.append(net_price)
                # print(f"ขอใบกำกับไหม? {result['is_tax'].get()}")
                # print(f"ที่อยู่: {result['address']}")
                # print("self.bot.get_tabs() ต้องถูกใช้ที่นี่")
                # self.update_log(f"ขอใบกำกับไหม? {result['is_tax'].get()}")
                # self.update_log(f"ที่อยู่: {result['address']}")
                # self.update_log("self.bot.get_tabs() ต้องถูกใช้ที่นี่")
                # loopสินค้า

                self.sum_price = sum(self.net_prices_list)
                self.show_products(self.items)
                print("จำนวนเงิน", self.f(
                    self.nondistortedData['จำนวนเงินทั้งหมด']))
                print('สินค้ารวมค่าส่ง: ', self.f(
                    self.nondistortedData['จำนวนเงินทั้งหมด'] + float(self.cus_ship_cost.get())))
                self.update_log(f"เวลาที่สั่ง: {self.cus_purchase_time.get()}")
                self.update_log(
                    f"ค่าขนส่ง: {self.f(self.cus_ship_cost.get())}")
                self.update_log(f"ราคาที่ต้องยิงทั้งหมด+ค่าส่ง: {self.f(
                    float(self.sum_price) + float(self.cus_ship_cost.get()))}")

                self.update_log(f" ")
                self.update_log(f"-↓↓↓↓↓↓-หน้าสุดท้าย-↓↓↓↓↓↓-")
                self.update_log(
                    f"seller voucher: -{self.f(self.cus_seller_voucher.get())}")

                # * จากปัญหาข้อที่ 37 // การอัพเดท LOG เนื่องจาก LAZ กับ Shopee มีเงื่อนไข การใส่ค่าขนส่งในการออกบิลไม่เหมือนกัน SHOPEE ใส่หมด แต่ LAZ ใส่เป็นบาง ORDER ขึ้นอยู่กับว่า ลูกค้า จะ inbox มาขอให้ใส่หรือไม่
                if self.marketplace_target.get() == "SHOPEE":
                    self.update_log(f"สินค้ารวมค่าส่ง หักseller: {self.f(
                        (self.sum_price+self.cus_ship_cost.get())-self.cus_seller_voucher.get())}")
                elif self.marketplace_target.get() == "LAZADA":
                    self.update_log(f"สินค้าเฉยๆ หักseller: {self.f(
                        (self.sum_price)-self.cus_seller_voucher.get())}")
                    self.update_log(f"---------------------------------")
                    self.update_log(f"สินค้ารวมค่าส่ง หักseller: {self.f(
                        (self.sum_price+self.cus_ship_cost.get())-self.cus_seller_voucher.get())}")

            else:
                print(f"Order ที่ยิงมา {
                      self.cus_order.get()} ไม่สามารถหาใน Export File ได้")
                print(
                    "อาจเกิดจาก เลข Order ที่กรอกเข้ามาผิดพลาด หรือไม่ก็ ไฟล์เก่าเกินไป")
                print("ถ้าไฟล์เก่าแนะนำให้ไป Export File มาใหม่ จาก Link ที่ให้ด้านล่าง")
                print("https://seller.shopee.co.th/portal/sale/shipment?type=toship")

                self.update_log(f"Order ที่ยิงมา {
                                self.cus_order.get()} ไม่สามารถหาใน Export File ได้")
                self.update_log(
                    "อาจเกิดจาก เลข Order ที่กรอกเข้ามาผิดพลาด หรือถ้า Order ไม่ผิด ก็แปลว่าไฟล์ไม่มีข้อมูล")
                self.update_log(
                    "ถ้าไฟล์เก่าแนะนำให้ไป Export File มาใหม่ จาก Link ที่ให้ด้านล่าง")
                self.update_log(
                    "https://seller.shopee.co.th/portal/sale/shipment?type=toship")
                self.reset_all_display()
                logger.info(
                    f"Order: {self.search_query} Not found in the shopee's Export File")

        else:
            self.reset_all_display()

        self.on_complete.set()

    def cusNameFixer5(self, name, account_name=":"):
        is_found = re.search(r"\[.*\]|\(.*\)|\{.*\}", name)
        name = re.sub(r"\[.*\]|\(.*\)|\{.*\}", '',
                      name).strip() if is_found else name.strip()
        # เช็คว่าถ้ามองชื่อเป็น list มันจะแบ่งได้กี่ส่วน
        name += " "+account_name if len(name.split()) == 1 else ""
        print("name:", name)
        return name

    def tax_name_standardizer(self, name):
        name_edited = name.replace('\u200b', '')
        name_edited = name_edited.strip()
        # name_edited = name_edited.replace(
        #     "สำนักงานใหญ่", "").replace("(สำนักงานใหญ่)", "")
        # print("name_editedทำไมมันเหมือนเดิมวะ", name_edited)

        if name_edited.startswith("หจก") or name_edited.startswith("ห้างหุ้นส่วนจำกัด") or name_edited.startswith("ห."):
            print("เงื่อนไขชื่อใบกำกับใน if", name_edited)
            name_edited = name_edited.replace("หจก.", "").replace(
                "ห้างหุ้นส่วนจำกัด", "").replace("ห.", "").strip()
            name_edited = f"""ห้างหุ้นส่วนจำกัด {name_edited}"""

        elif name_edited.startswith("บจก") or (name_edited.startswith("บริษัท") and "จำกัด" in name_edited) or name_edited.startswith("บ."):
            print("เงื่อนไขชื่อใบกำกับใน elif", name_edited)
            name_edited = name_edited.replace("บจก.", "").replace("บริษัท", "").replace(
                "จำกัด", "").replace("บ.", "").replace("จก.", "").strip()
            name_edited = f"""บริษัท {name_edited} จำกัด"""

        # * > ลบประเภทสาขาแล้วส่งค่าออก ค่าที่ออกจะไม่มี สำนักงาน สาขา เดี๋ยวไป add ทีหลังในขั้นตอน add ชื่อ (ส่วนท้ายของ code)
        # * >> สร้าง patterns ก่อน
        head_office_patterns = [
            r'\(สำนักงานใหญ่\)', r'สำนักงานใหญ่',
            r'\(สํานักงานใหญ่\)', r'สํานักงานใหญ่',
            r'\(สนญ\.\)', r'\(สนญ\)', r'สนญ\.', r'สนญ',
        ]

        # * >> ใช้ for-loop ดูว่า มีสัก pattern ไหม ที่อยู่ในชื่อลูกค้า แล้ว any จะจับค่า boolean ที่ได้ ว่ารอบไหนของ for-loop คืนค่า True บ้าง
        if any(pattern in name_edited for pattern in head_office_patterns):
            # * re.sub(pattern, คำที่เอามาแทน, ข้อความที่เป็นกรรม(ถูกกระทำ))
            # * r'|'.join(head_office_patterns) เป็นการ เอาคำทั้งหมดใน head_office_patterns มาต่อกันด้วยเครื่องหมาย "|" จะได้ r'x|y|z' ประมาณนี้
            name_edited = re.sub(
                r'|'.join(head_office_patterns), '', name_edited).strip()
        elif '(สาขา' in name_edited or 'สาขา' in name_edited:
            name_edited = re.sub(r'\(สาขา.*\)', '', name_edited)
            name_edited = re.sub(r'สาขา\d*', '', name_edited)

        name_edited = re.sub(r"\s{2,}", ' ', name_edited)
        return name_edited

    def on_thread_done(self):
        print("on_thread_done start")
        self.get_tabs_stat = self.get_tabs_thread.is_alive()
        self.search_thread_stat = self.search_thread.is_alive()
        print("ก่อนifเช็คตัวรัน tab", self.get_tabs_stat)
        print("ก่อนifเช็คตัวรัน excel", self.search_thread_stat)
        if self.get_tabs_thread.is_alive():
            print("self.get_tabs_thread.is_alive()")
            # self.search_complete.set()
            self.get_tabs_thread.join()

        print("Thread is done คงเหงาแย่")

        self.get_tabs_stat = self.get_tabs_thread.is_alive()
        self.search_thread_stat = self.search_thread.is_alive()
        print("หลังifเช็คตัวรัน tab", self.get_tabs_stat)
        print("หลังifเช็คตัวรัน excel", self.search_thread_stat)

        print("Thread is done")
        if self.get_tabs_stat == False and self.search_thread_stat == False:
            self.display_bot_status_label.config(
                text=f"Bot Status: ˶ᵔ ᵕ ᵔ˶ จบการทำงาน", bg="#d9f2ff", fg="#000")
            print("Bot Status: ˶ᵔ ᵕ ᵔ˶ จบการทำงาน (ตัวบน)")

        if self.get_tabs_thread.is_alive():
            print("มีthreadใหม่มาต่อ")
            self.display_bot_status_label.config(
                text=f"Bot Status: ᕦʕ •ᴥ•ʔᕤ กำลังทำงาน", bg="#cf1313", fg="#ffffff")

    def check_threads(self, shorter_thread_cycle, longer_thread_cycle, callback=None):
        # print(self.is_bot_running.get())
        # * เป็นการเช็ค thread ไปเรื่อยๆจนกว่า thread ทั้งคู่จะดับไป หาก Thread ใด Thread หนึ่ง ทำงานอยู่ ให้เช็คตัวเองอีกรอบ ภายในเวลา 100 millisec
        if (shorter_thread_cycle.is_alive() or longer_thread_cycle.is_alive()):
            # * after(เวลาmillisec, callbackfunction)
            self.root.after(1000, lambda: self.check_threads(
                shorter_thread_cycle, longer_thread_cycle, callback))

            # * เอาไว้แสดงสถานะของ bot gui ว่าทำงานอยู่หรือไม่
            if self.is_gui_busy.get() == True:
                self.display_bot_status_label.config(
                    text=f"Bot Status: ᕦʕ •ᴥ•ʔᕤ กำลังทำงาน", bg="#cf1313", fg="#ffffff")
            elif self.is_gui_busy.get() == False:
                self.display_bot_status_label.config(
                    text=f"Bot Status: Your Turn", bg="#21ff29", fg="#000")
        else:
            # * เมื่อ Thread ทั้งสองไม่ alive จะทำการรวม thread ย่อย เข้ากับ thread หลัก แล้วเรียกใช้ callback ถ้าหากมี callback มาด้วยน่ะนะ callbackนี้จะรับ operation_startเข้ามาให้ทำงานอีกรอบ
            shorter_thread_cycle.join()
            longer_thread_cycle.join()
            print("shorter_thread_cycle is alive?: ",
                  shorter_thread_cycle.is_alive())
            print("longer_thread_cycle is alive?: ",
                  longer_thread_cycle.is_alive())
            self.display_bot_status_label.config(
                text=f"Bot Status: ˶ᵔ ᵕ ᵔ˶ จบการทำงาน", bg="#d9f2ff", fg="#000")
            print("Bot Status: ˶ᵔ ᵕ ᵔ˶ จบการทำงาน (ตัวล่าง)")

            if callback:
                callback()

    def search_order(self, accel_order=None, callback=None):
        self.is_bot_running.set(True)
        self.autofinal = False
        # * ลบ result products list เก่า
        for widget in self.mp_products_list_frame.winfo_children()[6:]:
            widget.destroy()

        if self.is_accel_mode.get():
            self.search_query = accel_order
        else:
            self.search_query = self.entered_order.get()

        print("search() ทำงานและได้ผลลัพธ์: ", self.search_query)
        self.entered_order.set("")
        if self.search_query != "":
            self.report_log.config(state=NORMAL)
            self.report_log.delete("1.0", "end")
            # self.report_log.insert(END, self.search_query + "\n")
            self.report_log.config(state=DISABLED)
        else:
            self.report_log.config(state=NORMAL)
            self.report_log.delete("1.0", "end")
            self.report_log.config(state=DISABLED)

        self.search_complete = threading.Event()
        self.search_complete.set()

        # * สร้าง Thread
        self.shorter_thread_cycle = threading.Thread(target=self.bot.get_tabs)
        self.longer_thread_cycle = threading.Thread(
            target=lambda: self.order_search(self.search_query, self.search_complete))
        print("Thread Name: ", self.longer_thread_cycle.name)

        # * สั่ง Thread ให้เริ่มทำงาน
        self.shorter_thread_cycle.start()
        self.longer_thread_cycle.start()
        # self.shorter_thread_cycle.join()
        # self.longer_thread_cycle.join()

        # * ตรวจสอบว่า Thread ทั้งสองยังทำงานอยู่หรือไม่
        self.check_threads(self.shorter_thread_cycle,
                           self.longer_thread_cycle, callback)
        self.display_bot_status_label.config(
            text=f"Bot Status: ᕦʕ •ᴥ•ʔᕤ กำลังทำงาน", bg="#cf1313", fg="#ffffff")

    # * method accel_search() จะทำงานจากการกดปุ่ม
    def accel_search(self):
        self.is_accel_mode_activated.set(True)
        self.accel_orders_len = len(self.accel_orders_list)

        # * สร้าง recursive function
        def start_next_cycle(count):
            self.accel_df_state = pd.read_excel(self.accel_file_dir, dtype=str)
            if count < self.accel_orders_len:
                if self.is_accel_mode_activated.get():
                    self.search_order(
                        self.accel_orders_list[count], lambda: start_next_cycle(count+1))
                else:
                    raise ValueError("Accel mode has been destroyed")
            else:
                pass

        # * search รอบแรกใช้ตรงนี้
        self.search_order(
            self.accel_orders_list[0], lambda: start_next_cycle(1))

    def stop_operation(self):
        # self.is_accel_mode_activated.set(False) ตัวแปรนี้การการhandleที่ทำให้บัค แต่มันทำงานดี
        self.is_bot_running.set(False)

    def convert_text(self, text):
        result = []
        text = text.replace(" ", "")
        elements = text.split("+")

        for element in elements:
            prefix = element.split("-")[0]
            code = element.split("-")[1]
            code = code.zfill(6)
            result.append(prefix + "-00" + code)

        return result

    def demonic_cp(self):
        self.item_no = int(self.entered_item_no.get())-1
        self.cp_no = int(self.entered_cp_no.get())
        self.demonic_ordered_items_list = self.convert_text(
            self.items[self.item_no]['เลขอ้างอิง SKU (SKU Reference No.)'])
        print(self.demonic_ordered_items_list)
        print(self.cp_no)

        self.driver.switch_to.window(self.merged_dict['SMCO :: เปิดการขาย'])
        self.cp_no = 1
        # *>  element location
        # * >> ปุ่มคูปองด้านนอก ที่ตำแหน่ง [-4] จะเป็นตัวแยก element หรือ ตัวบอกตำแหน่งของ element ว่าเป็นลำดับที่เท่าไหร่ อย่างตัวอย่างนี้เป็น อันที่1
        cp_btn_xpath = '/html/body/div[1]/div[2]/div[2]/div[2]/div[1]/div[2]/div[1]/div/div[2]/div[3]/div[1]/a'
        green_agree_btn_xpath = '/html/body/div[1]/div[2]/div[9]/div/div[1]/div[2]/a'

        items_list = self.driver.find_elements(
            By.CSS_SELECTOR, '.col-sm-12.panel.panel-default.ng-scope')
        cp_list = self.driver.find_elements(
            By.XPATH, '/html/body/div[1]/div[2]/div[9]/div/div[2]/div[3]')

        # print("items_list", items_list)
        for idx, item in enumerate(self.demonic_ordered_items_list):
            print("มาถึงนี่ไหม")
            for idx2, div in enumerate(items_list):
                print("รอบ", idx2)
                # time.sleep(0.55)
                try:
                    is_found = div.text.find(item)
                except:
                    pass
                li_position = idx+1
                if is_found != -1:
                    print("เจอที่ ", li_position)
                    print("is_found: ", is_found)
                    cp_btn_xpath = f'/html/body/div[1]/div[2]/div[2]/div[2]/div[1]/div[2]/div[{
                        li_position}]/div/div[2]/div[3]/div[1]/a'
                    self.driver.find_element(By.XPATH, cp_btn_xpath).click()

                    # * เลือก cp เป้าหมาย
                    selected_btn = f'/html/body/div[1]/div[2]/div[9]/div/div[2]/div[3]/div[{
                        self.cp_no}]/div[1]/button'
                    self.driver.find_element(By.XPATH, selected_btn).click()

                    self.driver.find_element(
                        By.XPATH, green_agree_btn_xpath).click()
                    # time.sleep(0.55)
                    continue
                    # print(div.text)
                else:
                    print("ไม่เจอ", item, "นะ")
                    pass

    def open_subwindow(self):
        self.data_source_selector.create_subwindow()

    def get_dataframe(self):
        print("เรียกหา dataframe")


# สำหรับเลือกที่มาของแหล่งข้อมูล
class DataSourceSelector:
    def __init__(self, parent, app):
        self.parent = parent
        self.app = app
        self.create_subwindow()

    def create_subwindow(self):
        self.subwindow = Toplevel(self.parent)
        self.subwindow.transient(self.parent)
        self.subwindow.geometry("250x75+650+400")
        self.subwindow.title("Data Source")
        self.subwindow.grab_set()
        self.subwindow.resizable(False, False)

        self.api_btn = Button(self.subwindow, text="API",
                              command=self.select_api)
        self.api_btn.pack(side='left', expand=TRUE, fill="both")
        self.excel_btn = Button(
            self.subwindow, text="Excel", command=self.select_excel)
        self.excel_btn.pack(side='left', expand=TRUE, fill="both")

        self.subwindow.protocol("WM_DELETE_WINDOW", self.on_close)

    def select_api(self):
        self.app.result = "API"
        print("Select API")
        self.subwindow.destroy()

    def select_excel(self):
        self.app.result = "Excel"
        print("Select Excel")
        self.app.table_location = filedialog.askopenfilename(
            title="Select Shopee order toship file")
        self.app.display_location_result.config(
            text=f"{self.app.table_location.split('/')[-1]}")
        # target should come before get dataframe
        self.app.marketplace_target.set(self.app.define_marketplace())
        result = self.app.marketplace_target.get()
        print("ต้องตีเว็บไหน", result)
        # self.canvas.config(bg=f'{self.bg_by_market_place[self.app.marketplace_target.get()}')
        self.app.entry_frame.config(
            bg=f'{self.app.bg_by_market_place[str(result)]}')
        self.app.marketplace_label.config(
            bg=f'{self.app.bg_by_market_place[str(result)]}')
        # self.import_file_frame.config(
        #     bg=f'{self.bg_by_market_place[self.app.marketplace_target.get()]}')

        self.app.get_data_frame()
        print("Table Location:", self.app.table_location)
        self.subwindow.destroy()
        self.app.update_log("เพิ่มไฟล์แล้ว")

    def on_close(self):
        self.app.marketplace_target.set("")
        self.subwindow.destroy()


class PopUp:
    """
    Class PopUp use for create a pop-up for THE BOT GUI
    Parameters:
        - title (str): Title name of the pop-up.
        - message (str): For display a message in the pop-up.
        - parent (obj): class parent obj.
        - mode (str): มี 2 ทางเลือก "form" สำหรับ submit, "alert" สำหรับ alert
    """

    def __init__(self, title, message, parent, mode):
        self.mode_opt = {"form": "Submit", "alert": "Close"}
        self.mode = mode
        self.parent = parent
        self.title = title
        self.message = message
        self.create_subwindow()

    def delete(self):
        self.subwindow.destroy()

    def create_subwindow(self):
        self.subwindow = Toplevel(self.parent)
        self.subwindow.transient(self.parent)
        self.subwindow.geometry("400x140+650+400")
        self.subwindow.title(f"{self.title}")
        self.subwindow.grab_set()
        self.subwindow.resizable(True, False)

        # * สร้างเฟรม
        self.subwin_frame = Frame(self.subwindow)
        self.subwin_frame.pack(padx=10, pady=10, fill='x', expand=True)

        # * สร้าง Texted widget
        self.id_label = Text(
            self.subwin_frame, font=("bazooka", 9))
        self.id_label.insert(END, f'{self.message}')
        self.id_label.pack(fill=BOTH, expand=True)
        self.id_label.config(state=DISABLED)

        # * Submit Button
        self.submit_btn = Button(
            self.subwin_frame, text=f"{self.mode_opt[self.mode]}", command=self.delete)
        self.submit_btn.pack(fill='x', expand=True)

        # * ยก widget นี้ ขึ้นมาหน้าสุด
        # > กำหนดตำแหน่งเฉยๆ ยังไม่ขยับ ต้องไปสั่งขยับอีกที
        self.subwindow.attributes('-topmost', 1)
        # > ยกมาในตำแหน่งที่กำหนดจาก attribute ที่แล้ว
        self.subwindow.lift()

# * class สำหรับรับ ID PASS


class UserAccount:
    def __init__(self, parent, app):
        self.parent = parent
        self.app = app
        self.create_subwindow()

    def create_subwindow(self):
        self.subwindow = Toplevel(self.parent)
        self.subwindow.transient(self.parent)
        self.subwindow.geometry("250x140+650+400")
        self.subwindow.title("Loginปลอม")
        self.subwindow.grab_set()
        self.subwindow.resizable(False, False)

        # * Event Enter
        self.subwindow.bind(
            "<Return>", lambda event=None: self.submit_btn.invoke())
        self.subwindow.bind("<Key>", _onKeyRelease)

        # * สร้างเฟรม
        self.subwin_frame = Frame(self.subwindow)
        self.subwin_frame.pack(padx=10, pady=10, fill='x', expand=True)

        # * สร้าง widget
        self.id_label = Label(
            self.subwin_frame, text="SMCO ID", font=("bazooka", 9), anchor="w")
        self.id_label.pack(fill='x', expand=True)
        # self.id_input = Entry(self.subwin_frame, textvariable=self.app.user_id,
        #                       validate="key", validatecommand=(self.app.validate_input_variable, '%P'))
        self.id_input = Entry(self.subwin_frame, textvariable=self.app.user_id)
        self.id_input.pack(fill='x', expand=True)
        self.id_input.focus()

        self.pass_label = Label(
            self.subwin_frame, text="SMCO Password", font=("bazooka", 9), anchor="w")
        self.pass_label.pack(fill='x', expand=True)
        # self.pass_input = Entry(
        #     self.subwin_frame, textvariable=self.app.user_pw, show="*", validate="key", validatecommand=(self.app.validate_input_variable, '%P'))
        self.pass_input = Entry(
            self.subwin_frame, textvariable=self.app.user_pw, show="*")
        self.pass_input.pack(fill='x', expand=True)

        # * checkBox
        self.chk_bx_show_pw = Checkbutton(self.subwin_frame, text="Show Pass", font=(
            'bazooka', 9), command=self.show_and_hide)
        self.chk_bx_show_pw.pack()

        # * Submit Button
        self.submit_btn = Button(
            self.subwin_frame, text="Submit", command=self.update_btn)
        self.submit_btn.pack(fill='x', expand=True)

    def login(self):
        cookies = {
            'JSESSIONID': 'EA2AD7582A59949D14642F01ADF23832',
            'locale': 'en_US',
        }

        headers = {
            'Accept': '*/*',
            'Accept-Language': 'en-US,en;q=0.9',
            'Connection': 'keep-alive',
            'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
            # 'Cookie': 'JSESSIONID=EA2AD7582A59949D14642F01ADF23832; locale=en_US',
            'Origin': 'http://192.168.0.11:8080',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36',
            'X-Requested-With': 'XMLHttpRequest',
        }

        data = {
            'locale': 'en_US',
            'redirect': 'http://192.168.0.11:8080/smartcore/',
            'username': [
                f'{self.app.user_id.get()}',
            ],
            'password': [
                f'{self.app.user_pw.get()}',
            ],
            'branch': [
                '',
                '',
            ],
            'storeId': [
                '',
                '',
            ],
        }

        response = session.post(
            'http://192.168.0.11:8080/smartcore/loginssoauthen.htm',
            cookies=cookies,
            headers=headers,
            data=data,
            verify=False,
        )

        result = response.json()
        print("ได้ response ไรมา: ", response)
        print("ได้ result ไรมา: ", result)
        # * ตรวจสอบ response จากการ login
        if result['status'] == "MORE_BRANCH":
            print("Logged in")
            return True
        else:
            print("Incorrect username or password")
            self.login_alert = PopUp(
                "Login Fail!!", "พาสเวิร์ดผิดหรือป่าว~\nถ้าถูกแล้วก็อาจจะเป็นที่ SMCO\nลองเช็ค SMCO ดู", self.parent, "form")
            return False

    def update_btn(self):
        if self.app.user_id.get() and self.app.user_pw.get():

            # is_closable = self.login()
            is_closable = True
            print("ปิดได้ไหม ", is_closable)
            if is_closable:
                self.display_btn_txt = f"""Logged in !! ID : {
                    self.app.user_id.get()}"""
                self.app.display_acc_btn.config(text=self.display_btn_txt)
                self.subwindow.destroy()

            else:
                print("ไม่ติด")
                self.display_btn_txt = "Login"
                # self.subwindow.destroy()
                # return self.display_btn_txt

            if self.app.user_id.get() in self.app.dev_account:

                print("Accel mode approachable")
                if self.app.accel_mode_checkbox.winfo_ismapped():
                    pass
                else:
                    self.app.accel_mode_checkbox.grid(row=0, column=8, padx=5)
            else:
                print("Normal mode", self.app.user_id.get()
                      in self.app.dev_account)
                self.app.accel_mode_checkbox.grid_remove()
                print(self.app.user_id.get())
                print(self.app.dev_account)

            return self.display_btn_txt

    def show_and_hide(self):
        if self.pass_input['show'] == '*':
            self.pass_input['show'] = ''
        else:
            self.pass_input['show'] = '*'


class Bot_POS:
    def __init__(self, parent, app):
        # super().__init__(parent)
        self.parent = parent
        self.app = app
        self.wsh = comclt.Dispatch("WScript.Shell")
        self.setup_chrome()

    def setup_chrome(self):
        self.opt = Options()
        # * ใช้เพื่อเก็บที่อยู่ของไฟล์ที่ถูก execute ด้วย Python ผ่าน command line arguments ในตัวแปร exepath ซึ่ง sys.argv[0] คือชื่อของไฟล์ Python script ที่ถูกเรียกใช้งาน
        exepath = sys.argv[0]

        Dir_path = os.path.dirname(os.path.abspath(exepath))
        self.custom_path = r'D:\\bin\\'

        os.environ["WDM_LOCAL"] = self.custom_path
        # print("มีไรบ้างใน obj Options:", dir(self.opt))
        self.opt.add_experimental_option("debuggerAddress", "localhost:8989")
        # self.opt.add_argument("--disable-popup-blocking")
        # self.opt.add_experimental_option("prefs",{
        #     "download.default_directory" : Download_dir,
        #     "directory_upgrade": True
        # })

        #! อันเก่า
        # self.driver = webdriver.Chrome(
        #     service=Service(r'C:\bin\chromedriver.exe'),
        #     options=self.opt
        # )

        # ?? อันใหม่ทดลอง
        try:
            print("create driver")
            # * error มันจะเกิดแถวนี้
            self.driver = webdriver.Chrome(
                service=Service(r'C:\bin\chromedriver.exe'),
                options=self.opt
            )

            print("driver created")

        except:
            traceback_str = traceback.format_exc()
            print("Cannot Create Driver")
            print(traceback_str)
            chrome_app_utils = ChromeAppUtils()
            chrome_app_version = chrome_app_utils.get_chrome_version()
            print("Chrome version: ", chrome_app_version)

            # * Target directory to store chromedriver
            driver_directory = 'C:/bin'

            # * Create an inst of WebDriverManager
            driver_manager = WebDriverManager(driver_directory)

            # * Call the main method to manage chromdriver
            try:
                driver_manager.main()
                # * check_driver() ใช้ปุ๊บมันจะทำการตรวจและโหลดเลย
                driver_manager.check_driver()
            except Exception as err:

                print('error from driver_manager.main()')
                print(err)
                raise

            self.driver = webdriver.Chrome(
                service=Service(r'C:\bin\chromedriver.exe'),
                options=self.opt
            )

    def convert_text(self, text):
        result = []
        elements = text.split("+")

        for element in elements:
            prefix, code = element.split("-")
            code = code.zfill(6)
            result.append(prefix + "-" + code)

        return result

    def demonic_cp_bot(self, item_no, cp_no):
        self.item_no = int(item_no)-1
        self.cp_no = int(cp_no)
        print("ตอนแรกเปนงี้",
              self.app.items[self.item_no]['เลขอ้างอิง SKU (SKU Reference No.)'])
        self.demonic_ordered_items_list = self.convert_text(
            self.app.items[self.item_no]['เลขอ้างอิง SKU (SKU Reference No.)'])
        print(self.demonic_ordered_items_list)
        print(self.cp_no)

        self.driver.switch_to.window(self.merged_dict['SMCO :: เปิดการขาย'])
        # *>  element location
        # * >> ปุ่มคูปองด้านนอก ที่ตำแหน่ง [-4] จะเป็นตัวแยก element หรือ ตัวบอกตำแหน่งของ element ว่าเป็นลำดับที่เท่าไหร่ อย่างตัวอย่างนี้เป็น อันที่1
        cp_btn_xpath = '/html/body/div[1]/div[2]/div[2]/div[2]/div[1]/div[2]/div[1]/div/div[2]/div[3]/div[1]/a'
        green_agree_btn_xpath = '/html/body/div[1]/div[2]/div[9]/div/div[1]/div[2]/a'

        items_list = self.driver.find_elements(
            By.CSS_SELECTOR, '.col-sm-12.panel.panel-default.ng-scope')
        try:
            # * ก่อน SMCOver 6.3.3
            cp_list = self.driver.find_elements(
                By.XPATH, '/html/body/div[1]/div[2]/div[9]/div/div[2]/div[3]')
        except:
            # * ตั้งแต่ SMCOver 6.3.3
            cp_list = self.driver.find_elements(
                By.XPATH, '/html/body/div[1]/div[2]/div[9]/div/div[2]/div[2]')

        # print("items_list", items_list)
        for idx, item in enumerate(self.demonic_ordered_items_list):
            print("มาถึงนี่ไหม")
            for idx2, div in enumerate(items_list):
                try:
                    print("จำนนวน div ", len(items_list))
                    # print("รอบ", idx2)
                    # time.sleep(0.55)

                    is_found = div.text.find(item)

                    li_position = idx2+1
                    if is_found != -1:
                        print("เจอที่ ", li_position)
                        print("is_found: ", is_found)
                        cp_btn_xpath = f'''/html/body/div[1]/div[2]/div[2]/div[2]/div[1]/div[2]/div[{
                            li_position}]/div/div[2]/div[3]/div[1]/a'''
                        self.driver.find_element(
                            By.XPATH, cp_btn_xpath).click()

                        # * เลือก cp เป้าหมาย
                        try:
                            # * SMCO ให้ ตั้งแต่ v6.3.3
                            selected_btn = f'''/html/body/div[1]/div[2]/div[9]/div/div[2]/div[2]/div[{
                                self.cp_no+1}]/div[1]/button'''
                            self.driver.find_element(
                                By.XPATH, selected_btn).click()
                        except:
                            # * SMCO เก่า
                            selected_btn = f'''/html/body/div[1]/div[2]/div[9]/div/div[2]/div[3]/div[{
                                self.cp_no}]/div[1]/button'''
                            self.driver.find_element(
                                By.XPATH, selected_btn).click()

                        self.driver.find_element(
                            By.XPATH, green_agree_btn_xpath).click()
                        # time.sleep(1)
                        continue
                        # print(div.text)
                    else:
                        print("ไม่เจอ", item, "นะ")
                        pass
                except:
                    pass

    def get_tabs(self):
        try:
            if self.parent.winfo_exists():
                print("รายงานจำนวนtabs")

                self.title_list = []
                # self.title_list_Idx = [] #!เหมือนจะไม่ได้ใช้
                self.value_list = []
                # self.title_dict = {} #!เหมือนจะไม่ได้ใช้

                # * check ว่า self.driver เดิมยังทำงานได้ไหม
                try:
                    # * เช็คก่อนว่า driver ใช้ได้ไหม หรือการเชื่อมต่อ session หลุดไหม
                    self.driver.window_handles
                except:
                    # * driver หลุดก็ออก seesion เก่า
                    try:
                        print(
                            "Quit old driver, not sure if this process is auto or not")
                        self.driver.quit()
                    except:
                        print("No need to quit old driver, no driver found")
                        pass

                    self.driver = webdriver.Chrome(
                        service=Service(r'C:\bin\chromedriver.exe'),
                        options=self.opt
                    )

                for idx, handle in enumerate(self.driver.window_handles):
                    self.driver.switch_to.window(handle)
                    # self.title_list_Idx.append(
                    #     self.driver.title + "["+str(idx)+"]") #!เหมือนจะไม่ได้ใช้
                    self.title_list.append(self.driver.title)

                    self.value_list.append(self.driver.current_window_handle)
                    # self.title_dict.update(
                    #     {self.driver.title: self.driver.current_window_handle}) #!เหมือนจะไม่ได้ใช้

                self.unique_titles = []
                self.counter = {}
                for item in self.title_list:
                    if item in self.counter:
                        self.counter[item] += 1
                        print("counter[item] คือไร: ", self.counter[item])
                        self.unique_titles.append(
                            f"{item}{self.counter[item]-1}")
                    else:
                        self.counter[item] = 1
                        self.unique_titles.append(item)

                # เอาList มารวมกัน
                self.merged_dict = dict(
                    zip(self.unique_titles, self.value_list))
                print("มี tabs ไรบ้าง", self.merged_dict)

                # * เริ่มการทำงาน Operation Start

                if self.app.order != "":

                    logger.info(f"Order: {self.app.order} Start!!")
                    try:
                        self.operation_start()
                    except EXCEPTION as err:
                        logger.info(
                            f"Order: {self.app.order} outer_Exception_Error!! {err}")
                else:
                    self.app.update_log("กรุณากรอก Order ก่อน")
                    self.app.search_complete.set()

        except Exception as err:
            traceback_str = traceback.format_exc()
            print(f"An error occirred: {err}")
            print(traceback_str)
            logger.info(
                f"Order: {self.app.order} get_tabs_outer_Exception_Error!! {err}")

    def enter_cus_name(self, cus_search):
        # * เคลียและกรอกชื่อลูกค้า
        self.driver.find_element(By.XPATH, self.app.cusNameInput).clear()
        self.driver.find_element(
            By.XPATH, self.app.cusNameInput).send_keys(cus_search)

    def add_cusname(self):
        # * ขอใบกำกับป่าว
        if self.app.tax_bool.get():
            print("Tax_needed")
            if self.app.marketplace_target.get() == 'SHOPEE':
                self.addTaxInvCustomer()

            # * กำลังทำ กำลังปรับปรุง ยังไม่เสร็จ การหาลูกค้าของ laz มันมีกรณี excel และ api
            elif self.app.marketplace_target.get() == 'LAZADA':
                self.addTaxInvCustomerLaz()

        else:
            print("no_Tax_needed")
            self.addNormalCustomer(self.cus_search_input)

    # !66 WIP เปลี่ยนวิธีเลือกชื่อลูกค้า เดิมทีคือเลือก // ชิพหายมันเลือกค่าจาก i
    def select_cus_name_from_lis(self, names, cb=""):
        cus_desire_name = self.app.cus_name.get().replace(" ", "")

        # * ทำการคัดเอาเฉพาะชื่อลูกค้าไม่เอารหัส ลง array
        names_no_code = names.copy()
        for i in range(len(names)):
            prog = re.search(r'[^-]-(.*)', names_no_code[i])
            names_no_code[i] = prog.group(1).replace(" ", "")

        for i, name in enumerate(names_no_code):
            print("if ", cus_desire_name, " In ", name)
            if cus_desire_name in name:
                print("ชื่อที่ต้องการ อยู่ใน li")
                while True:
                    try:
                        print("เลือกชื่อลูกค้า", names[i])
                        self.driver.find_element(
                            By.XPATH, f"/html/body/span/span/span[2]/ul/li[{i+1}]").click()
                        break

                    except:
                        print("No customer found")
                        continue
                return
            # * ถ้ามันเจอก็จะ break ไม่เจอค่อย cb
        try:
            if cb:
                print("use callback")
                cb(names)

            # * cb ให้รอบนึงแล้วก็ไม่เจอ แอดใหม่ให้
            # print('ไม่เจอ แอดใหม่ เปลี่ยนชื่อให้ด้วย')
            # self.cus_search_input = self.app.cus_name.get()
            # self.add_cusname()
        except:
            print("cb doesn't works")

        # * มันจะมีกรณีที่ถ้าเลือกลูกค้าได้ในครั้งแรก cb จะไม่ทำงานในส่วนนี้
        try:
            if cb:
                cb(names)
        except:
            print("cb doesn't works")

        # * มันจะมีกรณีที่ถ้าเลือกลูกค้าได้ในครั้งแรก cb จะไม่ทำงานในส่วนนี้

    def printtingPage(self):
        time.sleep(1)
        self.printing_page = self.driver.find_element(By().XPATH, '/html/body')
        self.action01 = ActionChains(
            self.driver).context_click(self.printing_page)
        self.action01.perform()

    def justPressP(self):
        time.sleep(1)
        self.wsh.SendKeys("P")
        time.sleep(1.55)
        self.wsh.SendKeys("{Enter}")
        print("print แล้วโว้ย")
        time.sleep(2)
        # * กดข้างนอกแล้วส่ง event ปุ่ม  ESC จาก KB
        # self.driver.find_element(By.XPATH, '/html/body/div[3]/div/div/div').click()
        # self.wsh.SendKeys("{ESC}")

        # * ใช้ selenium กดปุ่มแดงโดยตรง ปุ่มแดงหน้า print ไม่เหมือน ปุ่มแดงหน้าแรก เพราะห้นปริ้นจะเป็น tag a ส่วนหน้าปกติมันเป็น button ใช้คนละ element แต่อยู๋ตำแหน่งเดียวกัน โคตรปั่น
        self.driver.find_element(By.XPATH, '/html/body/div[1]/div[2]/div[8]/div/div[1]/div/a').click()

        print("กด ปุ่มแดง ออก แล้ว")
        # ถ้าเขียนเป็น cb แล้วมันจะพัง

    def etax_reprint(self, inv_number):
        try:
            # * เก็บหน้าเก่าเพื่อ กลับไปหน้าเดิมก่อน reprint
            prev_window = self.driver.current_window_handle
            # * สลับหน้าไป reprint
            self.driver.switch_to.window(
                self.merged_dict['SMCO :: พิมพ์ใบเสร็จซ้ำ'])
            print("สลับไปหน้าพิม์ใบเสร็จซ้ำ")

        except:
            # * สลับไม่ได้เปิด reprint ใหม่
            print("ไม่มีหน้าให้สลับ เปิดใหม่")
            self.driver.get(
                "http://115.31.167.28:8080/smartcore/smartpos/payment/reprint_invoice.htm?mc=POS2050")
            all_window_handles = self.driver.window_handles
            latest_window_handle = all_window_handles[-1]
            self.driver.switch_to.window(latest_window_handle)
            print("ไม่มีเปิดใหม่")

        # * เริ่มทำการกรอกบิลล่าสุดในหน้า reprint หน้า พิมพ์ใบเสร็จซ้ำ
        try:
            print("Start reprint")
            time.sleep(0.75)
            # * > เปิด dropdownก่อน ไม่งั้นใช้ input ไม่ได้
            self.driver.find_element(
                By().XPATH, '/html/body/div[1]/div[2]/div[1]/div[2]/div/div[1]/div[1]/div/span/span[1]/span/span[1]').click()
            self.driver.find_element(
                By().XPATH, '/html/body/span/span/span[1]/input').clear()
            self.driver.find_element(
                By().XPATH, '/html/body/span/span/span[1]/input').send_keys(inv_number)
            self.driver.find_element(
                By().XPATH, '/html/body/div[1]/div[2]/div[1]/div[2]/div/div[2]/div[2]/div/textarea').clear()
            self.driver.find_element(
                By().XPATH, '/html/body/div[1]/div[2]/div[1]/div[2]/div/div[2]/div[2]/div/textarea').send_keys("Etax")
            # while True:
            #     # if self.driver.find_element(By.XPATH, '/html/body/span/span/span[2]/ul/li').text == "Searching...":
            #     #     continue
            #     time.sleep(0.75)
            #     if self.driver.find_element(By.XPATH, '/html/body/span/span/span[2]/ul/li').text == inv_number:
            #         self.driver.find_element(By.XPATH, '/html/body/span/span/span[2]/ul/li').click()
            #         self.driver.find_element(By.XPATH, '/html/body/div[1]/div[2]/div[1]/div[1]/span/button[2]').click()
            #         break

        except Exception as err:
            print("reprint พัง: ", err)

        # * กลับหน้าเดิม
        self.driver.switch_to.window(prev_window)

    def get_pdf_src_and_print(self):
        # self.pdf_src = self.driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div[2]/div/div[2]/div[2]/div/embed").get_attribute('src') ของ reprint
        self.pdf_src = self.driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div[8]/div/div[2]/div[2]/div/embed").get_attribute('src') #* ของ smco
        self.proc = re.search("(?<=,).*", self.pdf_src)
        self.base64_pdf_data = self.proc.group(0)
        #* แปลง base64 to binary data
        self.bin_pdf_data = base64.b64decode(self.base64_pdf_data)
        
        #* Get the current time in UTC
        self.utc_time = datetime.datetime.now()
        #* Specify the timezone
        self.tz = pytz.timezone('Asia/Bangkok')
        #* Convert the current time to Bangkok time and format it
        self.th_time = self.utc_time.astimezone(self.tz).strftime("%d_%m_%Y-%I_%M_%S_%p")
        
        try:
            with open(f"online_inv_output_{self.th_time}.pdf", "wb") as pdf_file:
                pdf_file.write(self.bin_pdf_data)
                os.startfile("online_inv_output_{self.th_time}.pdf", "print")
                print("Printing complete.")
        except OSError as err:
            print(f"No PDF Reader found: {err}")
            
        #* กดปุ่มแดงปิดหน้า print
        self.driver.find_element(By.XPATH, '/html/body/div[1]/div[2]/div[8]/div/div[1]/div/a').click()

    #! WIP accel_mode[1]หากใช้ accel_mode จะดูว่ามี SN ในไฟล์ที่นำเข้าหรือไม่ ถ้ามีให้ระบุว่าเป็นโหมดของเหมือน(uni-SKU) แล้วเอา SN ยัดลงไป เติม CP ให้เรียบร้อย
    def accel_fill_sku(self):
        self.used_serials = []
        # *  ดึง array items เก็บลงตัวแปร items
        ordered_items = self.app.items
        print('accel_fill_sku() ตรวจสอบ items = ', ordered_items)
        if len(ordered_items) > 0:
            for ordered_item in ordered_items:
                print("item ordered by customer", ordered_item)
                current_sku = ordered_item['เลขอ้างอิง SKU (SKU Reference No.)']
                print("current_sku: ", current_sku)
                sku_qtys = ordered_item['จำนวน']
                if current_sku in self.app.obj_data_from_accel_file:
                    for item in range(sku_qtys):
                        print(
                            "self.app.obj_data_from_accel_file[current_sku]: ", self.app.obj_data_from_accel_file[current_sku])
                        self.app.obj_data_from_accel_file[current_sku]

                        if self.app.obj_data_from_accel_file[current_sku]:
                            print("มี SN")
                            time.sleep(1)
                            while True:
                                try:
                                    # *รอให้ไอนี่ใช้ได้ชัวก่อนค่อยไปทำขั้นตอนต่อไป
                                    self.driver.find_element(
                                        By.XPATH, '/html/body/div[1]/div[2]/div[2]/div[2]/div[1]/div[1]/from/div/div/div[1]/div[1]/span/input')
                                    break
                                except:
                                    continue
                            sn = self.app.obj_data_from_accel_file[current_sku].pop(
                                0)
                            skuInput = self.driver.find_element(
                                By.XPATH, '/html/body/div[1]/div[2]/div[2]/div[2]/div[1]/div[1]/from/div/div/div[1]/div[1]/span/input')
                            skuInput.clear()

                            skuInput.send_keys(sn)
                            print("fill sn complete")

                            # while True:
                            skuInput.send_keys(Keys().ENTER)
                            #! wip ต้องมีตรวจสอบผลลัพธ์การกรอก sn ตรงนี้
                            print("pressed Enter at SKU-Input")
                            print(f'to_sent_dict = sku: {
                                  current_sku}, sn: {sn} ')
                            to_sent_dict = {'sku': current_sku, 'sn': sn}
                            self.used_serials.append(to_sent_dict)
                            time.sleep(2)

                        else:

                            print(
                                "ไม่มี SN, there are no functions available at this moment")
                            pass
                            # self.app.is_accel_mode_activated.set(False)
                            # raise ValueError(
                            #     "There's no SN in Accel File, no functions to handle at this moment.")
                else:
                    print("ไม่ได้เลือก sn:",
                          current_sku in self.app.obj_data_from_accel_file)
                    print("ไม่ได้เลือก sn:", current_sku,
                          self.app.obj_data_from_accel_file)
                    print("ไม่ได้เลือก sn:", type(current_sku),
                          type(self.app.obj_data_from_accel_file))
        else:
            print("No items, return!!")
            return

    def operation_start(self):
        self.app.is_gui_busy.set(True)
        self.is_forbid = False
        is_etax = False
        inv_number = ""
        if self.app.order != "":
            ### * MARKETPLACES Part ########################################################################################
            self.autofinal = False
            print("operation start!! ยังไม่มีไรจะใส่ใส่เป็น placeholderไว้ก่อน")
            self.wait1 = WebDriverWait(self.driver, 50)
            # * เปลี่ยนไปtab MARKETPLACES เพื่อเช็ค status (เพราะไม่มี API เลยต้องทำ และเพื่อดูรูปว่ามีของแถมหรือไม่)

            #### IF MARKETPLACE IS SHOPEE ###################################################################################################################################
            if self.app.marketplace_target.get() == 'SHOPEE':
                self.driver.switch_to.window(self.merged_dict['Seller Centre'])
                cur_url = self.driver.current_url

                # * เปลี่ยนไปใช้หน้า "ทั้งหมด" เพราะ ในที่หน้าต่างกัน css, elements มันต่างกัน บังคับให้มันใช้อันที่ถูก
                if cur_url != "https://seller.shopee.co.th/portal/sale/order":
                    # self.driver.get("https://seller.shopee.co.th/portal/sale/order")
                    self.driver.find_element(
                        By.XPATH, '/html/body/div[1]/div[2]/div[2]/div/div/div/div[2]/div[2]/div[1]/div/div/div/div[1]/div/div[1]/div[1]/div').click()
                    #! ตรงนี้มันไม่ใช้แล้ว
                    # self.wait1.until(EC.text_to_be_present_in_element(
                    #     (By.XPATH, '/html/body/div[1]/div[1]/div/div[1]/div/div[2]/div[1]/div/div[1]/div[1]/a'), 'การขายของฉัน'))
                else:
                    print("อยู๋ในหน้าทั้งหมดอยู่แล้ว ไม่ต้องเปลี่ยน")

                try:
                    # * กรอก order ลงในช่อง search
                    self.search_elmt = self.wait1.until(EC.visibility_of_element_located(
                        # (By.XPATH, '/html/body/div[1]/div[2]/div[2]/div/div/div/div[2]/div[2]/div/div/div[1]/div[1]/div[2]/div[1]/span[2]/div/div[1]/div/div/input'))) เก่า ไม่น่าจะกลับมาใช้แล้ว
                        # (By.XPATH, '/html/body/div[1]/div[2]/div[2]/div/div/div/div[2]/div[3]/div/div/div[2]/div[1]/div[1]/div[1]/div/span[2]/div/div[1]/div/div/input')))
                        # (By.XPATH, '/html/body/div[1]/div[2]/div[2]/div/div/div/div[2]/div[3]/div/div/div[2]/div[1]/div/div[1]/div[1]/div/div/span[2]/div/div[1]/div/div/input'))) พัง 28/08/2024 12:00 PM
                        # (By.XPATH, '/html/body/div[1]/div[2]/div[2]/div/div/div/div[2]/div[4]/div/div/div[2]/div[1]/div/div/div[1]/div[1]/div/div/span[2]/div/div[1]/div/div/input') พัง 19/09/2024 17:00
                        (By.XPATH, '/html/body/div[1]/div[2]/div[2]/div/div/div/div[2]/div[5]/div/div/div[2]/div/div[1]/div/div/div[1]/div[1]/div/div/span[2]/div/div[1]/div/div/input')
                    ))

                    self.search_elmt.clear()
                    self.search_elmt.send_keys(self.app.cus_order.get())

                    # * กด Search เพื่อ เก็บ Status
                    self.searchBtn = self.driver.find_element(
                        # By.XPATH, '/html/body/div[1]/div[2]/div[2]/div/div/div/div[2]/div[2]/div/div/div[1]/div[1]/div[2]/div[2]/button[1]') เก่า ไม่น่าจะกลับมาใช้แล้ว
                        # By.XPATH, '/html/body/div[1]/div[2]/div[2]/div/div/div/div[2]/div[3]/div/div/div[2]/div[1]/div[1]/div[2]/button[1]' พัง 28/08/2024 12:00
                        # By.XPATH, '/html/body/div[1]/div[2]/div[2]/div/div/div/div[2]/div[3]/div/div/div[2]/div[1]/div/div/div[2]/button[1]' พัง 18/09/2024 14:00
                        # By.XPATH, '/html/body/div[1]/div[2]/div[2]/div/div/div/div[2]/div[4]/div/div/div[2]/div[1]/div/div/div[2]/button[1]' พัง 19/09/2024 17:00
                        By.XPATH, '/html/body/div[1]/div[2]/div[2]/div/div/div/div[2]/div[5]/div/div/div[2]/div/div[1]/div/div/div[2]/button[1]'
                    )
                    self.searchBtn.click()
                except:
                    print("cannot search order")
                    raise ValueError(f"method operation_start Error : {
                                     traceback.format_exc()}")

                # * ตรวจสอบ Status และ update ของ MARKETPLACE
                time.sleep(0.75)
                try:
                    self.driver.find_element(
                        By.CLASS_NAME, 'status-wrapper').is_displayed()
                    print("Found element classed big-text")
                except:
                    print(
                        "Not found element classed big-text, try to wait and click element with XPATH")
                    self.wait1.until(EC.element_to_be_clickable(
                        # (By.XPATH, '/html/body/div[1]/div[2]/div[2]/div/div/div/div[2]/div[2]/div/div/div[3]/div/div[3]/a/div[2]/div/div/div'))) เก่า ไม่น่าจะกลับมาใช้แล้ว
                        # (By.XPATH, '/html/body/div[1]/div[2]/div[2]/div/div/div/div[2]/div[3]/div/div/div[4]/div/div[3]/a/div[2]/div/div/div') พัง 28/08/2024 12:00 PM
                        (By.XPATH,
                         '/html/body/div[1]/div[2]/div[2]/div/div/div/div[2]/div[3]/div/div/div[2]/div[4]/div/div[2]/a/div[2]/div/div/div')
                    ))

                # *>  ต้องใช้ try except เพราะ element ของ shopee มันดันแบ่งเป็นสองแบบหากมีสถานะ order ที่ต่างกัน แทนที่จะเขียนให้เหมือนกัน ยุ่งยากกว่าเดิม
                try:
                    # * สำหรับ หาข้อความ "ที่ต้องจัดส่ง" ต่อให้มี element ที่บรรจุคำว่า "จะถูกยกเลินใน x วัน" หรือ "การจัดส่งช้า" ตราบใดที่ข้างล่างมี ที่ต้องจัดส่ง จะมี class big-text เสมอ
                    self.app.cus_cur_status.set(self.driver.find_element(
                        By.CLASS_NAME, 'status-wrapper').text)

                except:
                    # * elementจะแสดงตาม DOM DIR นี้ ถ้าหาก ดูในหน้า ทั้งหมด สำหรับ Order ที่มีสถานะ "ส่งสินค้าแล้ว", "ยกเลิกแล้ว", "สำเร็จ"
                    self.app.cus_cur_status.set(self.driver.find_element(
                        # By.XPATH, '/html/body/div[1]/div[2]/div[2]/div/div/div/div[2]/div[2]/div/div/div[3]/div/div[3]/a/div[2]/div/div/div/div[3]/div[1]/span').text) เก่า ไม่น่าจะกลับมาใช้แล้ว
                        # By.XPATH, '/html/body/div[1]/div[2]/div[2]/div/div/div/div[2]/div[3]/div/div/div[4]/div/div[3]/a/div[2]/div/div/div/div[3]/div[1]/span').text) พัง 28/08/2024 12:00 PM
                        # By.XPATH, '/html/body/div[1]/div[2]/div[2]/div/div/div/div[2]/div[3]/div/div/div[2]/div[4]/div/div[2]/a/div[2]/div/div/div/div[3]/div/div[2]/span').text) พัง 18/09/2024
                        By.XPATH, '/html/body/div[1]/div[2]/div[2]/div/div/div/div[2]/div[4]/div/div/div[2]/div[4]/div/div[2]/a/div[2]/div/div/div[3]/div/div[1]/span').text)

                # * จะได้ element มา
                print("realtime_status_text", self.app.cus_cur_status.get())
                self.app.display_current_status.config(
                    fg="#000000", bg="#8fd4ff")
                if self.app.cus_cur_status.get() == "ส่งสินค้าแล้ว":
                    self.app.display_current_status.config(
                        bg="#00ff11", fg="#000000")
                    PopUp("Caution!!", f"Order นี้มีสถานะ '{
                          self.app.cus_cur_status.get()}' จะทำต่อจริงอ่อ?", self.parent, "alert")

                elif "ยกเลิก" in self.app.cus_cur_status.get():
                    self.app.display_current_status.config(
                        bg="#ff2b2b", fg="#FFF")
                    self.is_forbid = True
                    #! WIP accel_mode[3] ถ้าเป็น accel mode อาจจะไม่ต้องใช้ popup แต่ใช้เป็นการเก็บผลลัพธ์การทำงานแทน
                    PopUp("Caution!!", f"Order นี้มีสถานะ '{
                          self.app.cus_cur_status.get()}' จะทำต่อจริงอ่อ?", self.parent, "alert")

                self.is_status_true = self.app.order_status == self.app.cus_cur_status.get()
                if self.is_status_true:
                    print(self.app.order_status ==
                          self.app.cus_cur_status.get())
                    print("Status in the file is reliable")
                else:
                    print(self.app.order_status ==
                          self.app.cus_cur_status.get())
                    print(
                        "Status in the file is unreliable, suggest downloading a new Export File from the link below")
                    print(
                        "https://seller.shopee.co.th/portal/sale/shipment?type=toship")

            #### * IF MARKETPLACE IS LAZADA ###########################################################################################################################
            elif self.app.marketplace_target.get() == 'LAZADA':
                try:
                    self.driver.switch_to.window(
                        self.merged_dict['การจัดการคำสั่งซื้อ - Lazada Seller Center'])
                except:
                    self.driver.switch_to.window(
                        self.merged_dict['การจัดการคำสั่งซื้อ - Seller Center'])

                cur_url = self.driver.current_url

                # * เปลี่ยนไปใช้หน้า "ทั้งหมด" เพราะ ในที่หน้าต่างกัน css, elements มันต่างกัน บังคับให้มันใช้อันที่ถูก
                if cur_url != "https://sellercenter.lazada.co.th/apps/order/list?oldVersion=1&spm=a1zawg.23708326.navi_left_sidebar.droot_normal_ordersreviews_ordersnewui.3fa34edfUCdGFY&status=all":
                    self.driver.find_element(
                        By.XPATH, '/html/body/div/section/div[2]/div/div[1]/div/div/div[2]/div/div[1]/div/div/div/ul/li[1]/div').click()
                    time.sleep(0.75)
                    self.wait1.until(EC.element_to_be_clickable(
                        (By.XPATH, '/html/body/div/section/div[2]/div/div[1]/div/div/div[3]/div/div[3]/div[1]/div[1]/div[2]/div[2]/span[1]/span[2]/span/a')))
                # else:
                #     pass

                # * กรอก order ลงในช่อง search
                self.search_elmt = self.wait1.until(EC.visibility_of_element_located(
                    (By.XPATH, '/html/body/div/section/div[2]/div/div[1]/div/div/form/div[2]/div/div/div/div[1]/div[3]/div[1]/div/div/span/span[1]/span[1]/span/input')))

                self.driver.find_element(
                    By.XPATH, '/html/body/div[1]/section/div[2]/div/div[1]/div/div/form/div[2]/div/div/div/div[1]/div[3]/div[1]/div/div/span/span[1]/span[1]/span/input').clear()

                self.input_count = []

                try:
                    close_btn = self.driver.find_element(
                        By.XPATH, '/html/body/div/section/div[2]/div/div[1]/div/div/form/div[2]/div/div/div/div[1]/div[3]/div[1]/div/div/span/span[1]/span[1]/div[1]/span[2]')

                    try:
                        self.input_count = self.driver.find_element(
                            By.XPATH, '/html/body/div/section/div[2]/div/div[1]/div/div/form/div[2]/div/div/div/div[1]/div[3]/div[1]/div/div/span/span[1]/span[1]/div[2]/span/span')
                    except:
                        print("Have only one input")
                except:
                    print("Input is empty")

                try:
                    if self.input_count.is_displayed() and close_btn.is_displayed():
                        clicks = re.sub(r'\W', "", self.input_count.text)
                        print("จำนวนครั้งของการกด x ", clicks)
                        print(f"{clicks} times click")
                        for click in range(int(clicks)):
                            close_btn = self.driver.find_element(
                                By.XPATH, '/html/body/div/section/div[2]/div/div[1]/div/div/form/div[2]/div/div/div/div[1]/div[3]/div[1]/div/div/span/span[1]/span[1]/div[1]/span[2]')
                            print("click", click)

                            close_btn.click()
                    else:
                        print("1 times click")
                        close_btn.click()

                except Exception as err:
                    try:
                        close_btn.click()
                        print('1 Button closed ')
                    except:
                        print('No close Button')
                        pass

                self.search_elmt.clear()
                self.search_elmt.send_keys(self.app.cus_order.get())

                # * กด Search เพื่อ เก็บ Status
                self.searchBtn = self.driver.find_element(
                    By.XPATH, '/html/body/div/section/div[2]/div/div[1]/div/div/form/div[2]/div/div/div/div[1]/div[3]/div[1]/div/div/div[1]')
                self.searchBtn.click()
                time.sleep(0.75)

                # * ตรวจสอบ Status และ update
                # รอให้ btn element กดได้
                self.wait1.until(EC.element_to_be_clickable(
                    (By.XPATH, '/html/body/div/section/div[2]/div/div[1]/div/div/div[3]/div/div[3]/div/div[2]/div/div/div[5]/div[1]/button')))

                # เก็บ status order เข้าตัวแปรไปแสดงผลใน GUI
                self.app.cus_cur_status.set(self.driver.find_element(
                    By.XPATH, '/html/body/div/section/div[2]/div/div[1]/div/div/div[3]/div/div[3]/div/div[2]/div/div/div[5]/div[1]/button/span').text)

                # จะได้ element มา
                print("realtime_status_text", self.app.cus_cur_status.get())
                self.app.display_current_status.config(
                    fg="#000000", bg="#8fd4ff")
                if "พิมพ์ใบแจ้งหนี้" in self.app.cus_cur_status.get():
                    self.app.display_current_status.config(
                        bg="#ff2b2b", fg="#FFF")
                    self.is_forbid = True
                elif self.app.cus_cur_status.get() == "สถานะการจัดส่ง":
                    self.app.display_current_status.config(
                        bg="#00ff11", fg="#000000")

            #### * IF MARKETPLACE NON OF THEM ABOVE ###################################################################################################################
            else:
                self.driver.switch_to.window(self.merged_dict[''])
                print('Cannot Define What marketplace you are working with')

            # * ถ้าสถานะยกเลิก ก็หยุดเลย
            if self.is_forbid:
                print("This order was forbidden.")
                self.display_bot_status_label.config(
                    text=f"Bot Status: ˶ᵔ ᵕ ᵔ˶ จบการทำงาน", bg="#d9f2ff", fg="#000")
                return

            ### * SMCO PART ############################################################################
            # * เปลี่ยนไปtab SMCO0 เพื่อเช็ค ชื่อลูกค้า
            self.driver.switch_to.window(
                self.merged_dict['SMCO :: เปิดการขาย'])

            # * ดูก่อนว่าเคลียชื่อลูกค้าแล้วเหรอยัง
            self.cus_name_span_elmt = self.driver.find_element(
                By.XPATH, '/html/body/div[1]/div[2]/div[2]/div[2]/div[2]/div[1]/div[2]/div/div/div[6]/form/div/span/span[1]/span/span[1]/span')
            self.cus_name_span_text = self.cus_name_span_elmt.text
            if self.cus_name_span_text == 'Please select':
                self.is_reset = False
            elif self.cus_name_span_text == 'กรุณาเลือก':
                self.is_reset = False
            else:
                self.is_reset = True
                print("มีชื่อลูกค้าอยู่แล้ว")

            try:
                print("เช็คว่าต้องรีไหม", self.is_reset)
                if self.is_reset:
                    print("รีนี่หว่า, กดรีเลย")
                    self.driver.find_element(
                        By.XPATH, '/html/body/div[1]/div[2]/div[2]/div[2]/div[2]/div[1]/div[2]/div/div/div[6]/form/div/span/span[1]/span/span[1]/span').click()
                    try:
                        # คลิกเพื่อให้ปิด droprdown
                        self.driver.find_element(
                            By.XPATH, '/html/body/div[1]/div[2]/div[2]/div[2]/div[2]/div[1]/div[2]/div/div/div[6]/form/div/span/span[1]/span/span[1]').click()
                    except:
                        # * ถ้ามีสินค้าจะ error คลิกไม่ได้จะกลายเป็น except
                        try:
                            print("wait for pop-up(try)")
                            # ระบุปุ่ม ok
                            if self.driver.find_element(By.XPATH, '/html/body/div[16]/div[2]/button[1]'):
                                print("has pop-up(try)")
                                self.driver.find_element(
                                    By.XPATH, '/html/body/div[16]/div[2]/button[1]').click()
                                print("Click OK(try)")

                        except:
                            print("wait for pop-up(except)")
                            time.sleep(1)
                            # * ระบุปุ่ม ok
                            if self.driver.find_element(By.XPATH, '/html/body/div[16]/div[2]/button[1]'):
                                print("has pop-up(except)")
                                self.driver.find_element(
                                    By.XPATH, '/html/body/div[16]/div[2]/button[1]').click()
                                print("Click OK(except)")
                        # * ถ้ามีสินค้าแล้วกดลบชื่อ มันจะมีชื่อค้างอยู่แต่สินค้าหายต้องกดอีกรอบ
                        try:
                            self.driver.find_element(
                                By.XPATH, '/html/body/div[1]/div[2]/div[2]/div[2]/div[2]/div[1]/div[2]/div/div/div[6]/form/div/span/span[1]/span/span[1]/span').click()
                            print("Cusname still appear the btn 'x' is available.")
                        except:
                            print("Cusname has disappeared no 'x' to press.")

                    print("หน้าใหม่พร้อมแล้ว")
                elif self.is_reset == False:
                    print("ไม่ต้องรี")
            except EXCEPTION as err:
                print("Error From SMCO phase1 Resetting", err)
                while True:
                    print("รอ")
                    time.sleep(1)
                    if self.driver.find_element(By.XPATH, '/html/body/div[1]/div[2]/div[2]/div[1]/label/div/button'):
                        print("เจอแล้ว")
                        break
                    else:
                        continue

            print("ผ่านเคลียชื่อลูกค้า, รอ element โผล่")
            # while True:
            #     if self.driver.find_element(By.XPATH, '/html/body/div[1]/div[2]/div[2]/div[2]/div[2]/div[1]/div[2]/div/div/div[5]/div/div/button'):
            #         print("เจอแล้วออก")
            #         break
            #     else:
            #         continue

            self.wait1.until(EC.element_to_be_clickable(
                (By.XPATH, '/html/body/div[1]/div[2]/div[2]/div[2]/div[2]/div[1]/div[2]/div/div/div[5]/div/div/button')))

            time.sleep(1)
            # * เปลี่ยน auto เป็น name ไม่ก็ email โดยขึ้นอยู่กับว่าขอใบกำกับหรือไม่
            self.driver.find_element(
                By.XPATH, '/html/body/div[1]/div[2]/div[2]/div[2]/div[2]/div[1]/div[2]/div/div/div[5]/div/div/button').click()
            print("self.app.tax_bool: ", self.app.tax_bool.get())

            # * จากปัญหาข้อที่ 39 // รอให้ตัวเลือกภายใน click ได้ก่อน แล้วค่อย เลือก วิธีการ search
            self.wait1.until(EC.element_to_be_clickable(
                (By.XPATH, r'''/html/body/div[1]/div[2]/div[2]/div[2]/div[2]/div[1]/div[2]/div/div/div[5]/div/div/div/a[contains(@ng-click, "st='E'")]''')))
            if self.app.tax_bool.get() == True:
                # ขอใบกำกับ **Trick** สามารถใส่single qoute สามตัวได้ หากด้านในมีการใช้ qoute และ bouble qoute ไปแล้ว แต่ทั้งหมดต้องเป็น string อีกที >>  ('''function("vbvb, x='แมว'")''')
                if self.app.marketplace_target.get() == "SHOPEE":
                    print("ขอใบกำกับSHOPEE ใช้ T:")
                    self.driver.find_element(
                        By.XPATH, r'''/html/body/div[1]/div[2]/div[2]/div[2]/div[2]/div[1]/div[2]/div/div/div[5]/div/div/div/a[contains(@ng-click, "st='T'")]''').click()
                elif self.app.marketplace_target.get() == "LAZADA":
                    print("ขอใบกำกับLazada ใช้ T:")
                    self.driver.find_element(
                        By.XPATH, r'''/html/body/div[1]/div[2]/div[2]/div[2]/div[2]/div[1]/div[2]/div/div/div[5]/div/div/div/a[contains(@ng-click, "st='T'")]''').click()
            elif self.app.tax_bool.get() == False:
                # ไม่ขอใบกำกับ
                print("ไม่ขอใบกำกับใช้ N:")
                self.driver.find_element(
                    By.XPATH, r'''/html/body/div[1]/div[2]/div[2]/div[2]/div[2]/div[1]/div[2]/div/div/div[5]/div/div/div/a[contains(@ng-click,"st='N'")]''').click()

            # * ดูว่า self.cus_search_input จะต้องถูกกำหนดค่าเป็นเลขใบกำกับหรือชื่อ อิงจาก tax_bool choosing by ternary like conditional
            # 09/11/2023 ใช้เลขใบกำกับเสิชไม่ได้แล้ว ฉะนั้นไม่ต้องเลือกแล้ว เอาชื่อเสิชให้หมดเลย

            # if self.app.marketplace_target.get() == "SHOPEE":
            #     self.cus_search_input = self.app.cus_email.get() if self.app.tax_bool.get(
            #     ) else self.app.cusNameFixer5(self.app.cus_name.get(), self.app.cus_account_name.get())
            # elif self.app.marketplace_target.get() == "LAZADA":
            #     self.cus_search_input = self.app.tax_num.get() if self.app.tax_bool.get(
            #     ) else self.app.cusNameFixer5(self.app.cus_name.get(), self.app.cus_account_name.get())

            #! WIP
            # * 05/07/2024 Shopeeนั้นได้ลบ ชื่อลูกค้าแบบ ธรรมดา ออกไปอย่างถาวร จึงต้องปรับวิธีออกบิลให้กับแบบธรรมดาโดยการใช้ "account"+" ชื่อที่เป็นดอกจัน"+" หมายเลขโทรศัพท์"
            # self.cus_search_input = self.app.tax_num.get() if self.app.tax_bool.get(
            # ) else self.app.cusNameFixer5(self.app.cus_name.get(), self.app.cus_account_name.get())

            if self.app.marketplace_target.get() == "SHOPEE":
                self.cus_search_input = self.app.tax_num.get() if self.app.tax_bool.get(
                ) else self.app.cusNameFixer5(self.app.cus_name.get())
            elif self.app.marketplace_target.get() == "LAZADA":
                self.cus_search_input = self.app.tax_num.get() if self.app.tax_bool.get(
                ) else self.app.cusNameFixer5(self.app.cus_name.get(), self.app.cus_account_name.get())

            # * จับตาดูว่า ul เปิดอยู่ไหม
            self.is_ul_not_open = False if self.driver.find_elements(
                By.XPATH, self.app.cus_name_dropdown_ul) else True
            # * กรณีไม่ได้เปิดไว้ จะเปิดให้
            if self.is_ul_not_open:
                self.driver.find_element(
                    By.XPATH, self.app.cus_arrow_btn).click()

                self.wait1.until(EC.visibility_of_element_located(
                    (By.XPATH, self.app.cusNameInput)))

            # * ถ้าเปิดแล้วจะข้ามมานี่
            while True:
                self.enter_cus_name(self.cus_search_input)
                print("กรอกชื่อเสร็จ")
                # * wait_condition มันจะเจอ cusNameLi1 ที่ containค่า "Searching..."
                self.wait_condition = self.driver.find_element(
                    By.XPATH, self.app.cusNameLi1)
                # * มันจะได้ Searching...
                print("มันทำไม", self.wait_condition.text)

                # ? WIP แก้ละรอดูว่าพังไหม //pop-up เด้งแทรกตอนกรอกชื่อลูกค้าในช่อง search
                # pop-up อันนึงเด้งมาหลังจาก กรอกชื่อ  xpath : "/html/body/div[16]/div[2]/div[6]" text: "Reload data not complete,reload page verify data again." button:"/html/body/div[16]/div[2]/button[1]"
                try:
                    # * มี pop-upไหม
                    if self.driver.find_element(By.XPATH, "/html/body/div[16]/div[2]/div[6]").is_displayed():
                        # * ถ้ามี ปิด แล้วเริ่มไปกรอกชื่อใหม่
                        self.driver.find_element(
                            By.XPATH, "/html/body/div[16]/div[2]/button[1]").click()
                        continue
                    # * ไม่มี pop-up ให้ break
                    break
                except:
                    break

            # * ตาม Stepแล้วนั้น ขั้นตอนด้านบนจะทำให้ Dropdown UL มันโผล่ และมี li อย่างน้อย 1 อัน นั่นคือ li[0] โดย li[0] จะบอกสถานะของการ search ตั้งแต่ "Searching...", "No results found", ไม่แน่ใจมีอีกไหม และแสดง ผลลัพธ์ที่เจอลำดับแรก
            self.customer_added_times = 0
            self.customer_name_search_count = 0
            while True:
                if self.driver.find_element(By.XPATH, self.app.cus_name_dropdown_ul):
                    time.sleep(0.7)
                    # self.wait1.until(EC.visibility_of_element_located(
                    #     (By.XPATH, self.app.cusNameLi1)))

                    # * li[1] เป็นตัวที่แสดงผลแบบ dynamic เราจะตรวจจับ พฤติกรรมของ element นี้
                    self.wait_condition = self.driver.find_element(
                        By.XPATH, self.app.cusNameLi1)

                    # * ช่วงรอ ผลลัพของ Searching...
                    try:
                        if self.wait_condition.text == "Searching...":
                            continue
                        elif self.wait_condition.text:
                            print("text element not display Searching...")
                            pass
                    except:
                        pass

                    # * หลังจาก Searching... หายไป ๑๑๑๑๑๑๑๑๑๑๑๑๑๑๑๑๑๑๑๑๑๑๑๑๑๑๑๑๑๑๑๑๑๑๑๑๑๑
                    self.wait1.until(EC.visibility_of_element_located(
                        (By.XPATH, self.app.cusNameLi1)))
                    self.wait_condition = self.driver.find_element(
                        By.XPATH, self.app.cusNameLi1)

                    # * กรณี ไม่เจอผลลัพธ์ ทำการ Add ใหม่
                    if self.wait_condition.text == "No results found" and self.customer_added_times == 0:
                        print("No results found and NeverAdd")
                        #! ปิดไว้ก่อน จะเทสของใหม่
                        # # * ขอใบกำกับป่าว
                        # if self.app.tax_bool.get():
                        #     print("Tax_needed")
                        #     if self.app.marketplace_target.get() == 'SHOPEE':
                        #         self.addTaxInvCustomer()

                        #     # * กำลังทำ กำลังปรับปรุง ยังไม่เสร็จ การหาลูกค้าของ laz มันมีกรณี excel และ api
                        #     elif self.app.marketplace_target.get() == 'LAZADA':
                        #         self.addTaxInvCustomerLaz()

                        # else:
                        #     print("no_Tax_needed")
                        #     self.addNormalCustomer(self.cus_search_input)
                        self.add_cusname()

                        # * เพิ่มจำนวนครั้งที่ add
                        self.customer_added_times += 1
                        self.driver.switch_to.window(
                            self.merged_dict['SMCO :: เปิดการขาย'])
                        print("ก่อนRe Enter ชื่อลูกค้า")
                        self.enter_cus_name(self.cus_search_input)
                        print(f"Re enter name after add")
                        continue
                    # * หลังจาก Add ไปแล้วรอบนึง แล้วมาเสิชใหม่แล้วยังไม่เจอ ถึงจะเข้าเงื่อนไขนี้ เป็นการ search ให้อีกรอบนึง
                    elif self.wait_condition.text == "No results found" and self.customer_name_search_count < 1:
                        self.enter_cus_name(self.cus_search_input)
                        self.customer_name_search_count += 1
                        print(f"Re enter name after add extra times{
                              self.customer_name_search_count}")
                        continue
                    # * Add แล้ว รีเสิชให้สองรอบแล้ว ก็ยังไม่เจอ ลองแอดด้วยตัวเองดู
                    elif self.wait_condition.text == "No results found" and self.customer_added_times == 1:
                        print(
                            "I've already add it, but the element still shows 'No results found', you have to add by yourself")
                        break
                    else:
                        self.driver.switch_to.window(
                            self.merged_dict['SMCO :: เปิดการขาย'])
                        break
                print("addcustomer and select While end!")
                break

            # !66 WIP เปลี่ยนวิธีเลือกชื่อลูกค้า เดิมทีคือเลือก
            while True:
                try:
                    customer_name_input_ul = self.driver.find_element(
                        By.XPATH, self.app.cus_name_dropdown_ul)
                    customer_name_dropdown_lis = customer_name_input_ul.find_elements(
                        By.CSS_SELECTOR, '.select2-results__option')
                    print("หาจำนวน li ชื่อลูกค้าเท่ากับ:",
                          customer_name_dropdown_lis)
                    break

                except:
                    self.driver.find_element(
                        By.XPATH, self.app.cus_arrow_btn).click()
                    continue

            if len(customer_name_dropdown_lis) > 1:
                print("มากกว่า 1")
                li_names = [
                    element.text for element in customer_name_dropdown_lis]
                self.select_cus_name_from_lis(
                    li_names, self.select_cus_name_from_lis)
                print("click แล้ว")
            else:
                self.driver.find_element(By.XPATH, self.app.cusNameLi1).click()
                print("Click the cusname li result")

            # * กรณีมีสินค้ายิงไปแล้ว แล้วมีการเปลี่ยนชื่อลูกค้า มันจะมี alert // path นี้คือ element นอกของ alert /html/body/div[16]/div[2]
            if self.driver.find_element(By.XPATH, "/html/body/div[16]/div[2]").is_displayed():
                try:
                    self.driver.find_element(
                        By.XPATH, "/html/body/div[16]/div[2]/button[1]").click()
                    self.driver.find_element(
                        By.XPATH, self.app.cus_arrow_btn).click()
                    self.wait1.until(EC.visibility_of_element_located(
                        (By.XPATH, self.app.cusNameInput)))
                except:
                    print("Skip, Alert Element is appear but can not perform actions.")
            else:
                print("Skip, Alert Element is Not appear")
                pass

            print("search หายไปแล้ว")
            self.wait1.until(EC.invisibility_of_element_located(
                (By.XPATH, self.app.cusNameInput)))

            # * ใส่ค่าขนส่ง
            # * ค่าขนส่งเราจะใส่ให้ SHOPEE เท่านั้น
            if self.app.marketplace_target.get() == "SHOPEE":
                if int(self.app.cus_ship_cost.get()) != int(0):
                    try:
                        self.skuInput_element = self.wait1.until(EC.visibility_of_element_located(
                            (By.XPATH, '/html/body/div[1]/div[2]/div[2]/div[2]/div[1]/div[1]/from/div/div/div[1]/div[1]/span/input')))
                        # skuInput = driver.find_element(By().XPATH,'/html/body/div[1]/div[2]/div[2]/div[2]/div[1]/div[1]/from/div/div/div[1]/div[1]/span/input')
                        self.skuInput_element.clear()

                        self.skuInput_element.send_keys("SV0-000101")
                        print("กรอก Code ขนส่งสำเร็จ")

                        self.skuAddBtn = self.wait1.until(EC.visibility_of_element_located(
                            (By.XPATH, '/html/body/div[1]/div[2]/div[2]/div[2]/div[1]/div[1]/from/div/div/div[1]/div[1]/span/input')))
                        # skuAddBtn = driver.find_element(By().XPATH,'/html/body/div[1]/div[2]/div[2]/div[2]/div[1]/div[1]/from/div/div/div[1]/div[1]/span/input')
                        self.skuAddBtn.send_keys(Keys().ENTER)
                        print("กด Enter ที่ช่อง SKU Input สำเร็จ")

                        #! WIP ทดสอบ 1/2 หยุดเพื่อให้จบ if ก่อน แล้ว2/2 จะเป็นชั้นที่จบ scope จริงๆ รู้สึก return ตรนี้ใช้แล้วจะจบเลย ไม่ได้จบแค่ if งั้นเหรอ
                        # logger.info(f"Order: {self.app.order} 1/2Finished!!")
                        # return
                        time.sleep(2)

                        # ทำไมต้องใส่วงเล็บ คลุม BY.XPATH เพราะ ถ้าไม่ใส่ ฟังชัน visibility จะมอง xpath เป็น argument ที่สอง ของ method visibility
                        self.definePrice_btn_element = self.wait1.until(EC.visibility_of_element_located(
                            (By.XPATH, '/html/body/div[1]/div[2]/div[2]/div[2]/div[1]/div[2]/div[1]/div/div[2]/div[1]/div[1]/div/a[1]')))
                        # self.definePrice_btn_element = driver.find_element(By().XPATH,'/html/body/div[1]/div[2]/div[2]/div[2]/div[1]/div[2]/div[1]/div/div[2]/div[1]/div[1]/div/a[1]')
                        self.definePrice_btn_element.click()
                        time.sleep(1)
                        # ค่าขนส่งโดนข้า230208FX99FUGGมหลังจากตรงนี้
                        print("กดที่ SKU ELEMENT 1 สำเร็จ")

                        self.changePriceInput = self.driver.find_element(
                            By().XPATH, '/html/body/div[1]/div[2]/div[2]/div[2]/div[6]/div/div/div[2]/div[2]/div[1]/input')
                        self.changePriceInput.clear()
                        self.changePriceInput.send_keys(
                            self.app.cus_ship_cost.get())
                        self.driver.find_element(
                            By().XPATH, '/html/body/div[1]/div[2]/div[2]/div[2]/div[6]/div/div/div[2]/div[2]/div[2]/input').clear()
                        self.driver.find_element(
                            By().XPATH, '/html/body/div[1]/div[2]/div[2]/div[2]/div[6]/div/div/div[2]/div[2]/div[2]/input').send_keys(self.app.user_id.get())

                        self.driver.find_element(
                            By().XPATH, '/html/body/div[1]/div[2]/div[2]/div[2]/div[6]/div/div/div[2]/div[2]/div[3]/input').clear()
                        self.driver.find_element(
                            By().XPATH, '/html/body/div[1]/div[2]/div[2]/div[2]/div[6]/div/div/div[2]/div[2]/div[3]/input').send_keys(self.app.user_pw.get())

                        self.driver.find_element(
                            By().XPATH, '/html/body/div[1]/div[2]/div[2]/div[2]/div[6]/div/div/div[2]/div[5]/div/textarea').clear()
                        self.driver.find_element(
                            By().XPATH, '/html/body/div[1]/div[2]/div[2]/div[2]/div[6]/div/div/div[2]/div[5]/div/textarea').send_keys("Online")

                        self.driver.find_element(
                            By().XPATH, '/html/body/div[1]/div[2]/div[2]/div[2]/div[6]/div/div/div[2]/div[6]/a[1]').click()
                        try:
                            print("รอหาย")
                            self.wait1(EC.invisibility_of_element_located(
                                (By().XPATH, '/html/body/div[1]/div[2]/div[2]/div[2]/div[6]/div/div/div[2]/div[6]/a[1]')))
                        except:
                            print("ไม่มีให้รอ")
                    except Exception as err:
                        print("ค่าขนส่งโดนข้าม")
                        print(err)
                else:
                    print("เงื่อนไขค่าขนส่ง มี Boolean เป็น False")

            self.app.update_log(
                "Autoหน้าแรก มันจบแค่นี้ ยิงของ, ใส่คูปอง, กดไปหน้าถัดไปได้เลย")
            self.app.display_bot_status_label.config(
                text=f"Bot Status: Your Turn", bg="#21ff29", fg="#000")

            ### PHASE2 After Add Product###############################################################################################################
            # # #เช็คของเติม CP อัตโนมัติ กำลังทำ ถ้าเอาไปใส่ใน while loop ข้างล่างมันจะบัค ไม่สามารถแปลงเป็น float ได้
            # while True:
            #     self.phase1_net_price = self.driver.find_element(By.XPATH, '/html/body/div[1]/div[2]/div[2]/div[2]/div[2]/div[2]/div/div/div/div/span[1]')
            #     self.phase1_net_price = float(self.phase1_net_price.text)
            #     if self.app.phase1_sum_price != self.phase1_net_price:
            #         print("ราคาไม่ตรง", self.app.phase1_sum_price,  " = ",self.phase1_net_price)
            #         continue
            #     else:
            #         print("ราคาตรงแล้ว")
            #         break

            #! WIP ทดสอบ 2/2
            # logger.info(f"Order: {self.app.order} 2/2Finished!!")
            # self.app.search_complete.set()
            # return

            if self.app.is_accel_mode_activated.get():
                self.accel_fill_sku()

            self.autofinal = True
            while self.autofinal:
                self.app.is_gui_busy.set(False)
                print("เข้า final loop ")
                print("รอให้มันโผล่")
                while self.parent.winfo_exists() and self.autofinal:
                    time.sleep(0.55)
                    while True:
                        # * รอ elementก่อน ถ้ามีค่อยออกจาก loop
                        try:
                            # print("loop หลัก")
                            self.cus_name_input_element = self.driver.find_element(
                                By.XPATH, '/html/body/div[1]/div[2]/div[2]/div[2]/div[2]/div[1]/div[2]/div/div/div[6]/form/div/span/span[1]/span/span[1]')
                            title_attribute = self.cus_name_input_element.get_attribute(
                                "title")

                            self.is_final_displayed = self.driver.find_element(
                                By.XPATH, '/html/body/div[1]/div[2]/div[6]/form/div[1]/span[1]').is_displayed()
                            # self.is_input_empty = re.search(
                            #     "^C[0-9]+\-", self.cus_name_input_element.text)
                            break
                        except:
                            # * ไม่มี element ให้วนเรื่อยๆ
                            continue

                    # *ดึงตัวอักษรออกมา
                    x = re.search(
                        "^C[0-9]+", title_attribute)
                    try:
                        self.is_input_empty = x.group()
                    except:
                        self.is_input_empty = ""

                    # * แก้ bot ดับจาก alert
                    while True:
                        time.sleep(0.55)
                        try:

                            sn_window = self.driver.find_element(
                                By.XPATH, '/html/body/div[1]/div[2]/div[2]/div[2]/div[7]/div/div/div[1]')
                            # print("SN_window is still there")
                            if sn_window.is_displayed():
                                # print("หน้า SN กำลังโชว์")

                                # if self.driver.find_element(By.XPATH, '/html/body/div[1]/div[2]/div[2]/div[2]/div[2]/div[1]/div[2]/div/div/div[6]/form/div/span/span[1]/span/span[1]/span').is_displayed():
                                continue

                            else:
                                # print("หน้า SN ไม่ได้โ๙ว์")
                                break
                        except UnexpectedAlertPresentException as err:
                            # self.alert_text = self.driver.switch_to.alert.text ใช้ไม่ได้
                            # print("alertทั้งหมดคือไร", err)
                            print("แสดงผล obj err แค่ส่วนเดียว", err.alert_text)
                            PopUp("SNซ้ำ", f'{err.alert_text}',
                                  self.parent, "alert")
                            # self.driver.switch_to.window(self.merged_dict['SMCO :: เปิดการขาย'])
                            # WebDriverWait(self.driver, 3).until(EC.alert_is_present())
                            # print("Popupโผล่")
                            continue

                    # print("ว่างแล้วไม่ใช่เหรอวะ: ", self.is_input_empty)
                    # print("type(self.is_input_empty): ", type(self.is_input_empty))
                    # print("self.cus_name_input_element.text: ", self.cus_name_input_element.text)
                    if self.is_input_empty == "" and self.is_final_displayed == False:
                        print("ชื่อหาย")
                        break
                    elif (self.cus_name_input_element.text != "Select Customer" or self.cus_name_input_element.text != "กรุณาเลือก") and self.is_final_displayed == False:
                        continue
                    elif (self.cus_name_input_element.text != "Select Customer" or self.cus_name_input_element.text != "กรุณาเลือก") and self.is_final_displayed == True:
                        self.app.is_gui_busy.set(True)
                        time.sleep(0.55)
                        print("หน้า จ่ายตัง")
                        self.is_final_page2 = self.wait1.until(EC.visibility_of_element_located(
                            (By.XPATH, '/html/body/div[1]/div[2]/div[6]/form/div[1]/span[1]')))
                        self.last_page = self.driver.find_element(
                            By.XPATH, '/html/body/div[1]/div[2]/div[6]/form/div[1]/span[1]')
                        if (self.last_page.text == "Payment:") or (self.last_page.text == "ชำระเงิน:"):
                            # Auto หน้าท้าย ทำได้ครั้งเดียว
                            self.is_final_page2 = self.wait1.until(EC.visibility_of_element_located(
                                (By.XPATH, '/html/body/div[1]/div[2]/div[6]/form/div[1]/span[1]')))

                            # self.is_final_page = self.wait1.until(EC.visibility_of_element_located(
                            #     (By.XPATH, '/html/body/div[1]/div[2]/div[6]/form/div[2]/div/div[5]/div[1]/textarea')))
                            try:
                                if self.app.cus_seller_voucher.get():
                                    # ถ้ามี เซลเลอร์ให้ ให้กรอกให้ด้วย
                                    self.driver.find_element(
                                        By.XPATH, '/html/body/div[1]/div[2]/div[6]/form/div[2]/div/div[5]/div[3]/div[1]/div[2]/input').clear()
                                    self.driver.find_element(
                                        By.XPATH, '/html/body/div[1]/div[2]/div[6]/form/div[2]/div/div[5]/div[3]/div[1]/div[2]/input').send_keys(self.app.cus_seller_voucher.get())

                                # ถ้าไม่มี seller ก็ไปกรอก remark ได้เลย
                                self.driver.find_element(
                                    By.XPATH, "/html/body/div[1]/div[2]/div[6]/form/div[2]/div/div[5]/div[1]/textarea").clear()
                                self.driver.find_element(
                                    By.XPATH, "/html/body/div[1]/div[2]/div[6]/form/div[2]/div/div[5]/div[1]/textarea").send_keys(self.app.cus_order.get())

                                # เลือกประเภทชำระเงิน
                                if self.app.marketplace_target.get() == 'SHOPEE':
                                    # เลือก shopee
                                    self.driver.find_element(
                                        By.XPATH, "//a[contains(.,'SHOPEE')]").click()
                                elif self.app.marketplace_target.get() == 'LAZADA':
                                    # เลือก lazada
                                    self.driver.find_element(
                                        By.XPATH, "//a[contains(., 'LAZ')]").click()

                                self.driver.find_element(
                                    By.XPATH, '/html/body/div[1]/div[2]/div[6]/form/div[2]/div/div[7]/div/div[3]/div/div[2]/div[2]/div[3]/div[2]/div[1]/div[2]/input').clear()
                                self.driver.find_element(
                                    By.XPATH, '/html/body/div[1]/div[2]/div[6]/form/div[2]/div/div[7]/div/div[3]/div/div[2]/div[2]/div[3]/div[2]/div[1]/div[2]/input').send_keys(self.app.cus_order.get())

                                try:
                                    self.driver.find_element(
                                        # จู่ๆ brows()btn มันก็ทำงานเลยต้องคลิกเพื่อปิด
                                        By.XPATH, '/html/body/div[1]/div[2]/div[6]/form/div[2]/div/div[5]/div[3]/div[1]/div[1]/div/div/div/div/div[2]/center/button[2]').click()
                                except:
                                    print("ปุ่ม Brows() ไม่โผล่")
                                # ลูกค้ามีชื่อไหม ถ้าไม่มี ใส่ a
                                if self.app.cus_name.get():
                                    self.driver.find_element(
                                        By.XPATH, '/html/body/div[1]/div[2]/div[6]/form/div[2]/div/div[7]/div/div[3]/div/div[2]/div[2]/div[2]/div[2]/div[1]/div[2]/input').clear()
                                    self.driver.find_element(
                                        By.XPATH, '/html/body/div[1]/div[2]/div[6]/form/div[2]/div/div[7]/div/div[3]/div/div[2]/div[2]/div[2]/div[2]/div[1]/div[2]/input').send_keys(self.app.cus_name.get())
                                else:
                                    self.driver.find_element(
                                        By.XPATH, '/html/body/div[1]/div[2]/div[6]/form/div[2]/div/div[7]/div/div[3]/div/div[2]/div[2]/div[2]/div[2]/div[1]/div[2]/input').clear()
                                    self.driver.find_element(
                                        By.XPATH, '/html/body/div[1]/div[2]/div[6]/form/div[2]/div/div[7]/div/div[3]/div/div[2]/div[2]/div[2]/div[2]/div[1]/div[2]/input').send_keys("a")

                            except:
                                print("Auto หน้าท้ายพัง ข้ามไปรอราคาเลย")
                                break

                            if self.app.user_id.get() in self.app.dev_account:
                                try:
                                    print("ใส่ราคาอตัโนมัติ")
                                    print(
                                        (self.app.sum_price + self.app.cus_ship_cost.get()) - self.app.cus_seller_voucher.get())
                                    final_price = (
                                        self.app.sum_price + self.app.cus_ship_cost.get()) - self.app.cus_seller_voucher.get()
                                    if self.app.user_id.get() in self.app.dev_account:
                                        self.driver.find_element(
                                            By.XPATH, '/html/body/div[1]/div[2]/div[6]/form/div[2]/div/div[7]/div/div[3]/div/div[2]/div[2]/div[1]/div[1]/input').clear()
                                        self.driver.find_element(
                                            By.XPATH, '/html/body/div[1]/div[2]/div[6]/form/div[2]/div/div[7]/div/div[3]/div/div[2]/div[2]/div[1]/div[1]/input').send_keys(final_price)
                                except Exception as e:
                                    print("auto_final_price broken", e)

                            # * ค้นหา element โดยใช้ XPath
                            self.is_input_on = self.driver.find_element(
                                By.XPATH, '/html/body/div[1]/div[2]/div[2]/div[2]/div[2]/div[1]/div[2]/div/div/div[6]/form/div/span/span[1]/span/span[1]')

                            # * ดึงข้อความจาก element ที่ค้นหาได้
                            text_value = self.is_input_on.get_attribute(
                                "title")

                            # * พิมพ์ผลลัพธ์
                            print("ตรวจหาชื่อลูกค้า self.is_input_on:", text_value)

                            # # * สำหรับ prefinal  pop-up (optional by ETAX)
                            # # * > แบบเลือกให้ตาม ข้อมูลลูกค้า
                            # while True:

                            #     try:
                            #         self.etax_radio_printout = self.driver.find_element(
                            #             By.XPATH, '/html/body/div[1]/div[2]/div[6]/div[1]/div/div/div[2]/div/div[1]/label/input')
                            #         self.etax_radio_sendmail = self.driver.find_element(
                            #             By.XPATH, '/html/body/div[1]/div[2]/div[6]/div[1]/div/div/div[2]/div/div[2]/label/input')

                            #         if self.etax_radio_printout.is_displayed():
                            #             print("Click Print Out Radio")
                            #             self.etax_radio_printout.click()
                            #             print("Press Print, and break the loop")
                            #             break
                            #         elif self.etax_radio_sendmail.is_displayed():
                            #             print("Click Send Email Radio")
                            #             self.etax_radio_sendmail.click()
                            #             print(
                            #                 "Press Send Email, and break the loop")
                            #             break
                            #     except:
                            #         print("radio has Disappeared")
                            #         continue

                            # # * > แบบเลือกemail เป็น default
                            # ใช้ได้หรือป่าวไม่แน่ใจ
                            # while True:
                            #     self.final_popup = self.driver.find_element(By.XPATH, '/html/body/div[16]/div[2]/button[1]')

                            #     print("Radio while loop")
                            #     if self.final_popup.is_displayed() == True and self.etax_radio_sendmail.is_displayed() == False:
                            #         print("Radio ยังไม่โผล่")
                            #         continue
                            #     elif self.final_popup.is_displayed() == False:
                            #         print("หน้า final หายไป")
                            #         break
                            #     else:
                            #         try:
                            #             self.etax_radio_sendmail = self.driver.find_element(
                            #                 By.XPATH, '/html/body/div[1]/div[2]/div[6]/div[1]/div/div/div[2]/div/div[2]/label/input')
                            #             print("Radio appeared")
                            #             if self.etax_radio_sendmail.is_displayed():
                            #                 is_etax = True
                            #                 print("Click Send Email Radio")
                            #                 self.etax_radio_sendmail.click()
                            #                 print(
                            #                     "Press Send Email, and break the loop")
                            #                 break
                            #             elif not self.etax_radio_sendmail.is_displayed():
                            #                 print("ไม่โชว์ก็ออก")
                            #                 break
                            #         except:
                            #             print("radio has Disappeared")
                            #             break

                            # * สำหรับรอ final pop-up after click the green btn
                            self.app.is_gui_busy.set(False)
                            auto_radio_times = 0
                            while True:

                                time.sleep(1)
                                try:
                                    # print("auto click Before print loop")
                                    self.final_popup = self.driver.find_element(
                                        By.XPATH, '/html/body/div[16]/div[2]/button[1]')
                                    # self.is_final_page = self.wait1.until(EC.invisibility_of_element_located(
                                    #     (By.XPATH, '/html/body/div[1]/div[2]/div[6]/form/div[1]/span[1]')))
                                    self.is_final_page = self.driver.find_element(
                                        By.XPATH, '/html/body/div[1]/div[2]/div[6]/form/div[1]/span[1]')
                                    self.etax_radio_sendmail = self.driver.find_element(
                                        By.XPATH, '/html/body/div[1]/div[2]/div[6]/div[1]/div/div/div[2]/div/div[2]/label/input')
                                    # print("self.is_final_page= ",
                                    #       self.is_final_page)
                                except:
                                    print("Element not found, continuing loop...")
                                    continue

                                if self.final_popup.is_displayed():
                                    pass
                                elif self.is_final_page.is_displayed() == True and self.etax_radio_sendmail.is_displayed() == False:
                                    # print("Radio ยังไม่โผล่")
                                    continue
                                elif self.is_final_page.is_displayed() == False:
                                    print("หน้า final หายไป")
                                    pass
                                else:
                                    try:

                                        print("Radio appeared")
                                        if self.etax_radio_sendmail.is_displayed():
                                            is_etax = True
                                            print("Click Send Email Radio")
                                            if auto_radio_times < 1:
                                                self.etax_radio_sendmail.click()
                                                print(
                                                    "Press Send Email, and break the loop")
                                                auto_radio_times += 1
                                            else:
                                                print("เคยเลือกไปแล้ว")

                                        elif not self.etax_radio_sendmail.is_displayed():
                                            print("ไม่โชว์ก็ออก")

                                    except:
                                        print("radio has Disappeared")

                                if self.final_popup.is_displayed() == True:
                                    self.app.is_gui_busy.set(True)
                                    print("final pop-up has finally displayed!")
                                    try:
                                        self.final_popup_btn = self.wait1.until(EC.element_to_be_clickable(
                                            (By.XPATH, '/html/body/div[16]/div[2]/button[1]')))
                                        # *> ให้เวลาดูเลขบิล 1 วิ
                                        time.sleep(1)

                                        alert_text = self.driver.find_element(
                                            By().XPATH, '/html/body/div[16]/div[2]/div[6]').text
                                        match = re.search(
                                            r'B\d+-W\d+-\d+', alert_text)
                                        print("match: ", match)
                                        # * ถ้าไม่มีบิล, match จะ = none ทำให้ .group() ไม่ได้ แล้ว return error ห
                                        inv_number = match.group()
                                        print("inv_number: ", inv_number)
                                        self.app.update_log(
                                            f'เลขบิล: {inv_number}')

                                        # * สลับไปreprintก่อนแล้วค่อยกลับมากด เพราะมันช้ากรอกรอไว้เลย
                                        # * ไปหน้า Reprint ##########################################################################################
                                        if is_etax and inv_number != "":
                                            self.etax_reprint(inv_number)

                                        self.final_popup_btn.click()
                                        self.wait1.until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[1]/div[2]/div[8]/div/div[2]')))
                                        time.sleep(1)
                                        self.driver.find_element(By.XPATH, '/html/body/div[1]/div[2]/div[8]/div/div[2]')
                                        #* วิธี print แบบเก่า
                                        # self.printtingPage()
                                        # self.justPressP()
                                        #* วิธี print แบบใหม่
                                        self.get_pdf_src_and_print()
                                        
                                        # * Update Accel file //////////////////////
                                        self.app.deduct_accel_file_data(
                                            self.app.cus_order,
                                            self.used_serials
                                        )

                                    except:
                                        # time.sleep(1)
                                        # print("ไม่ได้เลขบิล")
                                        # self.final_popup.click()
                                        self.final_popup_btn.click()
                                        print("พัง ข้ามไปเลยละกัน")

                                    break

                                    # * > รอหน้า canvas โผล่ก่อน
                                    # * >> แบบไม่มีระบบ ETAX มันจะ Process ไปหน้า print มันเลย wait element ของ canvas ได้ แล้วมันจะจบ แค่นี้

                                    #! WIP ต้องเปลี่ยนเป็น while loop แทน เพราะถ้าหาก ขั้นตอนด้านบนเป็น except มันจะรอนาน เพราะใช้ self.wait1
                                    #! ย้ายไปข้างบนแล้ว ถ้าข้างบนใช้ได้ข้างล่างลบทิ้งได้เลย
                                    # self.wait1.until(EC.visibility_of_element_located(
                                    #     (By.XPATH, '/html/body/div[1]/div[2]/div[8]/div/div[2]')))
                                    # time.sleep(1)
                                    # self.driver.find_element(
                                    #     By.XPATH, '/html/body/div[1]/div[2]/div[8]/div/div[2]')
                                    # self.printtingPage()
                                    # self.justPressP()
                                    # break

                                # * >> แบบมี ETAX มันจะ redirect กลับไปหน้าเดิม
                                elif self.is_final_page.is_displayed() == False:
                                    print("End or back")
                                    if bool(re.search(r"\w{5}\-\w{3}-\w{10}", self.driver.find_element(By.XPATH, '/html/body/div[1]/div[2]/div[8]/div/div[1]/span').text)):
                                        print("ไปหน้าสุดท้าย จบ loop")
                                        break
                                    elif self.driver.find_element(By.XPATH, '/html/body/div[1]/div[2]/div[2]/div[2]/div[2]/div[1]/div[2]/div/div/div[1]/form/label') and self.is_input_empty == "":
                                        print("มันจบละ")
                                        break
                                    elif self.driver.find_element(By.XPATH, '/html/body/div[1]/div[2]/div[2]/div[2]/div[2]/div[1]/div[2]/div/div/div[1]/form/label'):
                                        print("กลับมาหน้าเดิม")
                                        break

                                else:
                                    continue

                                    # try:
                                    #     print('จุดจบ')
                                    #     # * กดปุ่มใน pop-up สุดท้าย
                                    #     self.driver.find_element(
                                    #         By.XPATH, '/html/body/div[16]/div[2]/button[1]')
                                    #     self.wait1.until(EC.visibility_of_element_located(
                                    #         (By.XPATH, '/html/body/div[16]/div[2]/button[1]')))
                                    #     self.wait1.until(EC.element_to_be_clickable(
                                    #         (By.XPATH, '/html/body/div[16]/div[2]/button[1]'))).click()
                                    #     # > รอหน้า canvas โผล่ก่อน
                                    #     self.wait1.until(EC.visibility_of_element_located(
                                    #         (By.XPATH, '/html/body/div[1]/div[2]/div[8]/div/div[2]')))
                                    #     self.printtingPage()
                                    #     break
                                    # except Exception as err:
                                    #     print('ไม่ใช่จุดจบ', err)
                                    #     pass

                            # * ต้องใช้จริงๆเหรอ?
                            # if self.final_popup.is_displayed() == True:
                            #     break
                            # else:
                            #     break
                                # break ที่แก้เป็น break ดูเหมือน code ด้านบนมันจะผิด และไม่สามารถรับมือกับเหตุการณืแบบ dynamic ได้ ทำให้ continue ตรงนี้ทำงานอย่างผิดปกติ แต่ตอนนี้แก้ถูกแล้ว

                            # * ไม่แน่ใจ
                            continue
                        else:
                            print("จบสูตร")
                        self.autofinal = False
                        break

                    print("Whileหลัก ถ้ามาถึงนี่แปลว่าต้องเริ่มใหม่")
                    continue
                break

            print("จบ auto_last_page")
            self.autofinal = False
            # self.driver.quit()

        else:
            print("ไม่มีOrder ไม่รู้จะทำอะไร")

        # # * จงใจให้ Error จงใจ Error // มันจะได้จบๆไป
        # if self.app.accel_search_thread:
        #     self.app.accel_timer = threading.Timer(
        #         0.2, self.app.on_accel_thread_done)
        #     self.app.accel_timer.start()
        # else:
        #     print("'MyApp' object has no attribute 'accel_search_thread'")
        #     pass

    def addNormalCustomer(self, cusname_fixed):
        is_functionworking = False
        is_functionworking = True
        while is_functionworking:
            self.driver.switch_to.window(
                self.merged_dict['SMCO :: เปิดการขาย1'])

            self.element = self.driver.find_element(
                By.XPATH, self.app.cusSearchSMCO)
            self.element.click()  # * กดแว่นขยาย
            self.btnElement = self.wait1.until(
                EC.element_to_be_clickable((By.XPATH, self.app.cusCreateBtn)))
            time.sleep(0.65)
            self.btnElement.click()  # * create

            # * > เลือกหมวดลูกค้า  เพิ่มมาตอน 6.3.1 24/04/2024
            try:
                self.driver.find_element(
                    By.XPATH, '/html/body/div[1]/div[2]/div[11]/div/div/div[3]/div/form/div[1]/div[3]/div/span/span[1]/span/span[1]').click()

                # * บางจังหวะ มันไม่ขึ้น "CM1-Domestic Customer" แล้วมันข้ามไปใส่ชื่อเลย แล้วมันจะไปต่อไม่ได้เพราะ CM1-Domestic Customer ไม่ได้ถูกใส่
                while True:
                    try:
                        choice_found = self.driver.find_element(
                            By.XPATH, "/html/body/div[1]/div[2]/div[11]/span/span/span[2]/ul/li").text
                        print("choice_found: ", choice_found)
                        if self.driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div[11]/span/span/span[2]/ul/li").text == "CM1-Domestic Customer":
                            break
                    except Exception as err:
                        time.sleep(1)
                        # print("except: ", err) # for develop inspection
                        continue

                self.driver.find_element(
                    By.XPATH, '/html/body/div[1]/div[2]/div[11]/span/span/span[1]/input').clear()
                time.sleep(0.75)
                self.driver.find_element(
                    By.XPATH, '/html/body/div[1]/div[2]/div[11]/span/span/span[2]/ul/li').click()
            except:
                print("No customer category, Pass")

            self.driver.find_element(
                By.XPATH, '/html/body/div[1]/div[2]/div[11]/div/div/div[3]/div/form/div[3]/div[1]/input').clear()
            self.driver.find_element(
                By.XPATH, '/html/body/div[1]/div[2]/div[11]/div/div/div[3]/div/form/div[3]/div[1]/input').send_keys(cusname_fixed)

            self.driver.find_element(
                By.XPATH, '/html/body/div[1]/div[2]/div[11]/div/div/div[3]/div/form/div[3]/div[2]/input').clear()
            self.driver.find_element(
                By.XPATH, '/html/body/div[1]/div[2]/div[11]/div/div/div[3]/div/form/div[3]/div[2]/input').send_keys(cusname_fixed)

            self.driver.find_element(
                By.XPATH, '/html/body/div[1]/div[2]/div[11]/div/div/div[3]/div/form/div[7]/div/textarea').clear()
            self.driver.find_element(
                By.XPATH, '/html/body/div[1]/div[2]/div[11]/div/div/div[3]/div/form/div[7]/div/textarea').send_keys(self.app.cus_address)

            self.driver.find_element(
                By.XPATH, '/html/body/div[1]/div[2]/div[11]/div/div/div[3]/div/form/div[14]/div[2]/input').clear()
            self.driver.find_element(
                By.XPATH, '/html/body/div[1]/div[2]/div[11]/div/div/div[3]/div/form/div[14]/div[2]/input').send_keys(1)

            self.driver.find_element(
                By.XPATH, '/html/body/div[1]/div[2]/div[11]/div/div/div[3]/div/form/div[16]/center/button[1]').click()

            # รอมันหายก่อน
            self.wait1.until(EC.invisibility_of_element_located(
                (By.XPATH, '/html/body/div[1]/div[2]/div[11]/div/div/div[3]/div/form/div[16]/center/button[1]')))
            is_functionworking = False

            # *  24/04/2023: กลับมาอีกแล้วทำให้เป็น try except ละกัน// 09/11/2023: partนี้ ทาง SMCO ลบออกไปแล้ว
            # self.wait1.until(EC.visibility_of_element_located(
            #     (By.XPATH, '/html/body/div[16]/div[2]/button[1]')))
            try:
                self.driver.find_element(
                    By.XPATH, '/html/body/div[16]/div[2]/button[1]').click()
            except:
                pass

    def addressExtractor(self, cusAddress):
        self.splited = cusAddress.split(",")
        return (self.splited)

    def addTaxInvCustomer(self):
        print("ชื่อลูกค้าเป็นไง SHOP: ", self.app.cus_name.get())
        name = self.app.cus_name.get()
        # * เติมสาขาให้เรียบร้อย
        if self.app.branch_type == 'สำนักงานใหญ่':
            # ! self.app.tax_branch.set(self.app.nondistortedData['ประเภทสาขา']) ใช้ทำไมวะรอลบ
            # ! name = f"""{name} ({self.app.tax_branch.get()})""" รอลบ
            name = f"""{name} ({self.app.branch_type})"""
        elif self.app.branch_type == "สาขาย่อย" and not pd.isna(self.app.data_frame[self.app.target_row]['รหัสประจำสาขา'].iloc[0]):
            name = f"""{name} (สาขา{self.app.tax_branch_num.get()})"""

        self.driver.switch_to.window(
            self.merged_dict['SMCO :: เปิดการขาย1'])
        self.driver.find_element(By.XPATH, self.app.cusSearchSMCO).click()
        time.sleep(0.75)
        self.driver.find_element(By.XPATH, self.app.cusCreateBtn).click()

        # * > เลือกหมวดลูกค้า  เพิ่มมาตอน 6.3.1 24/04/2024
        try:
            self.driver.find_element(
                By.XPATH, '/html/body/div[1]/div[2]/div[11]/div/div/div[3]/div/form/div[1]/div[3]/div/span/span[1]/span/span[1]').click()
            self.driver.find_element(
                By.XPATH, '/html/body/div[1]/div[2]/div[11]/span/span/span[1]/input').clear()
            time.sleep(0.75)
            self.driver.find_element(
                By.XPATH, '/html/body/div[1]/div[2]/div[11]/span/span/span[2]/ul/li').click()
        except:
            print("No customer category, Pass")

        # * > nameTH clear
        self.driver.find_element(
            By.XPATH, '/html/body/div[1]/div[2]/div[11]/div/div/div[3]/div/form/div[3]/div[1]/input').clear()
        # # * >nameTH fill input better style ปิดการใช้งาน
        # self.driver.find_element( By.XPATH, '/html/body/div[1]/div[2]/div[11]/div/div/div[3]/div/form/div[3]/div[1]/input').send_keys(f'{name} Tax ID: {self.app.tax_num.get()}')
        # * >nameTH SMCO style
        self.driver.find_element(
            By.XPATH, '/html/body/div[1]/div[2]/div[11]/div/div/div[3]/div/form/div[3]/div[1]/input').send_keys(f'{name}')

        # * >nameEN clear
        self.driver.find_element(
            By.XPATH, '/html/body/div[1]/div[2]/div[11]/div/div/div[3]/div/form/div[3]/div[2]/input').clear()
        # * >nameEN fill input better style ปิดการใช้งาน
        # self.driver.find_element(By.XPATH, '/html/body/div[1]/div[2]/div[11]/div/div/div[3]/div/form/div[3]/div[2]/input').send_keys(f'{name} Tax ID: {self.app.tax_num.get()}')
        # * >nameEN SMCO style
        self.driver.find_element(
            By.XPATH, '/html/body/div[1]/div[2]/div[11]/div/div/div[3]/div/form/div[3]/div[2]/input').send_keys(f'{name}')

        # ! เปิดใช้การออกใบกำกับ ตาม SMCO style (ถ้าไม่เปิดจะถือว่าเป็นการใช้ Better style)
        # * clear Identity ID
        self.driver.find_element(
            By.XPATH, '/html/body/div[1]/div[2]/div[11]/div/div/div[3]/div/form/div[3]/div[3]/input').clear()
        # * Identity ID
        self.driver.find_element(
            By.XPATH, '/html/body/div[1]/div[2]/div[11]/div/div/div[3]/div/form/div[3]/div[3]/input').send_keys(self.app.tax_num.get())

        # * กรอก Address
        self.driver.find_element(
            By.XPATH, '/html/body/div[1]/div[2]/div[11]/div/div/div[3]/div/form/div[7]/div/textarea').clear()
        # ! > การกรอก address แบบโกง bypass เขตแขวง SMCO แต่กลัวว่า สรรพากรจะกำหมัด
        # self.driver.find_element(By.XPATH, '/html/body/div[1]/div[2]/div[11]/div/div/div[3]/div/form/div[7]/div/textarea').send_keys(self.app.cus_address)
        # ! > การกรอก address แบบทำตามกฎเลือก เขตแขวง ตามระบบ SMCO แต่กลัวว่า สรรพากรจะกำหมัด
        address = self.app.address
        if self.app.tax_bool.get():
            address = self.app.get_pure_address(self.app.address)
        self.driver.find_element(
            By.XPATH, '/html/body/div[1]/div[2]/div[11]/div/div/div[3]/div/form/div[7]/div/textarea').send_keys(address)

        # * กรอก email
        self.email_input = self.driver.find_element(
            By.XPATH, '/html/body/div[1]/div[2]/div[11]/div/div/div[3]/div/form/div[13]/div[2]/input')
        self.email_input.clear()
        self.email_input.send_keys(self.app.cus_email.get())

        # * tel.
        self.driver.find_element(
            By.XPATH, '/html/body/div[1]/div[2]/div[11]/div/div/div[3]/div/form/div[14]/div[2]/input').clear()
        self.driver.find_element(
            By.XPATH, '/html/body/div[1]/div[2]/div[11]/div/div/div[3]/div/form/div[14]/div[2]/input').send_keys(self.app.cus_tel.get())

        ### * เป็นแบบกรอกแบบ DropDown ##########################################################################################################
        # dropdown Country
        self.driver.find_element(
            By.XPATH, '/html/body/div[1]/div[2]/div[11]/div/div/div[3]/div/form/div[9]/div[1]/div/span/span[1]/span/span[1]').click()
        time.sleep(1)
        # select thailand in dropdown
        self.driver.find_element(
            By.XPATH, '/html/body/div[1]/div[2]/div[11]/span/span/span[2]/ul/li[2]').click()

        # province dropdown
        self.driver.find_element(
            By.XPATH, '/html/body/div[1]/div[2]/div[11]/div/div/div[3]/div/form/div[9]/div[2]/div/span/span[1]/span/span[1]').click()
        self.driver.find_element(
            # province input
            By.XPATH, '/html/body/div[1]/div[2]/div[11]/span/span/span[1]/input').clear()
        self.driver.find_element(By.XPATH, '/html/body/div[1]/div[2]/div[11]/span/span/span[1]/input').send_keys(
            self.app.cus_province.get().replace("จังหวัด", ""))  # province input
        time.sleep(1.75)
        self.driver.find_element(
            By.XPATH, '/html/body/div[1]/div[2]/div[11]/span/span/span[1]/input').send_keys(Keys().ENTER)

        self.driver.find_element(
            # District drop
            By.XPATH, '/html/body/div[1]/div[2]/div[11]/div/div/div[3]/div/form/div[11]/div[1]/div/span/span[1]/span/span[1]').click()
        self.driver.find_element(
            # District
            By.XPATH, '/html/body/div[1]/div[2]/div[11]/span/span/span[1]/input').clear()
        self.driver.find_element(By.XPATH, '/html/body/div[1]/div[2]/div[11]/span/span/span[1]/input').send_keys(
            self.app.cus_district.get().replace("อำเภอ", "").replace("เขต", "").replace("ต.", ""))  # District
        time.sleep(1.75)
        self.driver.find_element(
            By.XPATH, '/html/body/div[1]/div[2]/div[11]/span/span/span[1]/input').send_keys(Keys().ENTER)

        # SubDistrict drop
        self.driver.find_element(
            By.XPATH, '/html/body/div[1]/div[2]/div[11]/div/div/div[3]/div/form/div[11]/div[3]/div/span/span[1]/span/span[1]').click()
        self.driver.find_element(
            By.XPATH, '/html/body/div[1]/div[2]/div[11]/div/div/div[3]/div/form/div[11]/div[3]/div/span/span[1]/span/span[1]').click()
        self.driver.find_element(
            By.XPATH, '/html/body/div[1]/div[2]/div[11]/div/div/div[3]/div/form/div[11]/div[3]/div/span/span[1]/span/span[1]').click()
        # SubDistrict
        self.driver.find_element(
            By.XPATH, '/html/body/div[1]/div[2]/div[11]/span/span/span[1]/input').clear()
        self.driver.find_element(By.XPATH, '/html/body/div[1]/div[2]/div[11]/span/span/span[1]/input').send_keys(
            self.app.cus_sub_district.get().replace("ตำบล", "").replace("แขวง", "").replace("ต.", ""))  # SubDistrict
        time.sleep(1.75)
        self.driver.find_element(
            By.XPATH, '/html/body/div[1]/div[2]/div[11]/span/span/span[1]/input').send_keys(Keys().ENTER)

        # # * กด Save
        # self.driver.find_element(
        #     By.XPATH, '/html/body/div[1]/div[2]/div[11]/div/div/div[3]/div/form/div[16]/center/button[1]').click()

        # รอมันหายก่อนแล้วค่อยจบ function เพื่อไม่ให้ขั้นตอนต่อไปทำงานเร็วเกินไป
        self.wait1.until(EC.invisibility_of_element_located(
            (By.XPATH, '/html/body/div[1]/div[2]/div[11]/div/div/div[3]/div/form/div[16]/center/button[1]')))

        # *  24/04/2023: กลับมาอีกแล้วทำให้เป็น try except ละกัน// 09/11/2023: partนี้ ทาง SMCO ลบออกไปแล้ว
        # self.wait1.until(EC.visibility_of_element_located(
        #     (By.XPATH, '/html/body/div[16]/div[2]/button[1]')))
        try:
            self.driver.find_element(
                By.XPATH, '/html/body/div[16]/div[2]/button[1]').click()
        except:
            pass

    def addTaxInvCustomerLaz(self):
        tax_info = self.get_vatinfo_data(
            self.app.tax_num.get(),
            self.app.tax_branch.get()
        )

        print("ชื่อลูกค้าเป็นไง LAZ: ", f"{tax_info['name']}")
        name = f"{tax_info['name']}"

        # * เติมสาขาให้เรียบร้อย
        if self.app.branch_type == 'สำนักงานใหญ่':
            self.app.tax_branch.set(self.app.nondistortedData['ประเภทสาขา'])
            name = name
            if name.startswith("บริษัท") or "จำกัด" in name:
                name += f" {tax_info['branch']}"
        elif self.app.branch_type == "สาขาย่อย" and not pd.isna(self.app.data_frame[self.app.target_row]['รหัสประจำสาขา'].iloc[0]):
            name = f"""{name} (สาขา{self.app.tax_branch.get()})"""

        self.driver.switch_to.window(
            self.merged_dict['SMCO :: เปิดการขาย1'])
        self.driver.find_element(By.XPATH, self.app.cusSearchSMCO).click()
        time.sleep(0.75)
        self.driver.find_element(By.XPATH, self.app.cusCreateBtn).click()

        # * > เลือกหมวดลูกค้า  เพิ่มมาตอน 6.3.1 24/04/2024
        try:
            self.driver.find_element(
                By.XPATH, '/html/body/div[1]/div[2]/div[11]/div/div/div[3]/div/form/div[1]/div[3]/div/span/span[1]/span/span[1]').click()
            self.driver.find_element(
                By.XPATH, '/html/body/div[1]/div[2]/div[11]/span/span/span[1]/input').clear()
            time.sleep(0.75)
            self.driver.find_element(
                By.XPATH, '/html/body/div[1]/div[2]/div[11]/span/span/span[2]/ul/li').click()
        except:
            print("No customer category, Pass")

        # * clear nameTH
        self.driver.find_element(
            By.XPATH, '/html/body/div[1]/div[2]/div[11]/div/div/div[3]/div/form/div[3]/div[1]/input').clear()

        # * nameTH
        self.driver.find_element(
            By.XPATH, '/html/body/div[1]/div[2]/div[11]/div/div/div[3]/div/form/div[3]/div[1]/input').send_keys(name)

        # * clear nameEN
        self.driver.find_element(
            By.XPATH, '/html/body/div[1]/div[2]/div[11]/div/div/div[3]/div/form/div[3]/div[2]/input').clear()
        # * nameEN
        self.driver.find_element(
            By.XPATH, '/html/body/div[1]/div[2]/div[11]/div/div/div[3]/div/form/div[3]/div[2]/input').send_keys(name)

        # * clear Identity ID
        self.driver.find_element(
            By.XPATH, '/html/body/div[1]/div[2]/div[11]/div/div/div[3]/div/form/div[3]/div[3]/input').clear()
        # * Identity ID
        self.driver.find_element(
            By.XPATH, '/html/body/div[1]/div[2]/div[11]/div/div/div[3]/div/form/div[3]/div[3]/input').send_keys(tax_info['tax_num'])

        # * กรอก Address

        self.driver.find_element(
            By.XPATH, '/html/body/div[1]/div[2]/div[11]/div/div/div[3]/div/form/div[7]/div/textarea').clear()
        self.driver.find_element(
            By.XPATH, '/html/body/div[1]/div[2]/div[11]/div/div/div[3]/div/form/div[7]/div/textarea').send_keys(tax_info['address_shortened'])

        # * กรอก email
        self.email_input = self.driver.find_element(
            By.XPATH, '/html/body/div[1]/div[2]/div[11]/div/div/div[3]/div/form/div[13]/div[2]/input')
        self.email_input.clear()
        self.email_input.send_keys(self.app.cus_email.get())

        ### * เป็นแบบกรอกแบบ DropDown ##########################################################################################################
        # * dropdown Country
        self.driver.find_element(
            By.XPATH, '/html/body/div[1]/div[2]/div[11]/div/div/div[3]/div/form/div[9]/div[1]/div/span/span[1]/span/span[1]').click()
        time.sleep(1)
        # * > select thailand in dropdown
        self.driver.find_element(
            By.XPATH, '/html/body/div[1]/div[2]/div[11]/span/span/span[2]/ul/li[2]').click()

        # * province dropdown
        self.driver.find_element(
            By.XPATH, '/html/body/div[1]/div[2]/div[11]/div/div/div[3]/div/form/div[9]/div[2]/div/span/span[1]/span/span[1]').click()
        self.driver.find_element(
            # * > province input
            By.XPATH, '/html/body/div[1]/div[2]/div[11]/span/span/span[1]/input').clear()
        self.driver.find_element(By.XPATH, '/html/body/div[1]/div[2]/div[11]/span/span/span[1]/input').send_keys(
            tax_info['province'].replace("จังหวัด", ""))  # province input
        time.sleep(1.55)
        self.driver.find_element(
            By.XPATH, '/html/body/div[1]/div[2]/div[11]/span/span/span[1]/input').send_keys(Keys().ENTER)

        # *> District drop
        self.driver.find_element(
            By.XPATH, '/html/body/div[1]/div[2]/div[11]/div/div/div[3]/div/form/div[11]/div[1]/div/span/span[1]/span/span[1]').click()
        # *> District clear
        self.driver.find_element(
            By.XPATH, '/html/body/div[1]/div[2]/div[11]/span/span/span[1]/input').clear()
        # *> District fill input
        self.driver.find_element(By.XPATH, '/html/body/div[1]/div[2]/div[11]/span/span/span[1]/input').send_keys(
            tax_info['district'].replace("อำเภอ", "").replace("เขต", "").replace("ต.", ""))
        time.sleep(1.55)
        # *> District Enter to submit District from the dropdown
        self.driver.find_element(
            By.XPATH, '/html/body/div[1]/div[2]/div[11]/span/span/span[1]/input').send_keys(Keys().ENTER)

        # * SubDistrict drop
        self.driver.find_element(
            By.XPATH, '/html/body/div[1]/div[2]/div[11]/div/div/div[3]/div/form/div[11]/div[3]/div/span/span[1]/span/span[1]').click()
        self.driver.find_element(
            # > SubDistrict
            By.XPATH, '/html/body/div[1]/div[2]/div[11]/span/span/span[1]/input').clear()
        self.driver.find_element(By.XPATH, '/html/body/div[1]/div[2]/div[11]/span/span/span[1]/input').send_keys(
            # > SubDistrict
            tax_info['sub_district'].replace("ตำบล", "").replace("แขวง", "").replace("ต.", ""))
        time.sleep(1.55)
        self.driver.find_element(
            By.XPATH, '/html/body/div[1]/div[2]/div[11]/span/span/span[1]/input').send_keys(Keys().ENTER)

        # * tel.
        self.driver.find_element(
            By.XPATH, '/html/body/div[1]/div[2]/div[11]/div/div/div[3]/div/form/div[14]/div[2]/input').clear()
        self.driver.find_element(
            By.XPATH, '/html/body/div[1]/div[2]/div[11]/div/div/div[3]/div/form/div[14]/div[2]/input').send_keys(self.app.cus_tel.get())

        # # * กด Save
        # self.driver.find_element(
        #     By.XPATH, '/html/body/div[1]/div[2]/div[11]/div/div/div[3]/div/form/div[16]/center/button[1]').click()

        # #* รอมันหายก่อนแล้วค่อยจบ function เพื่อไม่ให้ขั้นตอนต่อไปทำงานเร็วเกินไป ใช้ได้
        self.wait1.until(EC.invisibility_of_element_located(
            (By.XPATH, '/html/body/div[1]/div[2]/div[11]/div/div/div[3]/div/form/div[16]/center/button[1]')))

        # *  24/04/2023: กลับมาอีกแล้วทำให้เป็น try except ละกัน// 09/11/2023: partนี้ ทาง SMCO ลบออกไปแล้ว
        # self.wait1.until(EC.visibility_of_element_located(
        #     (By.XPATH, '/html/body/div[16]/div[2]/button[1]')))
        try:
            self.driver.find_element(
                By.XPATH, '/html/body/div[16]/div[2]/button[1]').click()
        except:
            pass

    # * function แยก address ของ output จาก vatinfo ให้เป็น part ย่อย (เขต, แขวง)

    def classify_vatinfo_address(self, input):
        try:
            # Create a copy of the output dictionary
            result = input
            print("resultสำหรับ classify คือไร :", result)

            # Remove the "ตำบล" and everything after it from the address
            address_only = re.compile(r'(?:ตำบล|ต\.).*')
            result['address_shortened'] = address_only.sub(
                '', result['address']).strip()

            # Define the regular expression pattern
            pattern = re.compile(
                r'ตำบล/แขวง\s+(\S+).*?เขต\s+(\S+).*?จังหวัด\s+(\S+)')

            # Use the pattern to find matches in the address
            matches = pattern.search(result['address'])

            # Extract the matched groups
            if matches:
                result['sub_district'] = matches.group(1)
                result['district'] = matches.group(2)
                result['province'] = matches.group(3)
            else:
                result['sub_district'] = None
                result['district'] = None
                result['province'] = None

            # Add a space after the word "บริษัท" in the company name
            result['name'] = re.sub(r'(บริษัท)\s*', r'\1 ', result['name'])

            return result
        except:
            print("Input is empty. Cannot be classify, result = ", result)
            return result

    def get_res_vatinfo(self, tax_num, tax_branch):
        tax_input = str(tax_num)
        branch = str(tax_branch)
        jsession_id = ''

        # เราจะไม่ใช้ cookies แต่จะใช้ค่าจาก class แรกสุด เพราะ
        # cookies = self.app.cookies['vatinfo']
        print("cookies for reqtaxinfo: ", self.app.cookies['vatinfo'])

        headers = {
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
            'Accept-Language': 'en-US,en;q=0.9',
            'Cache-Control': 'max-age=0',
            'Connection': 'keep-alive',
            'Content-Type': 'application/x-www-form-urlencoded',
            # 'Cookie': 'JSESSIONID=0000afl1mgz_VGdxFmh7f5mQJqf:-1',
            'Origin': 'https://vsreg.rd.go.th',
            'Referer': 'https://vsreg.rd.go.th/VATINFOWSWeb/jsp/VATInfoWSServlet',
            'Sec-Fetch-Dest': 'document',
            'Sec-Fetch-Mode': 'navigate',
            'Sec-Fetch-Site': 'same-origin',
            'Upgrade-Insecure-Requests': '1',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36',
            'sec-ch-ua': '"Google Chrome";v="119", "Chromium";v="119", "Not?A_Brand";v="24"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
        }

        params = ''

        data = {
            'operation': 'searchByTin',
            'goto_page': '',
            'tin': 'on',
            'txtTin': tax_input,
            'branotxt': '',
            'fname': 'null',
            'lname': 'null',
        }

        times = 1

        data2 = {
            'operation': 'GotoPage_Click',
            'goto_page': f'{times}',
            'tin': 'on',
            'txtTin': tax_input,
            'branotxt': '',
            'fname': 'null',
            'lname': 'null',
        }

        while True:
            if times == 1:
                print("times = 1")
                response = session.post(
                    'https://vsreg.rd.go.th/VATINFOWSWeb/jsp/VATInfoWSServlet', cookies=self.app.cookies['vatinfo'], params=params, headers=headers, data=data)

                # Todo มันมีการตรวจสอบ cookies ตลอดเวลา แต่ครั้งแรกreqไปมันจะตรวจสอบก่อน ถ้าไม่มีมันจะ return มาให้  ครั้งถัดไปมันจะตรวจอีกถ้ามี"แล้วยังใช้ได้" มันจะไม่ return ให้ ถ้าใช้ไม่ได้มันจะ return ตัวใหม่ให้
                try:
                    # * กรณี ที่ มี cookies returns กลับมา เพราะอันเก่ามันหมดอายุแล้ว หรือไม่เคยมีมาก่อน
                    print("response cookies ไรมา", response.cookies)
                    # * > เก็บค่า cookies จาก response เข้าไปใน cookies ที่มีอยู่แล้ว
                    jsession_id = response.cookies['JSESSIONID']
                    print(
                        "we never have usable cookies before that why the response has cookies. We'll use it like a state in app.cookies")
                    self.app.cookies['vatinfo']['JSESSIONID'] = f"""{
                        jsession_id}"""
                except Exception as err:
                    # * กรณี ที่ ไม่มี cookies returns กลับมา เพราะอันเก่าใช้ได้อยู่ ใช้ cookies เดิมได้เลย
                    print("if the response is '<RequestCookieJar[]>', it indicates that no cookies were returned. Therefore, we already have available cookies now.", response)

            elif times > 1:
                print("jsession_id", jsession_id)
                # รอบสองเราเอา cookies มาประกอบ request โดย data ที่ใช้ request รอบนี้เป็นอีกแบบนึงจะต้องมี cookie เป็นตัวยืนยันว่าเคย login มาแล้ว ถ้าไม่มี cookie จะผ่านไม่ได้ เหมือนจะเป็น authen

                data2['goto_page'] = f'{times}'
                response = session.post('https://vsreg.rd.go.th/VATINFOWSWeb/jsp/VATInfoWSServlet?',
                                        params=params, cookies=self.app.cookies['vatinfo'], headers=headers, data=data2)

            try:
                response.raise_for_status()
                soup = BeautifulSoup(response.content, 'html.parser')
                # print("ได้ไรออกมา", soup)
                # หาว่า response มี <tr> หรือไม่ มีเท่าไหร่
                menu_elements = soup.select('tr[class^="trMenu"]')
                is_many_page = soup.select("""span[onclick^="gotoPage('"]""")
                print("มีหลายหน้า?: ", bool(is_many_page))
                search_result = []
                output = ""

                # * ตรวจหา element รายการข้อมูลใบกำกับ ซึ่งมันจะมี class ชื่อ trmenu
                if len(menu_elements):
                    # * มี <tr>
                    for menu_element in menu_elements:
                        result_data = {
                            "no": "",
                            "tax_num": "",
                            "branch": "",
                            "name": "",
                            "address": "",
                            "postal_code": ""
                        }

                        # print(menu_element) <<หาทั้งหมด
                        # * tr = menu_element.find('tr')
                        # * ในแต่ละ <tr> มี <td> หลายอัน
                        tds = menu_element.find_all('td')
                        for idx, key in enumerate(result_data):
                            b = tds[idx].find('b')
                            result = b.find('font').text.strip()
                            result = re.sub("\s{2,}", " ", result)

                            # * ช่วงใบกำกับ จะตัดเอาค่า 13 หลักจากด้านหลัง เพราะไอ 10 หลักตอนแรกมันคือไรไม่รู้
                            if idx == 1 and len(result) > 13:
                                result = result[-13:]

                            print(result)
                            result_data[key] = result
                        print(" ")
                        search_result.append(result_data)

                    # * เอา search_result มาดูว่าตรงกับสาขาที่ต้องการหรือไม่
                    for item in search_result:
                        if item['branch'] == self.app.branch_type:
                            output = item
                            print("เกบค่าลง dict result ลง output", output)
                            break
                    if bool(output) == False:
                        print("ว่างต้องวนใหม่")
                        times += 1
                        continue
                    else:
                        print("ใช้ได้", output)
                        break

                elif bool(menu_elements) == False:
                    # ไม่มี <tr>
                    print("ไม่มีใบกำกับจาก request", output)
                    break

            except session.exceptions.HTTPError as e:
                print(f"HTTP Error occurred: {e}")
            except Exception as e:
                print(f"An error occured: {e}")
            break

        output = self.classify_vatinfo_address(output)
        print("output: ", output)
        return output

    def google_for_tambon(self, address, possible_tambons):
        # Todo address รับค่าเป็น dict
        # Todo possible_tambons รับค่าเป็น list

        input = f"{address['cleaned_address']} {address['amphoe']} {
            address['province']} {address['postal']}"

        session = requests.Session()

        params = {
            'q': f'{input}',
            'oq': f'{input}',
        }
        response = session.get('https://www.google.com/search', params=params)
        soup = BeautifulSoup(response.content, 'html.parser')
        # print("soup type:", type(soup))
        # print("soupหน้าตาเปนไง: ", soup)

        found_tambon = {}
        # * การหาแบบ alltag มาจาก https://www.skytowner.com/explore/finding_elements_that_contain_a_specific_text_in_beautiful_soup#:~:text=To%20find%20elements%20that%20contain,together%20with%20a%20lambda%20function.
        try:
            for possible_tambon in possible_tambons:
                found_tambon[f'{possible_tambon}'] = 0
                matched_tags = soup.find_all(lambda tag: len(
                    tag.find_all()) == 0 and possible_tambon in tag.text)
                for matched_tag in matched_tags:
                    # print("found_the_possible_tambon", possible_tambon,"matched_tag: ", matched_tag)
                    found_tambon[f'{possible_tambon}'] += 1
                # print(found_tambon)
                most_tambon = max(found_tambon, key=found_tambon.get)
            print("คนที่คะแนนเยอะสุด", most_tambon)
            return most_tambon
        except:
            print('ไม่มี element')
            return possible_tambons[0]

    def resource_path(self, relative_path):
        try:
            base_path = sys._MEIPASS
        except Exception:
            base_path = os.path.abspath(".")
        return os.path.join(base_path, relative_path)

    def assign_address(self, df, order):
        # * function ใช้สำหรับลูกค้าขอใบกำกับ เพราะมันต้องย้ายค่าตำบล ออกไปใส่ใบกำกับ
        print("assign_address order:", order)
        # เตรียมข้อมูล Pattern ที่อยู่คนไทย
        df['ที่อยู่สำหรับออกใบกำกับภาษีแบบเต็มรูป'] = df['ที่อยู่สำหรับออกใบกำกับภาษีแบบเต็มรูป'].astype(
            str)
        df['หมายเลขคำสั่งซื้อ'] = df['หมายเลขคำสั่งซื้อ'].astype(
            str)

        # * หาตำบลจากไฟล์
        target_row_index = df['หมายเลขคำสั่งซื้อ'] == order
        if any(target_row_index) == True:
            print("เจอ Order ใน ไฟล์")
            cus_address = df[target_row_index]['ที่อยู่สำหรับออกใบกำกับภาษีแบบเต็มรูป'].iloc[0]
            print("cus_address", cus_address)
            amphoe = str(df[target_row_index]['เขต/อำเภอ.1'].iloc[0])
            amphoe_short = amphoe.replace("อำเภอ", "").replace("เขต", "")
            province = str(df[target_row_index]['จังหวัด.1'].iloc[0])
            print("amphoe", amphoe)
            print("amphoe_short", amphoe_short)
            postal_code = str(df[target_row_index]['รหัสไปรษณีย์.1'].iloc[0])
            is_alert = False

            # เอาข้อมูลลูกค้ามาเทียบกับตาราง Pattern ที่อยู่คนไทย
            # จัวนี้ต้องผูกกับ exe
            # tambon_data_address = r'test\tkinter_test\Addresscleaner_TambonData.xlsx'

            tambon_data_address = self.resource_path(
                "Addresscleaner_TambonData.xlsx")
            df_thai_addr = pd.read_excel(tambon_data_address)
            allfiltered_df = df_thai_addr[(df_thai_addr['PostCodeMain'].astype(
                str) == postal_code) & (df_thai_addr['DistrictThaiShort'] == amphoe_short)]
            possible_tambons = list(allfiltered_df['TambonThaiShort'])
            possible_tambons.sort(key=len, reverse=True)

            print("ตำบลที่เป็นไปได้: ", possible_tambons)

            ##
            decent_tambon = []
            for tambon in possible_tambons:
                # * เขต แขวง อ ต ไรก็ตามเอาออกให้หมด

                if "ตำบล" in tambon:
                    tambon = re.sub(r'\bตำบล\b', '', tambon)
                elif "ต." in tambon:
                    tambon = re.sub(r'\bต.\b', '', tambon)
                elif "แขวง" in tambon:
                    tambon = re.sub(r'แขวง', '', tambon)
                # * ช่องว่างตั้งแต่ 1 อันขึ้นไป จะกลายเป็น โดนลบทั้งหมด
                tambon = re.sub(r'\s{1,}', '', tambon)

                # decent_tambon.append(tambon) เลิกใช้ list แล้ว เก็บค่าตรงๆไปเลย
                if tambon in cus_address:
                    print("ทำไมได้ตลาดยอดวะ", tambon)
                    decent_tambon = tambon
                    break

            # *ลบตำบลออก
            cleaned_address = self.app.get_pure_address(cus_address)

            # * หาค่าตัวแปรที่เหมาะสมลงใน decent_tambon
            if decent_tambon:
                # * เจอตำบลในไฟล์
                print("เลือกตำบลที่เหมาะสมมาแล้ว", decent_tambon)

            else:
                # * ตำบลในไฟล์ไม่มีต้อง Google เอา
                print("ไม่มีตำบลมาให้ต้อง search google")
                address_dict = {"cleaned_address": cleaned_address, "decent_tambon": decent_tambon,
                                "amphoe": amphoe_short, "province": province, "postal": postal_code}
                googled_tambon = self.google_for_tambon(
                    address_dict, possible_tambons)
                decent_tambon = googled_tambon
                is_alert = True
                PopUp(
                    "Caution!!", f""""ตำบล"อันนี้มั่วมาโปรดตรวจสอบก่อนออกบิล""", self.parent, "alert")

            # * บางคนไม่ใส่ ตำบล ต แขวง ต้องรู้ ชื่อตำบลก่อนค่อยลบ
            print("ก่อนลบ", cleaned_address)
            #! ตรงนี้ผิด กุลบทำไมวะ
            # prog = re.compile(fr'{re.escape(decent_tambon)}.*')
            # cleaned_address = prog.sub('', cleaned_address)
            # print("ลบไม่ได้วะ", cleaned_address)

            # * เลือกว่าจะ ตำบล หรหือ แขวง
            if decent_tambon in cus_address and ("กรุงเทพ" in cus_address or "กทม" in cus_address):
                decent_tambon = "แขวง" + tambon
            elif decent_tambon in cus_address:
                decent_tambon = "ตำบล" + tambon

            return {"cleaned_address": cleaned_address, "decent_tambon": decent_tambon, "amphoe": amphoe_short, "province": province, "postal": postal_code, "alert": is_alert}

        # * หาตำบลจากไฟล์ไม่เจอ
        elif any(target_row_index) == False:
            print("ไม่เจอOrder")

    def get_vatinfo_data(self, tax_num, branch):
        if branch == "":
            branch = "สำนักงานใหญ่"
        print(
            f'ใช้ vatinfo_req และส่ง data body ด้วย : {str(tax_num)}, สาขา {str(branch)}')

        # * หาชื่อใบกำกับจาก vatinfo
        result = self.get_res_vatinfo(str(tax_num), str(branch))

        # * กรณีหาจาก taxinfo ไม่มี ทำให้ต้อง หาจาก Excel ที่ import เข้ามา
        if bool(result) == False:
            # * หาตำบล จาก address ที่ลูกค้าให้มา
            cus_address_from_table = self.assign_address(
                self.app.data_frame, self.app.cus_order.get())

            manual_result_strcuture = {
                'tax_num': f'{self.app.tax_num.get()}',
                'branch': f'{self.app.branch_type}',
                'name': f'{self.app.cus_name.get()}',
                'address_shortened': f"{cus_address_from_table['cleaned_address']}",
                'province': f'{self.app.cus_province.get()}',
                'district': f'{self.app.cus_district.get()}',
                'sub_district': f"{cus_address_from_table['decent_tambon']}",
                'province': f'{self.app.cus_province.get()}',
                'address': f'{self.app.cus_address}',
                'postal_code': f"{self.app.nondistortedData['รหัสไปรษณีย์.1']}",
            }
            result = manual_result_strcuture
        else:
            result['name']
            # * ตรวจสอบดูว่า ค่าที่ response กลับมา มีช่องว่างตามเงื่อนไขหรือไม่
            x = re.search("^ห้างหุ้นส่วนจำกัด\s|^บริษัท\s", result['name'])

            if x:
                # * มีช่องว่าง แปลว่าดี
                print("เจอช่องว่าง response ไม่ต้องทำอะไร return ได้เลย",
                      result['name'])
            else:
                #! ไม่มีช่องว่าง แปลว่าอับปรีย์
                result['name'] = result['name'].replace("บริษัท", "บริษัท ").replace(
                    "ห้างหุ้นส่วนจำกัด", "ห้างหุ้นส่วนจำกัด ")
                print("ไม่เจอช่องว่างจาก response แต่เพิ่มให้แล้ว",
                      result['name'])

        return result


if __name__ == "__main__":
    def on_closing():
        print("Tkinter window is closing")
        root.destroy()

    # def ctrl_saraea_copy(event):
    #     ctrl_state = event.state & 0x4 != 0  # 0x4 คือ flag สำหรับ Control key
    #     # 67 คือรหัสสำหรับสระแอในภาษาไทย (อาจแตกต่างบนระบบอื่นๆ)
    #     if ctrl_state and event.keycode == 67:
    #         event.widget.event_generate("<<Copy>>")

    # * เทคนิคคือ เช็คว่า ascii คือไร แล้วดูด้วยว่า นอกจากรับแบบ ascii แล้วรับแบบ keysym(ตัวอักษรจริง)ว่าตรงกับ ascii ไหม ถ้าไม่ตรงแปลว่าคนละภาษาแน่นอน เพราะ มันจะได้ ??
    def _onKeyRelease(event):
        # print("press :", event.keysym)
        ctrl = (event.state & 0x4) != 0
        if event.keycode == 88 and ctrl and event.keysym.lower() != "x":
            event.widget.event_generate("<<Cut>>")

        if event.keycode == 86 and ctrl and event.keysym.lower() != "v":
            event.widget.event_generate("<<Paste>>")

        if event.keycode == 67 and ctrl and event.keysym.lower() != "c":
            event.widget.event_generate("<<Copy>>")

        if event.keycode == 65 and ctrl and event.keysym.lower() != "a":
            event.widget.event_generate("<<SelectAll>>")

    root = CTk()
    # * options
    # * change icon
    root.iconbitmap(icon_path)

    # * > ทำลาย root tkInter เมื่อguiถูกปิด เพื่อไม่ให้มีการทำงานตกค้าง
    root.protocol("WM_DELETE_WINDOW", on_closing)

    # * > ลืมไปละ น่าจะเกี่ยวกับช่องไฟ
    root.columnconfigure(0, weight=1)

    # * > ปรับขนาดจอ
    # root.resizable(False, False)

    # * > ทำให้กด copy, paste, cut จากภาษาอะไรก็ได้
    root.bind('<Key>', _onKeyRelease)

    # * Create Instance
    app = MyApp(root)
    root.mainloop()

# ปัญหาที่ต้องแก้
# *1แก้แล้ว** บรรทัดล่างสุด"สินค้ารวมค่าส่งหัก seller: " เลขยอดเงิน ที่แสดงผล เมื่อเจอ list mี่มีสมาชิกหลายตัว มันจะรวมแค่ตัวแรกอย่างเดียว ต้องใช้ forloop รวมราคาทุกตัว
# *2 fixed 0.394 // เวลา kb เป็น ภาษาไทย จะกcopy ข้อความใน log ไม่ได้ น่าจะเป็นเพราะ เครื่องไม่ได้รับค่า ctrl+c แต่เป็น ctrl+แ
# *3 แอดใบกำกับ บัค ตรงที่ เราเหลือ ปุ่มสุดท้ายยังไม่กด แต่พอยังไม่กด มันไม่รอ มัน error ไปเลย
# *4 ลูกค้าธรรมดาไม่ต้องกรอก address แต่ใบกำกับกรอกให้แม่น
# *5 แก้แล้ว**ใบกำกับ/บิล ลืมล้างค่าก่อน แอด จริงๆ ทุก input ต้องล้างก่อนแอด ควรทำเป็นนิสัย ยังไม่เสร็จ
# *6แก้แล้ว** ตัวอย่าง order ทำ scrollbar "231010GK0S3VV3" เพราะมีหลายรายการ
# *7 หน้าท้ายรัน Auto ไปด้วย จะได้ไม่ต้อง copy
# *8 ปิด thread หลังจบคำสั่งด้วย terminateไม่ได้ แก้ด้วยข้อ 22 แทน
# *9แก้แล้ว** Total LastPage in SMCO -> ราคาที่ต้องออก
# *10แก้แล้ว** "Auto หน้าท้าย ทำได้ครั้งเดียว ส่วนนี้มันจะดัก" เวลามี pop-up ขึ้นกรณียิงของแล้ว SN ไม่มี มันจะ BUG  wait มันจะ Error ทีก่อนหน้านี้ waitfail ดันไม่ error งงชิพไห
# * /html/body/div[16]/div[2]/div[6] มีข้อความ "Your data has been successfully saved doing print Invoice No : B0183-W06-2310130023"
# *11 มีบัคตรงที่ลูกค้าบางคนใส่ (สำนักงานใหญ่) บางคนไม่ใส่ (สำนักงานใหญ่)//เปลี่ยนวิธี Add ลูกค้า และ ใบกำกับใหม่ ใช้สูตร BigM
# *12 ทำแล้ว //ทำ input ID PASS
# *13 แก้แล้วแต่จะใส่ให้สองครั้ง//ยังแก้ไม่ได้ลูกสึก bigM ยังมีปัญหากับตรงนี้อยู่ อาจจะลองแก้ด้วย while True // searhลูกค้า ไม่เจอแล้วแอด มันมีโอกาสที่แอดแล้วไม่เสิชต่อ
# *14 ทำแล้ว // ทำแยกตารางใหม่โดยใช้ layout แบบ Shopee //ตัวอักษรใน LOG หรือ ทำให้ Log อ่านและแยกแยะง่ายขึ้น ใช่ มันอ่านยากจริงๆ
# ?15 ตรวจดูแล้วยังไม่เจอสาเหตุ** ข้อความ "เพิ่มไฟล์แล้ว" แสดงผลไม่ถูกต้อง เนื่องจาก แสดงผล แม้ไม่ได้ แอดไฟล์จริงๆ
# *16 รายงาน มาว่าไม่เจอ แก้แล้วไม่รู้ใช้ได้ยัง // U200b display as ?
# *17 สินค้าบางประเภทต้องใส่ Variations ของมันด้วย ใน log จะได้แยกได้ เช่น หมึก มันจะไม่บอกสีใน ชื่อสินค้า แต่บอกใน variations
# *18 มีเลขลำดับบอกใน productslist
# ?19 แก้แล้ว!!!ยากมาก!!!เลยไม่ชัวว่าแก้ได้จริงป่าว ///order ไม่มี แต่ยังทำงานอยู่ เกิดจากการทำงานมันแยก thread กัน ต้องเอาผลลัพจากการเสิช มาเป็นเงื่อนไขว่าจะทำต่อหรือไม่
# *20 แก้แล้ว //แก้แล้วรอทดสอบ//ใบกำกับไม่มีคำว่า ใน margetplace มีคำว่า (สำนักงานใหญ่) แต่พอแอดมาดันไม่มี
# *21 ใช้ได้แล้ว //ทำได้แล้วรอทดสอบ //หน้าสุดท้ายกรอกเบิ้ล หากมีการยกเลิก หรือ รันบอททับ (ยากชิพไห) แต่หลักๆแก้ด้วย while True
# *22 แก้แล้ว//พวกไม่ขอแต่มีเลข มันจะได้สาขา nan มา ต้องแก้ด้วย
# *23  เพราะลูกค้าไม่ได้บอกว่าเป็น หจก หรือ บจก ไง เลยทำเงื่อนไขไม่ได้ เพราะกูก็ไม่รู้ว่าต้องเขียนชื่อเป็นอะไร // 231021G8CWC1N5 คำว่า บริษัทไม่ขึ้น
# *24 แก้แล้ว//เวลาสินค้ามีมากกว่า 1 รายการ แล้วถัดไปมีน้อยลง element ที่แสดงรายการ ของ order ที่แล้วจะไม่หายไป
# ?25 แก้แล้วเมื่อมี error thread จะถูกปิดทันที //Threading ทำให้ chrome กิน ram หนักมาก จนทำให้ browser ค้าง
# *26 แก้แล้วเกิดจาก ใช้ตัวแปรผิด ลืมใช้ตัวแปรที่เก็บค่าที่ลบคำแล้ว แต่ใช้ค่าเดิมไปเติม (สำนักงานใหญ่) จึงทำให้คนที่ให้ชื่อที่มีคำว่า "(สำนักงานใหญ่)" จะได้รับการเพิ่มคำว่า "(สำนักงานใหญ่)" ทำให้เบิ้ล //คำว่า สำนักงานใหญ่ เบิ้ล
# *27 แก้แล้ว // ลูกค้าขอใบกำกับแต่ให้คำว่า สาขาย่อย แต่ไม่มีชื่อสาขา และไม่มีรหัสสาขา แต่code ให้ผลลัพธ์ว่า (สาขาnan)
# *28 เพิ่ม Bot Status ว่ากำลังทำไรอยู่
# *29 เพิ่มแล้ว //เพิ่มช่องหมาเหตุจากผู้ซื้อ และ บันทึก
# *30 ปรับการทำงานให้เข้ากับ SMCO v6.2 อันเดิมคือ 6.1.1
# *31 แก้แล้ว//เพิ่มหน่วงเวลาให้ตอนกดแอดลูกค้าดูเหมือนว่า element ที่แอดลูกค้า มันจะขึ้นมาช้า locator มันเจอ แต่ กดไม่ได้ ซึ่งcodeผมมันสั่งให้กดไวไป = กับว่า การใช้ wait elment โผล่ กับ clickable element โผล่มันจะไวกว่า ต้องใช้อะไรที่ช้ากว่านั้นก็คือ clickable
# *32 แก้แล้วมั้ง ปัจจุบันไม่มีปัญหา จำไม่ได้ว่าแก้ตอนไหน // ทำ auto ตอนเริ่ม phase2 แต่ตอนนี้มีปัญหา error data type ถ้าเอาตัวauto ไปใช้ ใน final whileloop
# !!33 ใน phase2 ก่อน final loop จะต้องเช็คก่อนว่าเข้า final ได้ไหม โดยการเช็ค "ราคารวมก่อนหักseller voucher"  ว่ามีค่าตรงกับ '/html/body/div[1]/div[2]/div[2]/div[2]/div[2]/div[2]/div/div/div/div/span[1]' หรือไม่ ถ้าไม่ตรงให้ finalloop ไม่ต้องทำงานแต่จะกด esc ย้อนกลับไปหน้าเก่า
# *34 แก้แล้วหายแล้ว//แต่ยังบัคอยู่//แก้แล้ว//ตัว auto print bug ย้อนกลับหน้าเดิมไม่ได้
# *35 แก้แล้วใช้ได้//SMCO เอา Auto ออกทำให้ใช้ไม่ได้
# !36 ใช้หาใบกำกับได้ดีกว่า vatinfo สะอีก https://www.dataforthai.com/company/{เลข13หลัก}/
# *37 แก้แล้ว //ใน log ด้านล่าง จะไม่ได้แยกการแสดงผลของ SHOPEE กับ LAZADA นะ
# ?38 แก้แล้ว // module แปลภาษา รู้สึกจะมีปัญหาเรื่อยๆ เพราะมันมีตัวอักษรพิเศษ แฝงในชื่อด้วย
# ?39 แก้แล้ว // โหมดเสิชลูกค้ารู้สึกว่าจะไม่มีเวลารอ หรือไม่ก็มีการออกแอคชั่นกด ที่เร็วเกินไป ยังหา elemtn ไม่เจอเลย
# *40 แก้แล้ว // pop up ของ browser ทำ element ใน DOM หาย ทำให้ while loop error ต้องหยุดในช่วงที่ elment หายส่งผลให้ BOT หยุดทำงาน
# ?41 หายแล้วแต่ไม่ได้แก้ แค่เดินไปก็หายเอง //810074145748076 วันที่ 16/01/2024 อันนี้เคสตัวอย่างเลขใบกำกับ dtype มันกลายเป็นเลข
# ?42 24011504S292UB แอดไม่ติด ได้ไงวะ? แต่ปั่นอยู่
# *43 แก้แล้ว 0.382 // File address lazada ที่ add เข้าไป มันใช้ไม่ได้ หาไม่เจอนั่นเอง
# *44 0.383 แก้แล้ว // pop-up ของ contextwindow browser มันยังคงทำให้บอทดับอยู่ดีน่าจะเกิดจากการที่เราใช้ time.sleep แต่เราแก้ด้วยการใช้ while + try,except
# *45 0.384 แก้แล้ว // ปรับความเร็วกรอกบิล ในขั้นตอน reprint
# *46 0.384 แสดงlogเลขบิล
# *47 0.385 เอาเลขบิลมาโชว์ที่ GUI
# Todo 48 กรอกก่อนที่ element จะ display ได้ ดูเหมือนจะเป็นเช่นนั้น
# * 8/2/2024
# *49 fixed 0.387 // lazada ราคารวม bug
# *50 fixed 0.387 // Ultimate CP prototype for หมึก
# *51 fixed 0.387 // lazada ลูกค้า ภาษาสเปน googletrans ช่วยไม่ได้ กรณีถ้าแปลแล้วไม่ได้จริงๆ return ค่าinput ไปแหละ
# *52 fixed 0.387 // Order ยกเลิกแสดงผลไม่ชัดเจน
# *53 fixed 0.387 // shopee ปรับเปลี่ยนวิธีหาชื่อลูกค้าจาก Email เป็น ใช้ เลขผู้เสียภาษี
# *54 fixed 0.387 // shopee ลดความเร็วในการกรอก แขวง/ตำบล ใน dropdown ตอน เพิ่มชื่อลูกค้าใหม่
# !55 เอาเป็นว่าใช้เขตบางบอนดีกว่า มีแขวง บางบอนอยู่ 5 อัน แนวทางการแก้ไขอาจจะต้องใช้ req/res เพื่อดึงค่าจาก SMCO มาใช้แล้วแหละไม่งั้นทำไม่ได้ //วัฒนา ทวีวัฒนา dropdown จะมีสองค่า แล้วมันจะเอาค่าที่ยาวกว่าขึ้นก่อน การเลือกอันที่ 1 มันจะ ผิด
# ?56 สังเกตุมาเป็นปีละไม่เคยเจอกับตัวแต่คนอื่นเจอตลอด // จาก กรุงเทพ กลายเป็น ภูเก็ตได้ order นี้ 2401309DCMAYCS มันคนละแบบกับที่เจอตอนแรกที่ผิดแค่ ตำบล แต่อันนี้ผิดที่จังหวัด ต้องไป recheck ที่จุดเริ่มต้น
# *57 หายแล้ว จำไม่ได้แก้แพทช์ไหน // ประเทศไม่เลือกไทยในบางกรณี บางกรณีเลือกเป็น china เป็นเพราะเลือกจาก index แต่ไม่ได้เลือกจากข้อความด้านใน
# !!58 สำคัญมาก ใบกำกับที่ print ออกมาจะ !!!แสดงผลด้วยภาษาไทย!!! แต่จะ !!!เสิชจากภาษาอังกิด!!! ถ้าจะใช้เพื่อ เสิช ต้องใช้ภาษาอังกิด อังกิดจะใส่ไรก็ใส่
# !59 เวลามีหลาย SKU มัรจะ sonic blow ช้า
# *60 fixed 0.388 // อัพเดท Path ของ Shopee เนื่องจาก Shopee อัพเดท path หน้าเว็บใหม่
# *61 fixed 0.388 // Sonic blow บัค
# *62 fixed 0.389 // Accel_mode มันจบที่หน้าท้ายหน้าปริ้น ทำให้เวลาขึ้น loop ใหม่มันจะ error
# *63 fixed 0.389 // แก้เป็น float แล้ว // seller voucher Lazada มันมีค่าทศนิยมด้วย เนื่องจากมีบัคเก็บค่าของ sellervoucher เป็น int ไม่ใช่ float
# *64 fixed 0.389 // เพิ่ม pattern แล้ว // ใน method cus_name_standardizer() นอกจากจะมี "สำนักงานใหญ่" ในชื่อแล้ว บางกรณีมีคำว่า สนญ. ด้วย
# *65 fixed 0.389 // ทำตัวโหลด chromedriver อัตโนมัติ
# !66 ยังพังอยู่มันยัง Add ลูกค้าใหม่ได้ยังไม่ดีพอ // fixed 0.390  // ลองแล้วแต่ยังไม่ชัวเพราะใส่ callback recursion ด้วย ซึ่งยังไม่เซียน //การเลือกลูกค้าบางทีข้อมูลลูกค้าไม่ตรงกับที่ขอมา
# *67 fixed 0.390 ยังไม่ชัว น่าจะยังไม่ได้แก้ // ข้อมูล ไม่ตรงกัน ในส่วนของอันบนและ อันล่าง(ในGUI) orderตัวอย่าง 240416U5DMC0E5 เนื่องจาก Order นี้ มีการใส่ข้อมูลใน column "บันทึก" เข้ามา แปลว่าที่ผ่านมาไม่เคยเจอเลยงั้นรึนี่
# *68 Fixed 0.390 // SMCO อัพ 6.3.1 24/04/2024 ทำให้ต้องเพิ่ม input ในส่วนของ ประเภทลูกค้า ไม่งั้น submit form ไม่ได้
# *69 Fixed 0.390 // pop-up หลัง add ชื่อลูกค้ากลับมาอีกครั้ง จัดการแล้ว
# *70 Update 0.390 // ปรับให้ Lazada ต้องกด save เองกรณีใบกำกับ
# ?71 Fixed 0.390 Lazada เลขใบกำกับlazada ไม่ยอมเป็น str แถม ตัด 0 ด้านหน้าออก หลังแปลงค่าด้วย
# *72 Fixed 0.390 // ปรับcode การเลือก ช่องทางชำระเงินให้แม่นยำยิ่งขึ้น
# *73 Fixed 0.390 // Shopee อัพเดท ui ใหม่ ทำให้ต้องเปลี่ยน path ใหม่
# Todo74 Fixed 0.391 // Last pop-up มีตัวรอ event ที่เป็น driver.wait ทำให้รอนาน ควรเปลี่ยนเป็น while loop จะได้จบ errror ทันที
# *75 Fixed 0.391 // Shopee อัพเดท ui ใหม่ ทำให้ต้องเปลี่ยน path ใหม่ อีกแล้วเรอะ
# *76 Fixed 0.391 // จาก ข้อ 66 ปรับวิธีเลือก li ใบกำกับ เนื่องจากอันเดิม เป็นการเลือกจาก "ชื่อเต็ม"จาก li  แต่มันมีปัญหาคือ หา element ไม่เจอ เปลี่ยนไปใช้หาโดย idx แทนทดสอบแล้ว แม่นอยู่ (แต่เดี๋ยว พอใช้จริงพัง 55555)
# *77 Fixed 0.392 // ปรับช่วงรับ Data ขาเข้าของ Sonicblow ให้ตัด space ออกก่อนแล้ว
# *78 Done but bug 0.392 // สรุปว่าพัง // ปรับ accel mode แบบ อัดทุก SKU รอทดลองว่าพังไหม
# !79 issue จากข้อ 78 มันพังเวลาloop หา sku อื่นหลังจากจบ sku ก่อนหน้า อันแรกของ sku ถัดไป จะพังเป็นบางรอบ
# !80 issue Accel mode ยัง ขาด ความสามารถในการตรวจผลลัพธ์ว่า SN ที่กรอก ถูกต้องหรือไม่ มันกรอกและจบไปเฉยๆ
# *81 Fixed 0.392 // สามารถใช้ copy shortcut ขณะที่ keyboard input เป็นภาษาอื่นนอกจากภาษาอังกฤษได้แล้ว
# Todo82 // WIP deduct_accel_file_data ยังไม่เสร็จ เหลือจัดการ sn ต้องเก็บ sn ที่ใช้เป็น array
# *83 Fixed 0.392 แก้ละ //Shopee ลบชื่อลูกค้าออกไปจาก Exported File แล้ว ทำให้ เพิ่มชื่อลูกค้าไม่ได้ // แนวทางคือ ใช้ชื่อ Account+\s+ชื่อที่มีแต่\* แทน
# * 84 Fixed // จากข้อ 83 มันจะมีลูกค้าบางคนใช้เครื่องหมาย "(" หรือ ")"ทำให้ชื่อลูกค้าใช้เสิชหาชื่อลูกค้าไม่ได้
# * 85 Fixed 0.392R2// จากการแก้ 83 ทำให้ lazadabug แก้แล้วรอทดสอบ
# * 86 Fixed 0.392R3// แก้ Path ของ Shopee เนื่องจาก Shopee อัพเดท path input หน้า "ทั้งหมด" ใหม่
# !87 // Accelmode ปัญหาน่าจะเกิด เมื่อมีการข้าม บิล sn จะถูกข้ามมั้ง มันมีโอกาสที่จะไม่ดึงSN ที่เหลือ
# ?88 // แอดแบบมี * น่าจะไม่เวิร์ค เพราะหลายๆค่าใน db มี* ทำให้้ช้ามั้ง ยังไม่เคยลองทดสอบ
# * 89 Added 0.393 // เพิ่ม ฟังชั่น Read transfer เพื่อเพิ่มลง accelmode_file
# ! 90 ช่วงถ้ายังเลือก dropdown ไม่ได้มันจะ error
# ? 91 Fixed 0.394 // ปรับลูกค้าที่มีบันทึกให้อ่านค่าจากบันทึกได้ แต่อาจจะต้องมีทำต่อ เพราะใช้ได้แค่กรณีไม่ได้ขอใบกำกับ ตรงนี้ไม่มีระบบรองรับ ทำให้ input Dynamic มาก
# * 92 Fixed 0.394 //  อ่านได้ละ //ตัวอ่าน PDF ยังแยก serial ได้ไม่ดี
# * 93 Fixed 0.394 // แก้ การดึงค่า สถานะ จาก UI ใหม่ Shopee กรณี ส่งสำเร็จ, ยกเลิก, ส่งแล้ว
# * 94 Fixed 0.394 // น่าจะได้มั้งไม่มี FeedBack // อ่าน sn จาก pdf แล้ว อัพเดทค่า sn ลง state ได้แล้ว  แต่ต้องเทสก่อนว่าตัดค่าได้อย่างเหมาะสมหรือไม่
# * 95 Fixed 0.395 // แก้ให้เลือกได้แล้ว // มีโอกาสที่จะเลือกประเภทลูกค้าไม่ได้
# * 96 Added 0.395 // Logger เอาไว้ตรวจสอบการทำงานว่าเริ่มแล้วจบไหม
# * 97 Fixed 0.395 // แก้ xpath แล้ว shopee ปรับ interface
# * 98 Fixed 0.395R1 // เพิ่ม regex ในการ read pdf
# ! 99 popup หลัง add ลูกค้ามันต้องการเวลารอนานกว่านี้ เหมือนมันจะหา element ในขณะที่ตอนกด submit ลูกค้ายังไม่เสร็จ เลยข้าม order // เรียกหาชื่อปกติก็เปน ช่วงก่อนกรอก ค่าส่ง หลังเจอชื่อ
# ? 100 Fixed 0.395R2 try to see the result after fixed if there still bug or not// duplicated orders will be remove, but removing duplicated order will affect sn columns  to be disappered
# * 101 Fixed 0.395R2 // shopee ปรับ interface พบการเปลี่ยนแปลงวันที่ 18/09/2024
# * 102 Fixed 0.395R3 // shopee ปรับ interface พบการเปลี่ยนแปลงวันที่ 19/09/2024 รู้สึกแปลกๆตั้งแต่ของ 0.395R2 แล้วละ เมื่อน element มันไม่ครบ
# * 103 Fixed 0.395.4  // เปลี่ยนชื่อแล้ว 27/09/2567 16:23 // Lazada เปลี่ยน title ใหม่ เลยเข้า tab lazada ไม่ได้ 27/09/2567 16:23
# * 104 Added 0.396.0  // update new version Printing 08/10/2024 15:41
# * 105 Added 0.396.1  // patch new version Printing 08/10/2024 16:46

# Todo ควรจะต้องแยก MODULE เป็นแบบ version ธรรมดา กับ version ETAX เพราะวิธีการทำงานค่อข้างแตกต่างกัน
#!--------------------- ETAX SAGA ------------------------------------
# *E1 อยากให้ display Email ใน gui ปัญหาจริงๆมาจาก Shopee ไม่รู้ว่ามีลูกค้าขอใบกำกับ เขาไม่ได้ขอมาโดยตรง แล้วมันขัดกับ วิธีการทำใบกำกับของ SHOPEE สูตร BigM ด้วย
# *E1.5 ในบิลมันมีคำว่า tax ID : ขึ้นรอไว้เลย แบบมันมัดมือชกเลยว่ามึงต้องแอดใบกำกับเท่านั้น 555+ ต้องกลับมาทำเวย์เดิมแล้ว
# *E2 The POS does not lead to the printing page. ทำให้ต้องไปเปิดหน้า print แยก ซึ่งอาจจะถูกแก้สักวันละมั้ง
# *E3 canvas มันไม่โผล่ ทำให้ bot status มันไม่จบ มันจะค้างที่ Your turn
# Todo ETAX ได้ข่าวมาว่าจะไม่ได้ใช้ตลอดไป แต่อาจจะเลิกใช้ที่เดือนกุมภา วันที่ 15 แปลว่า etax อาจจะเป็นโหมดชั่วคราว

# Todo Future Features
#!--------------------- Auto CP SAGA ------------------------------------
# * ACP1 CPs บางอันมันจะมีวันที่ทับกัน
# * ACP2 CPs แต่ละอันมันอาจจะมีราคาที่ต้องออก เท่ากัน บางอัน
# * ACP3 แต่ CPs แต่ละอันมันจะมีมูลค่า CP ไม่เท่ากัน ถึงแม้ปลายทางมันจะเหมือนกัน ผลรวมลด เท่ากัน

# เก็บข้อมูล
# รอให้ final pop-up poped up /html/body/div[16]/div[2]/div[6]
# หรือ
# กดปุ่ม รอจนกว่าปุ่มนี้จะกดได้ /html/body/div[16]/div[2]/button[1] then click


# note
# /html/body/div[1]/div[2]/div[2]/div[2]/div[1]/div[2] parent

# /html/body/div[1]/div[2]/div[2]/div[2]/div[1]/div[2]/div[1]/div/div[1]/div[2]/a sv0
# /html/body/div[1]/div[2]/div[2]/div[2]/div[1]/div[2]/div[2]/div/div[1]/div[2]/a MNL

# you should try this to find child 2 <div>
# div_elements = x.find_elements_by_xpath("./div")
