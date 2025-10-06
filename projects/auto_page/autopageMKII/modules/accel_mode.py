import pandas as pd
import os
import re
from tkinter import filedialog
from pypdf import PdfReader
from openpyxl import load_workbook
import time
from selenium.webdriver.common.by import By


class AccelMode:
    def __init__(self, main_app):
        self.main_app = main_app
        self.accel_file_dir = ""
        self.accel_df_state = {}
        self.accel_file_columns = []
        self.obj_data_from_accel_file = {}
        self.accel_orders_list = []
        self.CP_list = []
        self.accel_orders_count = 0

    def select_accel_file(self):
        self.accel_file_dir = filedialog.askopenfilename(title="Select Accel File")
        if self.accel_file_dir:
            self.main_app.accl_dir_namedisplay_on_btn.configure(text=f"{self.accel_file_dir.split('/')[-1]}")
        else:
            self.main_app.accl_dir_namedisplay_on_btn.configure(text=f"ยังไม่เลือก Accel File")

        self.read_accel_file_to_state(self.accel_file_dir)

    def read_accel_file_to_state(self, accel_file_dir):
        self.accel_file_dir = accel_file_dir
        self.accel_df_state = pd.read_excel(self.accel_file_dir, dtype=str)
        print("before self.accel_df_state: ", self.accel_df_state)
        self.accel_df_state.loc[self.accel_df_state.duplicated(subset=['orders']), 'orders'] = pd.NA
        self.accel_df_state['orders'].dropna(inplace=True)
        print("after self.accel_df_state: ", self.accel_df_state)

        self.accel_file_columns = self.accel_df_state.columns.dropna().tolist()
        self.obj_data_from_accel_file = {
            col: self.accel_df_state[col].replace(" ", '').dropna().tolist() for col in self.accel_file_columns
        }

        self.accel_orders_list = self.accel_df_state['orders'].dropna().tolist()
        self.CP_list = self.accel_df_state['cp'].dropna().tolist()
        print(self.accel_orders_list)
        print('self.obj_data_from_accel_file: ', self.obj_data_from_accel_file)
        print(self.CP_list)

    def deduct_accel_file_data(self, order, sku_serials=[]):
        order = order.get()
        df = self.accel_df_state
        print("deduct_accel_file_data df มีมาก่อนเหรอ: ", df)
        print("deduct_accel_file_data order: ", order)
        print("deduct_accel_file_data ref: ", df.loc[df['orders'] == order, 'orders'])

        has_order = df.loc[df['orders'] == order, 'orders']
        if not has_order.empty:
            print(f'remove {order} from state df')
            df.loc[df['orders'] == order, 'orders'] = pd.NA

        print("sku_serials ไม่ได้ได้ไง: ", sku_serials)
        if sku_serials:
            for sn in sku_serials:
                df.loc[df[sn['sku']] == sn['sn'], sn['sku']] = pd.NA

        print("form state df to new excel")
        print(f"Check if accel file is accesible {os.access(self.accel_file_dir, os.W_OK)}")

        try:
            df.to_excel(self.accel_file_dir, sheet_name='Sheet1', index=False)
            print(f"Successfully updated {self.accel_file_dir}")

            # อ่าน dataframe ใหม่หลังจากอัปเดต Excel file
            self.accel_df_state = pd.read_excel(self.accel_file_dir, dtype=str)
            self.obj_data_from_accel_file = {
                col: self.accel_df_state[col].replace(" ", '').dropna().tolist() for col in self.accel_file_columns
            }
        except PermissionError as e:
            print(f"Permission denied: {e}")
            print("ไฟล์ Excel อาจถูกเปิดอยู่ในโปรแกรมอื่น กรุณาปิดไฟล์แล้วลองใหม่")
            # ใช้ข้อมูลในหน่วยความจำแทน ไม่อัปเดตไฟล์
            self.accel_df_state = df
            self.obj_data_from_accel_file = {
                col: self.accel_df_state[col].replace(" ", '').dropna().tolist() for col in self.accel_file_columns
            }
        except Exception as e:
            print(f"Error updating Excel file: {e}")
            # ใช้ข้อมูลในหน่วยความจำแทน
            self.accel_df_state = df
            self.obj_data_from_accel_file = {
                col: self.accel_df_state[col].replace(" ", '').dropna().tolist() for col in self.accel_file_columns
            }

    def extract_sn_btn(self, accel_file_dir):
        if not accel_file_dir:
            print("select accel file first!!")
            return

        target_dirs = filedialog.askopenfilenames(title="Select SN PDF files")
        if len(target_dirs) != 0:
            for target_dir in target_dirs:
                self.sn_extractor(accel_file_dir, target_dir)
        else:
            print("You have not selected any transfer file, Extraction ends!!")
        self.accel_df_state = pd.read_excel(self.accel_file_dir, dtype=str)

        self.read_accel_file_to_state(self.accel_file_dir)

    def sn_extractor(self, output_excel, target_dir):
        extracted_txt = self._extract_text_from_pdf(target_dir)
        extracted_txt = self._clean_text(extracted_txt)

        product_codes = self._extract_skus(extracted_txt)
        serial_numbers = self._extract_serial_numbers(extracted_txt)

        cleaned_serial_numbers = self._clean_serial_numbers(serial_numbers)
        serial_numbers_grouped = self._group_serial_numbers(cleaned_serial_numbers)

        self._print_debug_info(product_codes, cleaned_serial_numbers, serial_numbers_grouped)
        self._write_to_excel(output_excel, product_codes, serial_numbers_grouped)

    def _extract_text_from_pdf(self, target_dir):
        reader = PdfReader(target_dir)
        extracted_txt = ""
        for page in reader.pages:
            extracted_txt += page.extract_text()
        return extracted_txt

    def _clean_text(self, text):
        pattern = r'^.*?(?=No\. Product Code Barcode Product Name Transfer No\. Order Ship Status)'
        text = re.sub(pattern, '', text, flags=re.DOTALL).lstrip()

        pattern2 = r'ผู้ส่งสินค้า.*?(?:No\. Product Code Barcode Product Name Transfer No\. Order Ship Status|วันที่ _ _ _ / _ _ _ / _ _ _)'
        text = re.sub(pattern2, '', text, flags=re.DOTALL)

        text = re.sub(r'Serial\s:', '', text, flags=re.DOTALL)
        text = re.sub(r'\d+\s{0,}(?=([A-Z0-9]{3}-[0-9]{6}))', '', text, flags=re.DOTALL)

        return text

    def _extract_skus(self, text):
        sku_pattern = r'([A-Z0-9]{3}-[0-9]{6})'
        return re.findall(sku_pattern, text)

    def _extract_serial_numbers(self, text):
        serial_pattern = r'(?:Shipped|Confirm)\s*([\w, \n, \/]+)(?=(?:[A-Z0-9]{3}-[0-9]{6}|\nผู้ส่งสินค้า|$))'
        return re.findall(serial_pattern, text, re.DOTALL)

    def _clean_serial_numbers(self, serial_numbers):
        cleaned = []
        for serial in serial_numbers:
            serial = re.sub(r'\n', '', serial).strip()
            cleaned.append(serial)
        return cleaned

    def _group_serial_numbers(self, cleaned_serial_numbers):
        return [serial.replace(" ", "").split(",") for serial in cleaned_serial_numbers]

    def _print_debug_info(self, product_codes, cleaned_serial_numbers, serial_numbers_grouped):
        print("Product Codes:")
        for i, code in enumerate(product_codes, 1):
            print(i, code)

        print("\nSerial Numbers:")
        for i, serial in enumerate(cleaned_serial_numbers, 1):
            serial_list = serial.replace(" ", "").split(",")
            print(f"{i} {len(serial_list)} [{serial}]")

        print("SKU Matches:", len(product_codes), product_codes)
        print("Serial Numbers Grouped:", len(serial_numbers_grouped), serial_numbers_grouped)

    def _write_to_excel(self, output_excel, product_codes, serial_numbers_grouped):
        try:
            book = load_workbook(output_excel)
            sheet = book.active

            existing_skus = {}
            for col in range(1, sheet.max_column + 1):
                sku = sheet.cell(row=1, column=col).value
                if sku:
                    existing_skus[sku] = col

            for sku, serials in zip(product_codes, serial_numbers_grouped):
                if sku in existing_skus:
                    col = existing_skus[sku]

                    existing_serials = []
                    for row in range(2, sheet.max_row + 1):
                        val = sheet.cell(row=row, column=col).value
                        if val:
                            existing_serials.append(val)

                    merged_serials = existing_serials.copy()
                    for s in serials:
                        if s not in merged_serials:
                            existing_serials.append(s)

                else:
                    col = sheet.max_column + 1
                    sheet.cell(row=1, column=col, value=sku)
                    for row, serial in enumerate(serials, start=2):
                        sheet.cell(row=row, column=col, value=serial)
                    existing_skus[sku] = col

            book.save(output_excel)
            print(f"ข้อมูลถูกเพิ่ม/อัปเดตลงใน {output_excel} เรียบร้อยแล้ว")

        except PermissionError as e:
            print(f"Permission denied: {e}")
            print("ไฟล์ Excel อาจถูกเปิดอยู่ในโปรแกรมอื่น กรุณาปิดไฟล์แล้วลองใหม่")
        except Exception as e:
            import traceback
            print(f"เกิดข้อผิดพลาด: {e}")
            traceback.print_exc()

    def accel_search(self):
        self.main_app.is_accel_mode_activated.set(True)
        self.accel_orders_count = len(self.accel_orders_list)

        def start_next_cycle(count):
            self.accel_df_state = pd.read_excel(self.accel_file_dir, dtype=str)
            if count < self.accel_orders_count:
                if self.main_app.is_accel_mode_activated.get():
                    self.main_app.search_order(self.accel_orders_list[count], lambda: start_next_cycle(count+1))
                else:
                    raise ValueError("Accel mode has been destroyed")
            else:
                pass

        self.main_app.search_order(self.accel_orders_list[0], lambda: start_next_cycle(1))

    # * เอาไว้ใช้กับ smco โดยการเอา sn จาก accel file มาใส่ในช่อง sku input บนเว็บ smco
    def accel_fill_sku(self, driver, operation_thread):
        from selenium.webdriver.common.keys import Keys
        from loguru import logger

        available_sn_skus_list = list(self.obj_data_from_accel_file.keys())
        self.used_serials = []
        ordered_items = self.main_app.items
        print('accel_fill_sku() ตรวจสอบ items = ', ordered_items)

        if len(ordered_items) > 0:
            for i, ordered_item in enumerate(ordered_items):
                print("item ordered by customer", ordered_item)
                current_sku = ordered_item['เลขอ้างอิง SKU (SKU Reference No.)']
                print("current_sku: ", current_sku)
                sku_qtys = ordered_item['จำนวน']
                is_sku_ready_to_pick = [key for key in available_sn_skus_list if key in current_sku]

                if len(is_sku_ready_to_pick) > 0:
                    for item in range(sku_qtys):
                        print("self.obj_data_from_accel_file[current_sku]: ",
                              self.obj_data_from_accel_file[current_sku])
                        self.obj_data_from_accel_file[current_sku]

                        if self.obj_data_from_accel_file[current_sku]:
                            print("มี SN")
                            time.sleep(1)
                            while not operation_thread.is_set():
                                try:
                                    driver.find_element(
                                        By.XPATH,
                                        '/html/body/div[2]/div[3]/div[2]/div[2]/div[1]/div[1]/from/div/div/div[1]/div[1]/span/input')
                                    break
                                except:
                                    continue

                            sn = self.obj_data_from_accel_file[current_sku][0]
                            skuInput = driver.find_element(
                                By.XPATH,
                                '/html/body/div[2]/div[3]/div[2]/div[2]/div[1]/div[1]/from/div/div/div[1]/div[1]/span/input')
                            skuInput.clear()
                            attempts = 10
                            while attempts > 0:
                                try:
                                    skuInput.send_keys(sn)
                                    self.obj_data_from_accel_file[current_sku].pop(0)
                                    break
                                except:
                                    time.sleep(0.5)
                                    attempts -= 1
                            else:
                                logger.error('sku input in smco cannot be interact from order: ',
                                             self.main_app.cus_order.get())
                                raise ValueError('sku input in smco cannot be interact')

                            print("fill sn complete")

                            skuInput.send_keys(Keys().ENTER)
                            print("pressed Enter at SKU-Input")
                            print(f'to_sent_dict = sku: {current_sku}, sn: {sn} ')
                            to_sent_dict = {'sku': current_sku, 'sn': sn}
                            self.used_serials.append(to_sent_dict)
                            print("current self.used_serials = ", self.used_serials)
                            time.sleep(2)

                        else:
                            print("ไม่มี SN, there are no functions available at this moment")
                            pass
                else:
                    print("ไม่ได้เลือก sn:", current_sku in self.obj_data_from_accel_file)
                    print("ไม่ได้เลือก sn:", current_sku, self.obj_data_from_accel_file)
                    print("ไม่ได้เลือก sn:", type(current_sku), type(self.obj_data_from_accel_file))
        else:
            print("No items, return!!")
            return
