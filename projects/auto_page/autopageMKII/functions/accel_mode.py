import os
import re
import time
from tkinter import filedialog

import pandas as pd
from loguru import logger
from openpyxl import load_workbook
from pypdf import PdfReader
from selenium.webdriver.common.by import By


class AccelMode:
    def __init__(self, main_app):
        self.main_app = main_app
        self.accel_file_dir = ""
        self.accel_df_state = {}
        self.accel_file_columns = []
        self.obj_data_from_accel_file = {}
        self.accel_orders_list = []
        self.CP_list = []
        self.accel_orders_count = 0
        self.excel_save_failed = False

    def select_accel_file(self):
        self.accel_file_dir = filedialog.askopenfilename(
            title="Select Accel File")
        if self.accel_file_dir:
            self.main_app.accl_dir_namedisplay_on_btn.configure(
                text=f"{os.path.basename(self.accel_file_dir)}")
        else:
            self.main_app.accl_dir_namedisplay_on_btn.configure(
                text=f"ยังไม่เลือก Accel File")

        self._read_accel_file_to_state(self.accel_file_dir)

    def _normalize_orders_col(self, df):
        """ตัดช่องว่างรอบเลข order โดยคงค่า missing ไว้เป็น pd.NA
        (ห้ามใช้ .astype(str).str.strip() ตรงๆ เพราะ NaN จะกลายเป็น string 'nan' แล้วถูกเขียนลง excel เป็นคำว่า nan)
        """
        if 'orders' not in df.columns:
            return
        orders = df['orders']
        missing = orders.isna() | orders.astype(str).str.strip(
        ).str.lower().isin(['nan', 'none', '<na>', 'null'])
        df.loc[:, 'orders'] = orders.astype(str).str.strip()
        df.loc[missing, 'orders'] = pd.NA

    def _read_accel_file_to_state(self, accel_file_dir):
        self.accel_file_dir = accel_file_dir
        self.accel_df_state = pd.read_excel(self.accel_file_dir, dtype=str)
        self.accel_df_state.columns = self.accel_df_state.columns.astype(str).str.strip()
        if 'orders' in self.accel_df_state.columns:
            self._normalize_orders_col(self.accel_df_state)
        print("before self.accel_df_state: ", self.accel_df_state)
        self.accel_df_state.loc[self.accel_df_state.duplicated(
            subset=['orders']), 'orders'] = pd.NA
        print("after self.accel_df_state: ", self.accel_df_state)

        self.accel_file_columns = self.accel_df_state.columns.dropna().tolist()
        self.obj_data_from_accel_file = {
            col: [str(x).strip() for x in self.accel_df_state[col].dropna(
            ).tolist() if str(x).strip() != 'nan']
            for col in self.accel_file_columns}

        self.accel_orders_list = self.accel_df_state['orders'].dropna(
        ).tolist()
        self.CP_list = self.accel_df_state['cp'].dropna().tolist()
        print(self.accel_orders_list)
        print('self.obj_data_from_accel_file: ', self.obj_data_from_accel_file)
        print(self.CP_list)

    def deduct_accel_file_data(self, order, sku_serials=[], remove_order=True, update_memory=True):
        if hasattr(order, 'get'):
            order = order.get()
        order = str(order).strip()
        df = self.accel_df_state
        print("deduct_accel_file_data df มีมาก่อนเหรอ: ", df)
        print("deduct_accel_file_data order: ", order)

        if remove_order and 'orders' in df.columns:
            mask_order = df['orders'].astype(str).str.strip() == order
            has_order = df.loc[mask_order, 'orders']
            if not has_order.empty:
                print(f'remove {order} from state df')
                df.loc[mask_order, 'orders'] = pd.NA

        print("deduct_accel_file_data sku_serials: ", sku_serials)
        if sku_serials:
            for sn in sku_serials:
                target_sku = str(sn.get('sku', '')).strip()
                target_sn = str(sn.get('sn', '')).strip()
                if not target_sku or not target_sn:
                    continue

                # ค้นหา column ที่ตรงกับ target_sku
                matched_col = None
                for col in df.columns:
                    col_str = str(col).strip()
                    if col_str.lower() == target_sku.lower() or target_sku.lower() in col_str.lower():
                        matched_col = col
                        break

                if matched_col is not None:
                    print(f"ตัด SN: {target_sn} ออกจากคอลัมน์ {matched_col}")
                    mask_sn = df[matched_col].astype(str).str.strip() == target_sn
                    df.loc[mask_sn, matched_col] = pd.NA
                else:
                    print(f"Warning: ไม่พบ คอลัมน์ SKU {target_sku} ใน DataFrame ของ Excel")

        print("form state df to new excel")
        print(
            f"Check if accel file is accesible {os.access(self.accel_file_dir, os.W_OK)}")

        try:
            self._save_df_to_excel(df, 'Sheet1')
            print(f"Successfully updated {self.accel_file_dir}")
            self.excel_save_failed = False

            if update_memory:
                # อ่าน dataframe ใหม่หลังจากอัปเดต Excel file
                self.accel_df_state = pd.read_excel(
                    self.accel_file_dir, dtype=str)
                self.accel_df_state.columns = self.accel_df_state.columns.astype(str).str.strip()
                if 'orders' in self.accel_df_state.columns:
                    self._normalize_orders_col(self.accel_df_state)
                self.obj_data_from_accel_file = {
                    col: [str(x).strip() for x in self.accel_df_state[col].dropna().tolist()
                          if str(x).strip() != 'nan'] for col in self.accel_file_columns}
        except PermissionError as e:
            print(f"Permission denied: {e}")
            logger.warning(
                f"ไฟล์ Excel ถูกเปิดอยู่ในโปรแกรมอื่น บันทึกไม่สำเร็จ จะใช้ข้อมูลในหน่วยความจำแทน: {e}")
            self.excel_save_failed = True
            if update_memory:
                # ใช้ข้อมูลในหน่วยความจำแทน ไม่อัปเดตไฟล์
                self.accel_df_state = df
                self.obj_data_from_accel_file = {
                    col: [str(x).strip() for x in self.accel_df_state[col].dropna().tolist()
                          if str(x).strip() != 'nan'] for col in self.accel_file_columns}
        except Exception as e:
            print(f"Error updating Excel file: {e}")
            self.excel_save_failed = True
            if update_memory:
                # ใช้ข้อมูลในหน่วยความจำแทน
                self.accel_df_state = df
                self.obj_data_from_accel_file = {
                    col: [str(x).strip() for x in self.accel_df_state[col].dropna().tolist()
                          if str(x).strip() != 'nan'] for col in self.accel_file_columns}

    def extract_sn_btn(self, accel_file_dir):
        if not accel_file_dir:
            print("select accel file first!!")
            return

        target_dirs = filedialog.askopenfilenames(title="Select SN PDF files")
        if len(target_dirs) != 0:
            for target_dir in target_dirs:
                self.sn_extractor(accel_file_dir, target_dir)
        else:
            print("You have not selected any transfer file, Extraction ends!!")
        self.accel_df_state = pd.read_excel(self.accel_file_dir, dtype=str)
        self.accel_df_state.columns = self.accel_df_state.columns.astype(str).str.strip()
        if 'orders' in self.accel_df_state.columns:
            self._normalize_orders_col(self.accel_df_state)

        self._read_accel_file_to_state(self.accel_file_dir)

    def sn_extractor(self, output_excel, target_dir):
        extracted_txt = self._extract_text_from_pdf(target_dir)
        extracted_txt = self._clean_text(extracted_txt)

        product_codes = self._extract_skus(extracted_txt)
        serial_numbers = self._extract_serial_numbers(extracted_txt)

        cleaned_serial_numbers = self._clean_serial_numbers(serial_numbers)
        serial_numbers_grouped = self._group_serial_numbers(
            cleaned_serial_numbers)

        self._print_debug_info(
            product_codes, cleaned_serial_numbers, serial_numbers_grouped)
        self._write_to_excel(output_excel, product_codes,
                             serial_numbers_grouped)

    def _extract_text_from_pdf(self, target_dir):
        reader = PdfReader(target_dir)
        extracted_txt = ""
        for page in reader.pages:
            extracted_txt += page.extract_text()
        return extracted_txt

    def _clean_text(self, text):
        pattern = r'^.*?(?=No\. Product Code Barcode Product Name Transfer No\. Order Ship Status)'
        text = re.sub(pattern, '', text, flags=re.DOTALL).lstrip()

        pattern2 = r'ผู้ส่งสินค้า.*?(?:No\. Product Code Barcode Product Name Transfer No\. Order Ship Status|วันที่ _ _ _ / _ _ _ / _ _ _)'
        text = re.sub(pattern2, '', text, flags=re.DOTALL)

        text = re.sub(r'Serial\s:', '', text, flags=re.DOTALL)
        text = re.sub(
            r'\d+\s{0,}(?=([A-Z0-9]{3}-[0-9]{6}))', '', text, flags=re.DOTALL)

        return text

    def _extract_skus(self, text):
        sku_pattern = r'([A-Z0-9]{3}-[0-9]{6})'
        return re.findall(sku_pattern, text)

    def _extract_serial_numbers(self, text):
        serial_pattern = r'(?:Shipped|Confirm)\s*([\w, \n, \/]+)(?=(?:[A-Z0-9]{3}-[0-9]{6}|\nผู้ส่งสินค้า|$))'
        return re.findall(serial_pattern, text, re.DOTALL)

    def _clean_serial_numbers(self, serial_numbers):
        cleaned = []
        for serial in serial_numbers:
            serial = re.sub(r'\n', '', serial).strip()
            cleaned.append(serial)
        return cleaned

    def _group_serial_numbers(self, cleaned_serial_numbers):
        return [serial.replace(" ", "").split(",") for serial in cleaned_serial_numbers]

    def _print_debug_info(self, product_codes, cleaned_serial_numbers, serial_numbers_grouped):
        print("Product Codes:")
        for i, code in enumerate(product_codes, 1):
            print(i, code)

        print("\nSerial Numbers:")
        for i, serial in enumerate(cleaned_serial_numbers, 1):
            serial_list = serial.replace(" ", "").split(",")
            print(f"{i} {len(serial_list)} [{serial}]")

        print("SKU Matches:", len(product_codes), product_codes)
        print("Serial Numbers Grouped:", len(
            serial_numbers_grouped), serial_numbers_grouped)

    def _write_to_excel(self, output_excel, product_codes, serial_numbers_grouped):
        try:
            book = load_workbook(output_excel)

            # Target 'Sheet1' (case-insensitive) instead of active sheet, fallback to first sheet if not found
            sheet_names_lower = [s.lower() for s in book.sheetnames]
            if 'sheet1' in sheet_names_lower:
                idx = sheet_names_lower.index('sheet1')
                sheet = book[book.sheetnames[idx]]
            else:
                sheet = book.worksheets[0]

            # * Map existing SKUs to their columns to avoid duplicates
            existing_skus = {}
            for col in range(1, sheet.max_column + 1):
                sku = sheet.cell(row=1, column=col).value
                if sku:
                    existing_skus[sku] = col

            # * Add new SKUs and their serial numbers, avoiding duplicates
            # *the incoming new data from PDF
            for sku, serials in zip(product_codes, serial_numbers_grouped):
                if sku in existing_skus:
                    col = existing_skus[sku]

                    existing_serials = []
                    for row in range(2, sheet.max_row + 1):
                        val = sheet.cell(row=row, column=col).value
                        if val:
                            existing_serials.append(val)

                    merged_serials = existing_serials.copy()
                    for s in serials:
                        if s not in merged_serials:
                            existing_serials.append(s)

                else:
                    # * ถ้าเป็น SKU ใหม่ ให้เพิ่มคอลัมน์ใหม่ แล้วใส่ serials ลงไปเรื่อยๆจนหมด
                    col = sheet.max_column + 1
                    sheet.cell(row=1, column=col, value=sku)
                    for row, serial in enumerate(serials, start=2):
                        sheet.cell(row=row, column=col, value=serial)
                    existing_skus[sku] = col

            book.save(output_excel)
            print(f"ข้อมูลถูกเพิ่ม/อัปเดตลงใน {output_excel} เรียบร้อยแล้ว")

        except PermissionError as e:
            print(f"Permission denied: {e}")
            print("ไฟล์ Excel อาจถูกเปิดอยู่ในโปรแกรมอื่น กรุณาปิดไฟล์แล้วลองใหม่")
        except Exception as e:
            import traceback
            print(f"เกิดข้อผิดพลาด: {e}")
            traceback.print_exc()

    # * start searching orders from accel file to check if order needed to be processed
    def accel_search(self):
        self.main_app.is_accel_mode_activated.set(True)

        # Reload Excel file state to get the latest status, orders, and SNs
        if self.accel_file_dir:
            try:
                self._read_accel_file_to_state(self.accel_file_dir)
            except Exception as e:
                logger.error(
                    f"เกิดข้อผิดพลาดในการโหลดไฟล์ Excel ที่เริ่มต้นค้นหา: {e}")

        self.accel_orders_count = len(self.accel_orders_list)
        if self.accel_orders_count == 0:
            logger.warning("ไม่มีออเดอร์ในไฟล์ Excel ให้ดำเนินการ")
            self.main_app.is_accel_mode_activated.set(False)
            return

        def start_next_cycle(count):
            # ดึงข้อมูลจาก Excel ใหม่ทุกรอบเพื่อให้ได้ SN บนสุดที่ยังเหลืออยู่ (เหมือน reload magazine)
            # แต่ถ้าเซฟครั้งก่อนไม่สำเร็จ (PermissionError) ให้ใช้ข้อมูลในหน่วยความจำล่าสุดแทนการไปดึงจากไฟล์เดิมบนดิสก์
            if getattr(self, 'excel_save_failed', False):
                logger.warning(
                    "ตรวจพบการบันทึก Excel ล้มเหลวก่อนหน้า จะใช้ข้อมูลในหน่วยความจำล่าสุดแทนการโหลดใหม่จากไฟล์ดิสก์")
            else:
                try:
                    self.accel_df_state = pd.read_excel(
                        self.accel_file_dir, dtype=str)
                    self.accel_df_state.columns = self.accel_df_state.columns.astype(str).str.strip()
                    if 'orders' in self.accel_df_state.columns:
                        self._normalize_orders_col(self.accel_df_state)
                    self.obj_data_from_accel_file = {
                        col: [str(x).strip() for x in self.accel_df_state[col].dropna(
                        ).tolist() if str(x).strip() != 'nan']
                        for col in self.accel_file_columns}
                except Exception as e:
                    logger.error(f"เกิดข้อผิดพลาดในการโหลดไฟล์ Excel: {e}")
            if count < self.accel_orders_count:
                if self.main_app.is_accel_mode_activated.get():
                    self.main_app.search_order(
                        self.accel_orders_list[count], lambda: start_next_cycle(count+1))
                else:
                    logger.info("Accel mode has been stopped by user.")
            else:
                pass

        self.main_app.search_order(
            self.accel_orders_list[0], lambda: start_next_cycle(1))

    # * ดึงรายการ SN ที่พร้อมใช้งานในสต็อกของ SMCO ทั้งหมดสำหรับ SKU นี้
    def get_available_sns_from_smco(self, driver, sku):
        """
        ดึงรายการ SN ที่พร้อมใช้งานในสต็อกของ SMCO ทั้งหมดสำหรับ SKU นี้
        """
        try:
            import re

            from loguru import logger

            # 1. เตรียม Cookies และ Origin จาก Browser
            cookies = self.main_app.bot.get_cookies_from_driver()
            current_url = driver.current_url
            matched_str = re.search(r'\/[A-z].*', current_url).group()
            origin = current_url.replace(matched_str, '')

            # 2. ค้นหา Product ID จาก SKU
            resp_prod = self.main_app.smco_api.get_product_info(
                origin, sku, cookies)
            product_data = resp_prod.json()

            if not product_data or len(product_data) == 0:
                logger.warning(f"ไม่พบข้อมูลสินค้าสำหรับ SKU: {sku}")
                return []

            product_id = product_data[0].get(
                'productId') or product_data[0].get('id')
            master_id = 180
            parent_id = 441

            if not product_id:
                return []

            # 3. ค้นหา Serial List จาก Product ID
            resp_sn = self.main_app.smco_api.get_serial_list(
                origin, product_id, master_id, parent_id, cookies
            )
            sn_list_data = resp_sn.json().get('data', [])

            # ดึงเฉพาะ serialNo ของตัวที่พร้อมใช้งาน
            found_sns = [str(item.get('serialNo')).strip()
                         for item in sn_list_data if item.get('serialNo')]
            logger.debug(
                f"SMCO API ส่งกลับมารวม {len(found_sns)} รายการสำหรับ product {product_id}")
            return found_sns

        except Exception as e:
            from loguru import logger
            logger.error(f"Error while fetching available SNs via API: {e}")
            return "API_ERROR"

    # * ตรวจสอบว่า SN นั้นมีอยู่ในระบบ SMCO จริงหรือไม่ (รักษา signature เดิมไว้เผื่อใช้ร่วมกับส่วนอื่น)
    def is_sn_in_smco(self, driver, sku, sn_to_check):
        """
        ตรวจสอบ SN ผ่าน API ว่ามีในสต็อกของ SMCO จริงไหม
        """
        res = self.get_available_sns_from_smco(driver, sku)
        if res == "API_ERROR":
            return "API_ERROR"
        return sn_to_check in res

    def _filter_invalid_sns(self, driver, current_sku):
        logger.info(
            f"จำนวนครั้งที่ใช้งานไม่ได้ของ SKU {current_sku} เกิน 2 ครั้ง (> 2) จะดึงข้อมูลสต็อกของ SKU นี้จาก SMCO...")
        available_smco_sns = self.get_available_sns_from_smco(driver, current_sku)
        if available_smco_sns != "API_ERROR":
            logger.info(f"ดึงข้อมูลสำเร็จ: พบ SN ที่ใช้งานได้ใน SMCO ทั้งหมด {len(available_smco_sns)} รายการ")
            current_candidate_sns = self.obj_data_from_accel_file.get(current_sku, [])
            sns_to_remove = [sn for sn in current_candidate_sns if sn not in available_smco_sns]

            if sns_to_remove:
                logger.info(
                    f"พบ SN ใน Excel ที่ไม่มีในสต็อกของ SMCO {len(sns_to_remove)} ตัว: {sns_to_remove} จะทำการตัดออก...")
                self.deduct_accel_file_data(
                    self.main_app.cus_order,
                    [{'sku': current_sku, 'sn': sn} for sn in sns_to_remove],
                    remove_order=False,
                    update_memory=True
                )
                logger.info(
                    f"อัปเดต Excel และสถานะความทรงจำเรียบร้อยแล้ว คงเหลือ SN ในระบบ: {self.obj_data_from_accel_file.get(current_sku, [])}")
            else:
                logger.info(
                    "SN ทั้งหมดใน Excel สอดคล้องกับสต็อกของ SMCO ไม่มีตัวต้องตัดออก")
        else:
            logger.warning(
                "ไม่สามารถดึงข้อมูลสต็อก SN จาก SMCO API ได้ (API_ERROR)")

    # * เอาไว้ใช้กับ smco โดยการเอา sn จาก accel file มาใส่ในช่อง sku input บนเว็บ smco และทำการ verify บนเว็บ
    def accel_fill_sku(self, driver, operation_thread):
        from loguru import logger
        from selenium.webdriver.common.keys import Keys
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.webdriver.support.ui import WebDriverWait

        def adjust_qty_down(sku, target_qty):
            print(
                f"กำลังตรวจสอบและปรับลดจำนวนสินค้า SKU: {sku} ให้เท่ากับ {target_qty}")
            for attempt in range(10):
                if operation_thread.is_set():
                    break

                try:
                    # ค้นหา index ของ SKU บนหน้าเว็บ
                    sku_elements = driver.find_elements(
                        By.XPATH,
                        "//span[(contains(@ng-click, 'productNameChangeChk(x)')) and not(contains(@class, 'ng-hide'))]//u")
                    target_idx = None
                    for idx, elem in enumerate(sku_elements):
                        if elem.text.strip() == sku.strip():
                            target_idx = idx
                            break

                    if target_idx is None:
                        print(
                            f"ไม่พบ SKU: {sku} บนหน้าเว็บ จึงไม่ต้องปรับลดจำนวน")
                        break

                    # ดึงจำนวนสินค้าปัจจุบันบนหน้าเว็บ
                    current_qty_elements = driver.find_elements(
                        By.XPATH, "//span[@class='col-sm-4 ng-binding' and not(contains(@class, 'ng-hide'))]"
                    )
                    if target_idx >= len(current_qty_elements):
                        print(
                            f"Warning: target_idx {target_idx} เกินจำนวนของ element แสดงจำนวนสินค้า")
                        break

                    current_qty = int(
                        current_qty_elements[target_idx].text.strip())
                    print(
                        f"จำนวนปัจจุบันบนหน้าเว็บ: {current_qty}, จำนวนที่ควรจะเป็น (target): {target_qty}")

                    if current_qty <= target_qty:
                        # จำนวนเหมาะสมแล้ว หรือน้อยกว่า/เท่ากับเป้าหมาย ไม่ต้องปรับลด
                        break

                    # ดึงปุ่มปรับลดจำนวน
                    decrease_buttons = driver.find_elements(
                        By.XPATH, "//button[@ng-click='incrementMainQty(false, x)' and not(contains(@class, 'ng-hide'))]"
                    )
                    if target_idx >= len(decrease_buttons):
                        print(
                            f"Warning: target_idx {target_idx} เกินจำนวนของปุ่มปรับลดจำนวน")
                        break

                    # กดปุ่มปรับลดจำนวน
                    print(
                        f"กดปุ่มปรับลดจำนวนของ SKU: {sku} จาก {current_qty} เหลือ {current_qty - 1}")
                    driver.execute_script(
                        "arguments[0].click();", decrease_buttons[target_idx])
                    time.sleep(0.5)  # รอให้หน้าจออัปเดต

                except Exception as e:
                    print(f"เกิดข้อผิดพลาดระหว่างปรับลดจำนวน: {e}")
                    time.sleep(0.5)

        accel_available_skus_list = list(self.obj_data_from_accel_file.keys())
        self.used_serials = []
        ordered_product_data_rows: list = self.main_app.items
        print('accel_fill_sku() ตรวจสอบ items = ', ordered_product_data_rows)

        sku_input_xpath = '/html/body/div[2]/div[3]/div[2]/div[2]/div[1]/div[1]/from/div/div/div[1]/div[1]/span/input'

        if len(ordered_product_data_rows) > 0:
            for i, ordered_item in enumerate(ordered_product_data_rows):
                print("item ordered by customer", ordered_item)
                current_ordered_sku = ordered_item['เลขอ้างอิง SKU (SKU Reference No.)']
                print("current_sku: ", current_ordered_sku)
                sku_qtys = int(ordered_item['จำนวน'])
                is_sku_ready_to_pick = [key for key in accel_available_skus_list
                                        if str(key) in str(current_ordered_sku)]

                if len(is_sku_ready_to_pick) > 0:
                    successful_count = 0
                    sku_fail_count = 0
                    while successful_count < sku_qtys and not operation_thread.is_set():
                        # ปรับลดจำนวนให้เท่ากับ successful_count ก่อนลอง SN ตัวใหม่
                        adjust_qty_down(current_ordered_sku, successful_count)

                        # ตรวจสอบว่ายังมี SN ในหน่วยความจำไหม
                        candidates = self.obj_data_from_accel_file.get(
                            current_ordered_sku, [])
                        if not candidates:
                            logger.warning(
                                f"ไม่มี SN เหลือใน Excel สำหรับ SKU: {current_ordered_sku}")
                            break

                        candidate_sn = candidates[0]
                        print(f"ลองใช้งาน SN จาก Excel: {candidate_sn}")

                        # รอช่อง Input SKU แสดงขึ้นมา
                        while not operation_thread.is_set():
                            try:
                                skuInput = driver.find_element(
                                    By.XPATH, sku_input_xpath)
                                if skuInput.is_displayed():
                                    break
                            except:
                                time.sleep(0.5)
                                continue

                        skuInput = driver.find_element(
                            By.XPATH, sku_input_xpath)
                        skuInput.clear()

                        attempts = 10
                        while attempts > 0:
                            try:
                                skuInput.send_keys(candidate_sn)
                                break
                            except:
                                time.sleep(0.5)
                                attempts -= 1
                        else:
                            logger.error(
                                f'sku input in smco cannot be interacted with from order: {self.main_app.cus_order.get()}')
                            raise ValueError(
                                'sku input in smco cannot be interacted with')

                        print(f"กรอก SN: {candidate_sn} สำเร็จ")
                        skuInput.send_keys(Keys.ENTER)
                        print("กด Enter ที่ช่อง input สำเร็จ")

                        # รอ SKU element และปุ่ม //i[@class='fa fa-check-square-o'] แสดงขึ้นมา
                        sku_elem_xpath = f"//span[@ng-click='productNameChangeChk(x)']/a/u[text()='{current_ordered_sku}']"
                        check_btn_xpath = "//i[@class='fa fa-check-square-o']"

                        wait_timeout = 44
                        start_wait = time.time()
                        check_btn = None
                        sku_elem = None

                        print(
                            "กำลังรอปุ่มยืนยัน SN (fa-check-square-o) และ SKU element...")
                        while (time.time() - start_wait) < wait_timeout and not operation_thread.is_set():
                            try:
                                sku_elem = driver.find_element(
                                    By.XPATH, sku_elem_xpath)
                                check_btn = driver.find_element(
                                    By.XPATH, check_btn_xpath)
                                if sku_elem.is_displayed() and check_btn.is_displayed():
                                    break
                            except:
                                pass
                            time.sleep(0.5)
                        else:
                            logger.error(
                                f"หมดเวลารอปุ่มยืนยัน SN หรือ SKU element สำหรับ {candidate_sn}")
                            # หากรอไม่เจอ ถือว่า SN นั้นมีปัญหา ให้เอาออกแล้วลองตัวถัดไป
                            self.deduct_accel_file_data(
                                self.main_app.cus_order, [
                                    {'sku': current_ordered_sku, 'sn': candidate_sn}],
                                remove_order=False, update_memory=False)
                            if candidate_sn in self.obj_data_from_accel_file[current_ordered_sku]:
                                self.obj_data_from_accel_file[current_ordered_sku].remove(
                                    candidate_sn)

                            sku_fail_count += 1
                            if sku_fail_count > 2:
                                self._filter_invalid_sns(driver, current_ordered_sku)
                            continue

                        # ดัก req response api
                        # เคลียร์ logs ก่อนกดเพื่อความถูกต้อง
                        try:
                            self.main_app.bot.network_capture.clear_logs()
                        except Exception as log_err:
                            logger.warning(
                                f"ไม่สามารถเคลียร์ log performance ได้: {log_err}")

                        # กดปุ่มยืนยัน
                        print(f"กำลังกดปุ่มยืนยัน SN...")
                        try:
                            driver.execute_script(
                                "arguments[0].click();", check_btn)
                        except Exception as click_err:
                            logger.warning(
                                f"JS click ล้มเหลว จะลองคลิกแบบปกติ: {click_err}")
                            check_btn.click()

                        # รอและดัก response
                        print("กำลังรอ response จาก /verifySerialFullBill.htm ...")
                        response = self.main_app.bot.network_capture.capture_response(
                            'verifySerialFullBill.htm', max_attempts=40, wait_interval=0.5
                        )

                        # ตรวจสอบความถูกต้อง
                        is_invalid = False
                        reasons = []

                        if response is not None:
                            if isinstance(response, list):
                                for item in response:
                                    if isinstance(item, dict):
                                        if 'reasonNameEn' in item or 'reasonNameTh' in item:
                                            is_invalid = True
                                            reasons.append(item.get('reasonNameEn') or item.get(
                                                'reasonNameTh') or 'Unknown Reason')
                            elif isinstance(response, dict):
                                if 'reasonNameEn' in response or 'reasonNameTh' in response:
                                    is_invalid = True
                                    reasons.append(response.get('reasonNameEn') or response.get(
                                        'reasonNameTh') or 'Unknown Reason')

                        # ตรวจสอบว่าปุ่ม //i[@class='fa fa-check-square-o'] หายไปหรือไม่
                        button_disappeared = False
                        for _ in range(10):
                            try:
                                btn = driver.find_element(
                                    By.XPATH, check_btn_xpath)
                                if not btn.is_displayed():
                                    button_disappeared = True
                                    break
                            except:
                                button_disappeared = True
                                break
                            time.sleep(0.5)

                        # ถ้ามี reasonNameEn/Th หรือปุ่มไม่ยอมหายไป แสดงว่าใช้งานไม่ได้
                        if is_invalid or not button_disappeared:
                            print(
                                f"SN {candidate_sn} ใช้งานไม่ได้! เหตุผล: {reasons if is_invalid else 'ปุ่มไม่หายไป'}")

                            # 1. ปิด Swal popup หรือ alert ที่เด้งขึ้นมา
                            try:
                                swal_ok = driver.find_element(
                                    By.XPATH,
                                    "//button[@class='swal2-confirm styled' and (text()='OK' or text()='ตกลง')]")
                                if swal_ok.is_displayed():
                                    swal_ok.click()
                                    print("ปิด Swal popup สำเร็จ")
                            except:
                                pass

                            try:
                                alert = driver.switch_to.alert
                                alert_text = alert.text
                                alert.accept()
                                print(f"ยอมรับ browser alert: {alert_text}")
                            except:
                                pass

                            # 2. ค้นหาลำดับของ SKU ที่มีปัญหาใน DOM แล้วกดปุ่มลบ (btn-danger) และเช็คว่าหายไปแล้วจริงๆ
                            try:
                                target_idx = None
                                sku_elements = driver.find_elements(
                                    By.XPATH,
                                    "//span[(contains(@ng-click, 'productNameChangeChk(x)')) and not(contains(@class, 'ng-hide'))]//u"
                                )
                                for idx, elem in enumerate(sku_elements):
                                    text_content = elem.text.strip()
                                    if (current_ordered_sku.strip().lower() in text_content.lower() or 
                                        text_content.lower() in current_ordered_sku.strip().lower()):
                                        target_idx = idx
                                        break

                                delete_buttons = driver.find_elements(
                                    By.XPATH,
                                    "//button[@class='btn btn-danger btn-sm ng-scope']"
                                )

                                deleted_button_clicked = False
                                if target_idx is not None and target_idx < len(delete_buttons):
                                    driver.execute_script("arguments[0].click();", delete_buttons[target_idx])
                                    print(f"กดปุ่มลบรายการลำดับที่ {target_idx} (SKU: {current_ordered_sku}) ที่ตรวจสอบไม่ผ่านสำเร็จ")
                                    deleted_button_clicked = True
                                else:
                                    # Fallback: ลองหาปุ่มลบผ่าน panel ancestor
                                    for elem in sku_elements:
                                        if (current_ordered_sku.strip().lower() in elem.text.strip().lower() or 
                                            elem.text.strip().lower() in current_ordered_sku.strip().lower()):
                                            try:
                                                panel = elem.find_element(By.XPATH, "./ancestor::div[contains(@class, 'panel')][1]")
                                                del_btns = panel.find_elements(By.XPATH, ".//button[contains(@class, 'btn-danger')]")
                                                if del_btns:
                                                    driver.execute_script("arguments[0].click();", del_btns[0])
                                                    print(f"กดปุ่มลบรายการ SKU: {current_ordered_sku} ผ่าน panel ancestor สำเร็จ")
                                                    deleted_button_clicked = True
                                                    break
                                            except:
                                                pass

                                if not deleted_button_clicked:
                                    print(f"ไม่สามารถหาปุ่มลบสำหรับ SKU {current_ordered_sku} ได้ ลองปรับลดจำนวนด้วย adjust_qty_down")
                                    adjust_qty_down(current_ordered_sku, successful_count)

                                # เช็คว่า SKU ของ SN นั้น หายไปจาก DOM แล้วจริงๆ ก่อนเติม SN ถัดไป
                                print(f"กำลังตรวจสอบว่า SKU {current_ordered_sku} หายไปจาก DOM หรือยัง...")
                                start_check = time.time()
                                while (time.time() - start_check) < 6:
                                    current_skus = [
                                        e.text.strip() for e in driver.find_elements(
                                            By.XPATH,
                                            "//span[(contains(@ng-click, 'productNameChangeChk(x)')) and not(contains(@class, 'ng-hide'))]//u"
                                        )
                                    ]
                                    is_still_in_dom = any(
                                        current_ordered_sku.strip().lower() in s.lower() or s.lower() in current_ordered_sku.strip().lower()
                                        for s in current_skus
                                    )
                                    if not is_still_in_dom:
                                        print(f"ยืนยันเรียบร้อย: SKU {current_ordered_sku} หายไปจาก DOM แล้วจริงๆ!")
                                        break
                                    time.sleep(0.5)

                            except Exception as del_err:
                                print(f"เกิดข้อผิดพลาดในการลบรายการ: {del_err}")
                                adjust_qty_down(current_ordered_sku, successful_count)

                            # 3. ลบ SN ตัวที่มีปัญหาออกจาก Excel และหน่วยความจำ
                            self.deduct_accel_file_data(
                                self.main_app.cus_order, [
                                    {'sku': current_ordered_sku, 'sn': candidate_sn}],
                                remove_order=False, update_memory=False
                            )
                            if candidate_sn in self.obj_data_from_accel_file[current_ordered_sku]:
                                self.obj_data_from_accel_file[current_ordered_sku].remove(
                                    candidate_sn)

                            sku_fail_count += 1
                            if sku_fail_count > 2:
                                self._filter_invalid_sns(driver, current_ordered_sku)
                            time.sleep(1)
                            # วนกลับไปรันใหม่โดยไม่เพิ่ม successful_count
                        else:
                            print(f"SN {candidate_sn} ใช้งานได้สำเร็จ!")
                            # แอดเข้า used serials
                            self.used_serials.append({'sku': current_ordered_sku, 'sn': candidate_sn})
                            # เอาออกจากหน่วยความจำ (เพราะใช้ได้แล้ว)
                            if candidate_sn in self.obj_data_from_accel_file[current_ordered_sku]:
                                self.obj_data_from_accel_file[current_ordered_sku].remove(
                                    candidate_sn)
                            successful_count += 1
                            time.sleep(1)

                    # เมื่อจบ loop ของ SKU นี้ ปรับลดจำนวนสินค้าให้ตรงตามจริงที่สำเร็จอีกครั้งเพื่อความถูกต้อง
                    adjust_qty_down(current_ordered_sku, successful_count)
                else:
                    logger.info(
                        f"มี current_sku ใน Accel_File หรือไม่?: {current_ordered_sku in self.obj_data_from_accel_file}")
                    print("มี current_sku ใน Accel_File หรือไม่?:",
                          current_ordered_sku in self.obj_data_from_accel_file)
        else:
            print("No items, return!!")
            return

    def _save_df_to_excel(self, target_df, sheet_name):
        if not os.path.exists(self.accel_file_dir):
            target_df.to_excel(self.accel_file_dir,
                               sheet_name=sheet_name, index=False)
            return

        try:
            with pd.ExcelWriter(self.accel_file_dir, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                target_df.to_excel(writer, sheet_name=sheet_name, index=False)
        except (TypeError, ValueError):
            try:
                book = load_workbook(self.accel_file_dir)
                if sheet_name in book.sheetnames:
                    del book[sheet_name]
                    book.save(self.accel_file_dir)
                book.close()
                with pd.ExcelWriter(self.accel_file_dir, engine='openpyxl', mode='a') as writer:
                    target_df.to_excel(
                        writer, sheet_name=sheet_name, index=False)
            except Exception as ex:
                print(f"Append failed, overwriting entire excel file: {ex}")
                target_df.to_excel(self.accel_file_dir,
                                   sheet_name=sheet_name, index=False)

    def record_failed_order(self, order, reason):
        """Record failed order

        Args:
            order (_type_): _order to record
            reason (_type_): 
        """
        if hasattr(order, 'get'):
            order_str = order.get()
        else:
            order_str = str(order)

        print(f"Recording failed order: {order_str} due to: {reason}")

        if not self.accel_file_dir:
            print("No accel file selected, cannot record failed order.")
            return

        try:
            if not os.path.exists(self.accel_file_dir):
                print(
                    f"Accel file {self.accel_file_dir} does not exist, cannot record failed order.")
                return

            failed_df = pd.DataFrame(
                columns=['orders', 'failed_reason', 'timestamp'])

            try:
                failed_df = pd.read_excel(
                    self.accel_file_dir, sheet_name='Failed_Orders', dtype=str)
            except Exception:
                print("Failed_Orders sheet does not exist yet. Creating a new one.")

            new_row = pd.DataFrame([{
                'orders': order_str,
                'failed_reason': str(reason),
                'timestamp': time.strftime("%Y-%m-%d %H:%M:%S")
            }])

            failed_df = failed_df[failed_df['orders'] != order_str]
            failed_df = pd.concat([failed_df, new_row], ignore_index=True)

            self._save_df_to_excel(failed_df, 'Failed_Orders')
            print(
                f"Successfully recorded failed order {order_str} to Failed_Orders sheet.")
        except PermissionError as e:
            print(f"Permission denied while recording failed order: {e}")
            logger.warning(
                f"ไฟล์ Excel ถูกเปิดอยู่ในโปรแกรมอื่น บันทึก Failed Order ไม่สำเร็จ: {e}")
        except Exception as e:
            print(f"Error recording failed order to Excel: {e}")
            logger.error(f"Error recording failed order to Excel: {e}")

    def record_completed_order(self, order, tracking="", bill_no="", status="Completed"):
        """Record completed order into Completed_Orders sheet in Accel Excel file

        Args:
            order: order object or string
            tracking: tracking number string
            bill_no: bill/receipt number string
            status: completion status string (e.g. Completed, TEST_SUCCESS)
        """
        if hasattr(order, 'get'):
            order_str = order.get()
        else:
            order_str = str(order)

        print(f"Recording completed order: {order_str} (Status: {status})")

        if not self.accel_file_dir:
            print("No accel file selected, cannot record completed order.")
            return

        try:
            if not os.path.exists(self.accel_file_dir):
                print(
                    f"Accel file {self.accel_file_dir} does not exist, cannot record completed order.")
                return

            completed_df = pd.DataFrame(
                columns=['tracking', 'orders', 'bill_no', 'timestamp', 'status'])

            try:
                completed_df = pd.read_excel(
                    self.accel_file_dir, sheet_name='Completed_Orders', dtype=str)
            except Exception:
                print("Completed_Orders sheet does not exist yet. Creating a new one.")

            # จัดลำดับ column ให้เป็นแบบใหม่เสมอ (ไฟล์เก่าที่เรียงแบบเดิมจะถูก reorder ด้วย)
            # align ด้วยชื่อ column ไม่ใช้ตำแหน่ง -> ข้อมูลไม่หลุดหาย
            _col_order = ['tracking', 'orders', 'bill_no', 'timestamp', 'status']
            for _col in _col_order:
                if _col not in completed_df.columns:
                    completed_df[_col] = ""
            completed_df = completed_df[_col_order]

            new_row = pd.DataFrame([{
                'tracking': str(tracking),
                'orders': order_str,
                'bill_no': str(bill_no),
                'timestamp': time.strftime("%Y-%m-%d %H:%M:%S"),
                'status': str(status)
            }])

            completed_df = completed_df[completed_df['orders'] != order_str]
            completed_df = pd.concat(
                [completed_df, new_row], ignore_index=True)
            completed_df = completed_df[_col_order]

            self._save_df_to_excel(completed_df, 'Completed_Orders')
            print(
                f"Successfully recorded completed order {order_str} to Completed_Orders sheet.")
        except PermissionError as e:
            print(f"Permission denied while recording completed order: {e}")
            logger.warning(
                f"ไฟล์ Excel ถูกเปิดอยู่ในโปรแกรมอื่น บันทึก Completed Order ไม่สำเร็จ: {e}")
        except Exception as e:
            print(f"Error recording completed order to Excel: {e}")
            logger.error(f"Error recording completed order to Excel: {e}")
